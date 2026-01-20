#!/usr/bin/env python3
"""
Create Excel template for Favela Brass staff sheets.
Upload to Google Drive and open with Google Sheets.
"""

import sqlite3
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter

DB_PATH = "/Users/tom/Documents/HQ/favelabrass/data/favelabrass.db"
OUTPUT_PATH = "/Users/tom/Documents/HQ/favelabrass/outputs/FavelaBrass_Database.xlsx"

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, num_cols):
    """Apply header styling to first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

def add_dropdown(ws, col_letter, options, start_row=2, end_row=500):
    """Add dropdown validation to a column."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True
    )
    dv.error = "Selecione uma opção da lista"
    dv.errorTitle = "Valor inválido"
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')


def add_next_id_helper(ws, prefix, col_letter='R'):
    """Add a 'next ID' helper cell to help staff generate new IDs.

    Places a formula in column R (or specified) row 1 that calculates the next available ID.
    """
    # Label in col R row 1
    ws.cell(row=1, column=ord(col_letter) - ord('A') + 1, value="Próximo ID →")
    ws.cell(row=1, column=ord(col_letter) - ord('A') + 1).font = Font(bold=True, color="FF0000")

    # Formula in col S row 1 - extracts number from ID like "AV-123", finds max, adds 1
    next_col = chr(ord(col_letter) + 1)
    formula = f'="{prefix}-"&TEXT(MAX(IFERROR(VALUE(SUBSTITUTE(A:A,"{prefix}-","")),0))+1,"000")'
    ws.cell(row=1, column=ord(next_col) - ord('A') + 1, value=formula)
    ws.cell(row=1, column=ord(next_col) - ord('A') + 1).font = Font(bold=True, color="FF0000")
    ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions[next_col].width = 10

def create_alunos_sheet(wb, conn):
    """Create Alunos (Students) sheet - matches Google Form + admin fields."""
    ws = wb.active
    ws.title = "Alunos"

    # Matches Google Form column order exactly, plus admin fields at start
    headers = [
        # Admin fields (not from form)
        ("id", 8),
        ("status", 12),
        ("program_type", 12),
        ("confirmado_2026", 13),
        ("primeira_matricula", 13),               # First enrollment date (from historical data)
        ("ultima_matricula", 13),                 # Last enrollment date (when last re-registered)
        # Form fields (exact order from Banco de Dados)
        ("serie_escolar", 12),                    # 4. Está em que série escolar?
        ("escola", 28),                           # 5. Selecione a escola
        ("nome", 35),                             # 6. Nome completo da criança
        ("data_nascimento", 13),                  # 7. Data de nascimento
        ("rg_crianca", 15),                       # 8. Documento de identidade da criança
        ("cpf_crianca", 15),                      # 9. CPF da criança
        ("responsavel1_nome", 35),                # 10. Nome completo (responsável 1)
        ("responsavel1_rg", 15),                  # 11. RG Responsável 1
        ("responsavel1_cpf", 15),                 # 12. CPF Responsável 1
        ("responsavel1_tel", 15),                 # 13. Celular de contato (responsável 1)
        ("responsavel2_nome", 30),                # 14. Nome completo (responsável 2)
        ("responsavel2_rg", 15),                  # 15. RG (Responsável 2)
        ("responsavel2_cpf", 15),                 # 16. CPF (Responsável 2)
        ("responsavel2_tel", 15),                 # 17. Celular de contato (responsável 2)
        ("endereco", 40),                         # 18. Endereço
        ("bairro", 18),                           # 19. Bairro de residência
        ("cep", 12),                              # 20. CEP
        ("mora_em_comunidade", 15),               # 21. Em comunidade/favela?
        ("comunidade", 20),                       # NEW: Which community/favela
        ("tamanho_uniforme", 13),                 # 22. Tamanho de camiseta
        ("unico_na_familia", 13),                 # 23. É o único da família no projeto?
        ("bolsa_familia", 12),                    # 24. Recebe Bolsa Família?
        ("tipo_imovel", 15),                      # 25. Tipo de imóvel
        ("pessoas_na_casa", 12),                  # 26. Quantas pessoas moram na casa?
        ("contribuintes_renda", 12),              # 27. Quantas contribuem com renda?
        ("renda_familiar", 18),                   # 28. Renda familiar total
        ("condicao_medica", 35),                  # 29. Condição médica ou alergia
        ("usa_medicamento", 12),                  # 30. Faz uso de medicamento?
        ("qual_medicamento", 20),                 # 31. Qual medicamento?
        ("deficiencia", 20),                      # 32. Apresenta deficiência?
    ]

    # Write headers
    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    # Add dropdowns (E=primeira_matricula, F=ultima_matricula, then form fields)
    add_dropdown(ws, 'B', ['Ativo', 'Evadido', 'Ex-Aluno'])  # status
    add_dropdown(ws, 'C', ['avancado', 'escola'])  # program_type
    add_dropdown(ws, 'D', ['Sim', 'Não', 'Evadido'])  # confirmado_2026
    add_dropdown(ws, 'X', ['Sim', 'Não'])  # mora_em_comunidade (shifted +1)
    communities = ['Pereira da Silva', 'Fallet', 'Complexo do Turano', 'Guararapes', 'Morro dos Cabritos',
                   'Chapéu Mangueira', 'Babilônia', 'Santa Marta', 'Cantagalo', 'Pavão-Pavãozinho', 'Outra']
    add_dropdown(ws, 'Y', communities)  # comunidade (shifted +1)
    add_dropdown(ws, 'Z', ['PP', 'P', 'M', 'G', 'GG', 'XGG'])  # tamanho_uniforme (shifted +1)
    add_dropdown(ws, 'AA', ['Sim', 'Não'])  # unico_na_familia (shifted +1)
    add_dropdown(ws, 'AB', ['Sim', 'Não'])  # bolsa_familia (shifted +1)
    add_dropdown(ws, 'AG', ['Sim', 'Não'])  # usa_medicamento (shifted +1)

    # Load existing data
    c = conn.cursor()
    c.execute("""
        SELECT
            id, status, program_type, confirmed_2026, first_enrollment_date, last_enrollment_date,
            school_grade, school, name, birth_date,
            child_rg, child_cpf,
            guardian1_name, guardian1_rg, guardian1_cpf, guardian1_phone,
            guardian2_name, guardian2_rg, guardian2_cpf, guardian2_phone,
            address, neighborhood, cep, lives_in_community, community,
            uniform_size, only_child_in_project, bolsa_familia, housing_type,
            household_size, income_contributors, family_income,
            medical_condition, medications IS NOT NULL, medications, has_disability
        FROM students
        ORDER BY name
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if value == 1:
                cell.value = "Sim"
            elif value == 0:
                cell.value = "Não"
            else:
                cell.value = value

def create_instrumentos_sheet(wb, conn):
    """Create Instrumentos (Instruments) sheet."""
    ws = wb.create_sheet("Instrumentos")

    headers = [
        ("id", 8),
        ("tipo", 15),
        ("marca_modelo", 25),
        ("numero_serie", 20),
        ("qualidade", 12),
        ("tem_case", 10),
        ("baixado", 10),
        ("motivo_baixa", 25),
        ("anotacoes", 40),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    # Dropdowns
    instrument_types = ['Trompete', 'Trombone', 'Sax Alto', 'Sax Tenor', 'Sax Barítono',
                        'Percussão', 'Escaleta', 'Trombonito', 'Cornet', 'Tuba', 'Clarinete']
    add_dropdown(ws, 'B', instrument_types)  # tipo
    add_dropdown(ws, 'E', ['Bom', 'Regular', 'Ruim'])  # qualidade
    add_dropdown(ws, 'F', ['Sim', 'Não'])  # tem_case
    add_dropdown(ws, 'G', ['Sim', 'Não'])  # baixado

    # Load data
    c = conn.cursor()
    c.execute("""
        SELECT id, type, brand_model, serial_number, quality, has_case,
               is_written_off, writeoff_reason, notes
        FROM instruments
        ORDER BY id
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if value == 1 and col_num in (6, 7):
                cell.value = "Sim"
            elif value == 0 and col_num in (6, 7):
                cell.value = "Não"
            else:
                cell.value = value

def create_emprestimos_sheet(wb, conn):
    """Create Emprestimos (Loans) sheet with proper ID references."""
    ws = wb.create_sheet("Emprestimos")

    headers = [
        ("id", 10),
        ("instrumento_id", 15),
        ("categoria", 12),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP for students
        ("pessoa_nome", 35),  # For ex-students/teachers (manual entry)
        ("data_emprestimo", 15),
        ("data_devolucao", 15),
        ("status", 12),
        ("obs", 40),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    add_dropdown(ws, 'C', ['Aluno', 'Ex-aluno', 'Professor'])  # categoria
    add_dropdown(ws, 'I', ['Ativo', 'Devolvido'])  # status

    # Load data
    c = conn.cursor()
    c.execute("""
        SELECT id, instrument_id, category, student_id, person_name,
               loan_date, return_date, status, notes
        FROM instrument_loans
        ORDER BY loan_date DESC
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        loan_id, instr_id, category, student_id, person_name, loan_date, return_date, status, notes = row
        ws.cell(row=row_num, column=1, value=loan_id)
        ws.cell(row=row_num, column=2, value=instr_id)
        ws.cell(row=row_num, column=3, value=category)
        ws.cell(row=row_num, column=4, value=student_id)
        # VLOOKUP for student name (only works if aluno_id is filled)
        ws.cell(row=row_num, column=5, value=f'=IFERROR(VLOOKUP(D{row_num},Alunos!A:B,2,FALSE),"")')
        # pessoa_nome for ex-students and teachers
        ws.cell(row=row_num, column=6, value=person_name if category != 'Aluno' else '')
        ws.cell(row=row_num, column=7, value=loan_date)
        ws.cell(row=row_num, column=8, value=return_date)
        ws.cell(row=row_num, column=9, value=status)
        ws.cell(row=row_num, column=10, value=notes)

def create_avaliacoes_pratica_sheet(wb, conn):
    """Create Avaliacoes_Pratica (Practical Assessments) sheet with ID references.

    Only includes internal exams - MTB external exams are managed separately via /mtb command.
    """
    ws = wb.create_sheet("Avaliacoes_Pratica")

    headers = [
        ("id", 10),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP
        ("data", 12),
        ("nivel", 8),
        ("instrumento", 15),
        ("peca_1", 8),
        ("peca_2", 8),
        ("peca_3", 8),
        ("escalas", 8),
        ("leitura", 8),
        ("tecnica", 8),
        ("total", 8),
        ("resultado", 12),
        ("avaliador", 20),
        ("obs", 30),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    instrument_types = ['Trompete', 'Trombone', 'Sax Alto', 'Sax Tenor', 'Sax Barítono', 'Percussão']
    add_dropdown(ws, 'E', ['1', '2', '3', '4'])  # nivel (now col E)
    add_dropdown(ws, 'F', instrument_types)  # instrumento (now col F)
    add_dropdown(ws, 'N', ['Distinção', 'Mérito', 'Aprovado', 'Reprovado'])  # resultado (now col N)

    # Load data - internal exams only (MTB external exams handled via /mtb command)
    c = conn.cursor()
    c.execute("""
        SELECT id, student_id, assessment_date, level_tested, instrument,
               score_piece_1, score_piece_2, score_piece_3, score_scales,
               score_sight_reading, score_technical, total_score, result, examiner, notes
        FROM assessments_practical
        WHERE exam_type = 'internal' OR exam_type IS NULL
        ORDER BY assessment_date DESC
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        (assess_id, student_id, date, level, instrument, p1, p2, p3,
         scales, reading, tech, total, result, examiner, notes) = row
        ws.cell(row=row_num, column=1, value=assess_id)
        ws.cell(row=row_num, column=2, value=student_id)
        ws.cell(row=row_num, column=3, value=f'=IFERROR(VLOOKUP(B{row_num},Alunos!A:B,2,FALSE),"")')
        ws.cell(row=row_num, column=4, value=date)
        ws.cell(row=row_num, column=5, value=level)
        ws.cell(row=row_num, column=6, value=instrument)
        ws.cell(row=row_num, column=7, value=p1)
        ws.cell(row=row_num, column=8, value=p2)
        ws.cell(row=row_num, column=9, value=p3)
        ws.cell(row=row_num, column=10, value=scales)
        ws.cell(row=row_num, column=11, value=reading)
        ws.cell(row=row_num, column=12, value=tech)
        ws.cell(row=row_num, column=13, value=total)
        ws.cell(row=row_num, column=14, value=result)
        ws.cell(row=row_num, column=15, value=examiner)
        ws.cell(row=row_num, column=16, value=notes)

    # Add next ID helper for Wesley
    add_next_id_helper(ws, 'AV', col_letter='R')


def create_avaliacoes_teoria_sheet(wb, conn):
    """Create Avaliacoes_Teoria (Theory Assessments) sheet with ID references."""
    ws = wb.create_sheet("Avaliacoes_Teoria")

    headers = [
        ("id", 10),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP
        ("data", 12),
        ("nivel", 8),
        ("pontuacao", 10),
        ("resultado", 12),
        ("avaliador", 20),
        ("obs", 30),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    add_dropdown(ws, 'E', ['1', '2', '3', '4'])  # nivel (now col E)
    add_dropdown(ws, 'G', ['Aprovado', 'Reprovado'])  # resultado (now col G)

    # Load data
    c = conn.cursor()
    c.execute("""
        SELECT id, student_id, assessment_date, level_tested, score, result, examiner, notes
        FROM assessments_theory
        ORDER BY assessment_date DESC
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        assess_id, student_id, date, level, score, result, examiner, notes = row
        ws.cell(row=row_num, column=1, value=assess_id)
        ws.cell(row=row_num, column=2, value=student_id)
        ws.cell(row=row_num, column=3, value=f'=IFERROR(VLOOKUP(B{row_num},Alunos!A:B,2,FALSE),"")')
        ws.cell(row=row_num, column=4, value=date)
        ws.cell(row=row_num, column=5, value=level)
        ws.cell(row=row_num, column=6, value=score)
        ws.cell(row=row_num, column=7, value=result)
        ws.cell(row=row_num, column=8, value=examiner)
        ws.cell(row=row_num, column=9, value=notes)

    # Add next ID helper
    add_next_id_helper(ws, 'AT', col_letter='K')


def create_grupos_sheet(wb, conn):
    """Create Grupos sheet (reference for bands/groups)."""
    ws = wb.create_sheet("Grupos")

    headers = [
        ("id", 12),
        ("nome", 20),
        ("regente", 25),
        ("ativa", 8),
        ("tamanho_previsto", 15),
        ("descricao", 40),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    add_dropdown(ws, 'D', ['Sim', 'Não'])  # ativa

    c = conn.cursor()
    c.execute("SELECT id, name, conductor, is_active, target_size, description FROM groups ORDER BY name")

    for row_num, row in enumerate(c.fetchall(), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if value == 1 and col_num == 4:
                cell.value = "Sim"
            elif value == 0 and col_num == 4:
                cell.value = "Não"
            else:
                cell.value = value

def create_atribuicao_grupos_sheet(wb, conn):
    """Create Atribuicao_Grupos sheet with proper ID references."""
    ws = wb.create_sheet("Atribuicao_Grupos")

    headers = [
        ("id", 12),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP formula
        ("grupo", 20),
        ("grupo_anterior", 20),  # Previous band (2025)
        ("instrumento", 15),
        ("turma_teoria", 15),
        ("ano_formatura", 12),
        ("obs", 25),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    groups = ['Banda Preta', 'Banda Roxa', 'Banda Verde', 'Banda Amarela',
              'Banda Branca', 'Grupo Semente', 'Iniciação Musical']
    instruments = ['Trompete', 'Trombone', 'Sax Alto', 'Sax Tenor', 'Sax Barítono',
                   'Percussão', 'Caixa', 'Surdo', 'Flauta Doce']
    theory_classes = ['Teoria Nível 1', 'Teoria Nível 2', 'Teoria Nível 3', 'Teoria Nível 4']

    add_dropdown(ws, 'D', groups)  # grupo
    add_dropdown(ws, 'E', groups)  # grupo_anterior
    add_dropdown(ws, 'F', instruments)  # instrumento
    add_dropdown(ws, 'G', theory_classes)  # turma_teoria

    c = conn.cursor()
    c.execute("""
        SELECT id, student_id, group_id, previous_band, current_instrument, theory_class,
               graduation_year, notes
        FROM group_assignments
        ORDER BY group_id, student_id
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        assign_id, student_id, group, prev_band, instrument, theory, grad_year, notes = row
        ws.cell(row=row_num, column=1, value=assign_id)
        ws.cell(row=row_num, column=2, value=student_id)
        ws.cell(row=row_num, column=3, value=f'=IFERROR(VLOOKUP(B{row_num},Alunos!A:I,9,FALSE),"")')
        ws.cell(row=row_num, column=4, value=group)
        ws.cell(row=row_num, column=5, value=prev_band)
        ws.cell(row=row_num, column=6, value=instrument)
        ws.cell(row=row_num, column=7, value=theory)
        ws.cell(row=row_num, column=8, value=grad_year)
        ws.cell(row=row_num, column=9, value=notes)

def create_professores_sheet(wb, conn):
    """Create Professores sheet."""
    ws = wb.create_sheet("Professores")

    headers = [
        ("nome", 30),
        ("funcao", 25),
        ("situacao", 15),
        ("valor_hora", 12),
        ("instrumentos", 40),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    add_dropdown(ws, 'C', ['Ativo', 'Inativo', 'Em Experiência', 'Afastado'])  # situacao

    c = conn.cursor()
    c.execute("SELECT name, role, status, hourly_rate, instruments_taught FROM teachers ORDER BY name")

    for row_num, row in enumerate(c.fetchall(), 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)

def create_atividades_sheet(wb, conn):
    """Create Atividades (Schedule) sheet - granular time slots."""
    ws = wb.create_sheet("Atividades")

    headers = [
        ("id", 10),
        ("dia_semana", 15),
        ("nome", 40),
        ("tipo", 18),
        ("horario_inicio", 14),
        ("horario_fim", 14),
        ("duracao_horas", 12),
        ("local", 20),
        ("professor", 30),
        ("obs", 25),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    days = ['2a - Segunda', '3a - Terça', '4a - Quarta', '5a - Quinta', '6a - Sexta', 'Sábado']
    types = ['ensaio_banda', 'ensaio_percussao', 'aula_individual', 'aula_em_grupo',
             'aula_escola', 'aula_teoria', 'iniciacao']
    locations = ['Casa Tom', 'Silvia', 'Curvelo', 'E.M. Pereira Passos', 'E.M Estados Unidos',
                 'E.M. Mem de Sá', 'E.M. Vital Brasil', 'E.M. Maria Leopoldina']

    add_dropdown(ws, 'B', days)  # dia_semana
    add_dropdown(ws, 'D', types)  # tipo
    add_dropdown(ws, 'H', locations)  # local

    # Load from database
    c = conn.cursor()
    c.execute("""
        SELECT id, day_of_week, name, type, start_time, end_time, duration_hours, location, teacher
        FROM activities
        ORDER BY
            CASE day_of_week
                WHEN '2a - Segunda' THEN 1
                WHEN '3a - Terça' THEN 2
                WHEN '4a - Quarta' THEN 3
                WHEN '5a - Quinta' THEN 4
                WHEN '6a - Sexta' THEN 5
                WHEN 'Sábado' THEN 6
            END,
            start_time, name
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        # Empty obs column
        ws.cell(row=row_num, column=10, value='')


def create_atribuicao_atividades_sheet(wb, conn):
    """Create Atribuicao_Atividades (Activity Assignments) sheet with ID references."""
    ws = wb.create_sheet("Atribuicao_Atividades")

    headers = [
        ("id", 10),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP
        ("atividade_id", 12),
        ("atividade_nome", 35),
        ("data_inicio", 12),
        ("data_fim", 12),
        ("obs", 25),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    # Load from database
    c = conn.cursor()
    c.execute("""
        SELECT aa.id, aa.student_id, aa.activity_id, a.name, aa.start_date, aa.end_date, aa.notes
        FROM activity_assignments aa
        JOIN activities a ON a.id = aa.activity_id
        ORDER BY aa.activity_id, aa.student_id
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        aa_id, student_id, act_id, act_name, start, end, notes = row
        ws.cell(row=row_num, column=1, value=aa_id)
        ws.cell(row=row_num, column=2, value=student_id)
        ws.cell(row=row_num, column=3, value=f'=IFERROR(VLOOKUP(B{row_num},Alunos!A:B,2,FALSE),"")')
        ws.cell(row=row_num, column=4, value=act_id)
        ws.cell(row=row_num, column=5, value=act_name)
        ws.cell(row=row_num, column=6, value=start)
        ws.cell(row=row_num, column=7, value=end)
        ws.cell(row=row_num, column=8, value=notes)


def create_reparos_sheet(wb, conn):
    """Create Reparos (Repairs) sheet."""
    ws = wb.create_sheet("Reparos")

    headers = [
        ("id", 8),
        ("instrumento_id", 14),
        ("severidade", 15),
        ("problema", 45),
        ("orcamento", 12),
        ("status", 12),
        ("reportado_por", 20),
        ("data", 12),
        ("notas", 30),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    add_dropdown(ws, 'C', ['Não Utilizável', 'Utilizável com Restrições', 'Pequeno Reparo'])
    add_dropdown(ws, 'F', ['Pendente', 'Em Orçamento', 'Aprovado', 'Em Reparo', 'Concluído'])

    c = conn.cursor()
    c.execute("""
        SELECT id, instrument_id, severity, problem_description, budget,
               status, reported_by, date, notes
        FROM repairs
        ORDER BY date DESC
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Add next ID helper
    add_next_id_helper(ws, 'RE', col_letter='K')


def create_saidas_sheet(wb, conn):
    """Create Saidas (Departures) sheet for logging when students leave with explanations."""
    ws = wb.create_sheet("Saidas")

    headers = [
        ("id", 8),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP
        ("data", 12),
        ("motivo", 30),
        ("explicacao", 45),  # Required written explanation
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    # Dropdown for departure reasons
    reasons = [
        'Desmotivação', 'Questões de saúde', 'Distância/Transporte',
        'Questões familiares', 'Comportamento/Advertências', 'Maior de 18',
        'Responsáveis em desacordo', 'Intolerância religiosa',
        'Rendimento escolar', 'Outro'
    ]
    add_dropdown(ws, 'E', reasons)

    # Load existing departures (only actual departures, not bulk entries)
    c = conn.cursor()
    c.execute("""
        SELECT h.id, h.student_id, h.date, h.reason, h.notes
        FROM student_status_history h
        ORDER BY h.date DESC, h.student_id
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        hist_id, student_id, date, reason, notes = row
        ws.cell(row=row_num, column=1, value=hist_id)
        ws.cell(row=row_num, column=2, value=student_id)
        ws.cell(row=row_num, column=3, value=f'=IFERROR(VLOOKUP(B{row_num},Alunos!A:I,9,FALSE),"")')
        ws.cell(row=row_num, column=4, value=date)
        ws.cell(row=row_num, column=5, value=reason)
        ws.cell(row=row_num, column=6, value=notes)

    # Add next ID helper
    add_next_id_helper(ws, 'SA', col_letter='H')


def create_presenca_sheet(wb, conn):
    """Create Presenca (Attendance) sheet with ID references."""
    ws = wb.create_sheet("Presenca")

    headers = [
        ("id", 10),
        ("data", 12),
        ("atividade_id", 12),
        ("atividade_nome", 35),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP
        ("presente", 10),
        ("obs", 25),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    # Dropdown for attendance status (now col G)
    add_dropdown(ws, 'G', ['P', 'F', 'J', 'A'])  # Presente, Falta, Justificado, Atrasado

    # This sheet starts empty - populated as attendance is recorded
    # Staff enters aluno_id, name auto-fills via VLOOKUP

    # Add next ID helper
    add_next_id_helper(ws, 'PR', col_letter='J')


def create_aulas_particulares_sheet(wb, conn):
    """Create Aulas_Particulares (Private Lessons) sheet for tracking 1:1 lessons."""
    ws = wb.create_sheet("Aulas_Particulares")

    headers = [
        ("id", 10),
        ("professor", 20),
        ("aluno_id", 10),
        ("aluno_nome", 35),  # VLOOKUP
        ("dia", 12),
        ("horario", 10),
        ("duracao_min", 10),
        ("local", 15),
        ("ativo", 8),
    ]

    for col, (header, width) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_header(ws, len(headers))

    # Dropdowns
    teachers = ['Joe', 'Joana', 'Wesley', 'Raffaella', 'Antonio']
    days = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado']
    add_dropdown(ws, 'B', teachers)
    add_dropdown(ws, 'E', days)
    add_dropdown(ws, 'I', ['Sim', 'Não'])

    # Load existing data
    c = conn.cursor()
    c.execute("""
        SELECT id, teacher, student_id, day_of_week, start_time,
               duration_minutes, location, active
        FROM private_lessons
        ORDER BY teacher, day_of_week, start_time
    """)

    for row_num, row in enumerate(c.fetchall(), 2):
        lesson_id, teacher, student_id, day, time, duration, location, active = row
        ws.cell(row=row_num, column=1, value=lesson_id)
        ws.cell(row=row_num, column=2, value=teacher)
        ws.cell(row=row_num, column=3, value=student_id)
        ws.cell(row=row_num, column=4, value=f'=IFERROR(VLOOKUP(C{row_num},Alunos!A:I,9,FALSE),"")')
        ws.cell(row=row_num, column=5, value=day)
        ws.cell(row=row_num, column=6, value=time)
        ws.cell(row=row_num, column=7, value=duration)
        ws.cell(row=row_num, column=8, value=location)
        ws.cell(row=row_num, column=9, value='Sim' if active else 'Não')

    # Add next ID helper
    add_next_id_helper(ws, 'AP', col_letter='K')


def create_folha_pagamento_sheet(wb, conn):
    """Create Folha_Pagamento (Payroll) sheet with SUMIF formulas."""
    ws = wb.create_sheet("Folha_Pagamento")

    # Header section - months across the top
    months = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    weeks_per_month = [0, 4, 4, 4, 4, 4, 2, 4, 4, 4, 4, 3]  # 2026 estimates

    # Headers
    ws.cell(row=1, column=1, value="Professor")
    ws.cell(row=1, column=2, value="Valor/Hora")
    ws.cell(row=1, column=3, value="Horas/Semana")

    for col, month in enumerate(months, 4):
        ws.cell(row=1, column=col, value=month)
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.cell(row=1, column=16, value="Total Ano")

    # Weeks per month row
    ws.cell(row=2, column=1, value="Semanas no mês:")
    for col, weeks in enumerate(weeks_per_month, 4):
        ws.cell(row=2, column=col, value=weeks)

    style_header(ws, 16)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14

    # Get active teachers with hourly rates
    c = conn.cursor()
    c.execute("""
        SELECT name, hourly_rate
        FROM teachers
        WHERE status = 'Ativo' AND hourly_rate > 0
        ORDER BY name
    """)

    row_num = 3
    for name, rate in c.fetchall():
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=rate)
        # SUMIF formula: sum duracao_horas from Atividades where professor contains this name
        # Atividades!I:I is professor column, Atividades!G:G is duracao_horas
        ws.cell(row=row_num, column=3, value=f'=SUMIF(Atividades!I:I,"*"&A{row_num}&"*",Atividades!G:G)')

        # Monthly calculations: hours/week * weeks in month * hourly rate
        for col in range(4, 16):
            week_cell = f'{get_column_letter(col)}$2'
            hours_cell = f'$C{row_num}'
            rate_cell = f'$B{row_num}'
            ws.cell(row=row_num, column=col, value=f'={hours_cell}*{week_cell}*{rate_cell}')

        # Total year: sum of all months
        ws.cell(row=row_num, column=16, value=f'=SUM(D{row_num}:O{row_num})')
        row_num += 1

    # Total row
    last_row = row_num
    ws.cell(row=last_row, column=1, value="TOTAL")
    for col in range(4, 17):
        ws.cell(row=last_row, column=col, value=f'=SUM({get_column_letter(col)}3:{get_column_letter(col)}{row_num-1})')


def main():
    print("Connecting to database...")
    conn = sqlite3.connect(DB_PATH)

    print("Creating workbook...")
    wb = Workbook()

    print("Creating sheets...")
    create_alunos_sheet(wb, conn)
    create_instrumentos_sheet(wb, conn)
    create_emprestimos_sheet(wb, conn)
    create_reparos_sheet(wb, conn)
    create_avaliacoes_pratica_sheet(wb, conn)
    create_avaliacoes_teoria_sheet(wb, conn)
    create_grupos_sheet(wb, conn)
    create_atribuicao_grupos_sheet(wb, conn)
    create_professores_sheet(wb, conn)
    create_atividades_sheet(wb, conn)
    create_atribuicao_atividades_sheet(wb, conn)
    create_saidas_sheet(wb, conn)
    create_presenca_sheet(wb, conn)
    create_aulas_particulares_sheet(wb, conn)
    create_folha_pagamento_sheet(wb, conn)

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)

    conn.close()
    print("Done!")
    print(f"\nNext steps:")
    print(f"1. Open {OUTPUT_PATH}")
    print(f"2. Upload to Google Drive")
    print(f"3. Right-click → Open with → Google Sheets")
    print(f"4. Share with staff (edit permissions)")

if __name__ == "__main__":
    main()
