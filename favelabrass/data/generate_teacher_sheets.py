#!/usr/bin/env python3
"""
Generate per-teacher attendance sheets for Favela Brass.

Creates individual Excel files for each teacher with their assigned students,
ready for attendance tracking. These sheets link to the master database via
student IDs.

Usage:
    python generate_teacher_sheets.py

Output:
    outputs/attendance/Presenca_[TeacherName].xlsx

Workflow:
    1. Wesley fills in Atividades with actual timetable (GRP/IND activities)
    2. Wesley fills in Atribuicao_Atividades linking students to activities
    3. Run this script to generate per-teacher sheets
    4. Upload sheets to Google Drive, share with respective teachers
    5. Teachers record attendance weekly
    6. Iris/Wesley periodically import attendance back to master DB
"""

import sqlite3
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

DB_PATH = "/Users/tom/Documents/HQ/favelabrass/data/favelabrass.db"
OUTPUT_DIR = "/Users/tom/Documents/HQ/favelabrass/outputs/attendance"

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_teachers_with_activities(conn):
    """Get all teachers who have activities assigned."""
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT a.teacher, t.status
        FROM activities a
        LEFT JOIN teachers t ON t.name = a.teacher
        WHERE a.teacher IS NOT NULL AND a.teacher != ''
        ORDER BY a.teacher
    """)
    return [(row[0], row[1] or 'Ativo') for row in c.fetchall()]


def get_teacher_activities(conn, teacher_name):
    """Get all activities for a specific teacher."""
    c = conn.cursor()
    c.execute("""
        SELECT id, day_of_week, name, type, start_time, end_time, duration_hours, location
        FROM activities
        WHERE teacher = ?
        ORDER BY
            CASE day_of_week
                WHEN '2a - Segunda' THEN 1
                WHEN '3a - Terça' THEN 2
                WHEN '4a - Quarta' THEN 3
                WHEN '5a - Quinta' THEN 4
                WHEN '6a - Sexta' THEN 5
                WHEN 'Sábado' THEN 6
            END,
            start_time
    """, (teacher_name,))
    return c.fetchall()


def get_students_for_activity(conn, activity_id):
    """Get students assigned to an activity."""
    c = conn.cursor()
    c.execute("""
        SELECT aa.student_id, s.name
        FROM activity_assignments aa
        JOIN students s ON s.id = aa.student_id
        WHERE aa.activity_id = ?
          AND s.status = 'Ativo'
          AND (aa.end_date IS NULL OR aa.end_date = '' OR aa.end_date > date('now'))
        ORDER BY s.name
    """, (activity_id,))
    return c.fetchall()


def get_group_students_for_activity(conn, activity_name, activity_type):
    """
    For group activities, get students from group_assignments based on activity name.
    This is a fallback if activity_assignments isn't filled in for group activities.
    """
    # Map activity names to groups
    group_mapping = {
        'preta': 'Banda Preta',
        'roxa': 'Banda Roxa',
        'verde': 'Banda Verde',
        'amarela': 'Banda Amarela',
        'branca': 'Banda Branca',
        'semente': 'Grupo Semente',
    }

    group_id = None
    activity_lower = activity_name.lower()
    for keyword, group in group_mapping.items():
        if keyword in activity_lower:
            group_id = group
            break

    if not group_id:
        return []

    c = conn.cursor()
    c.execute("""
        SELECT ga.student_id, s.name
        FROM group_assignments ga
        JOIN students s ON s.id = ga.student_id
        WHERE ga.group_id = ?
          AND s.status = 'Ativo'
        ORDER BY s.name
    """, (group_id,))
    return c.fetchall()


def generate_date_columns(start_date=None, weeks=12):
    """Generate weekly date columns for attendance."""
    if start_date is None:
        # Start from next Monday
        today = datetime.now()
        days_ahead = 7 - today.weekday()  # Days until next Monday
        if days_ahead == 7:
            days_ahead = 0
        start_date = today + timedelta(days=days_ahead)

    dates = []
    for week in range(weeks):
        date = start_date + timedelta(weeks=week)
        dates.append(date.strftime('%d/%m'))
    return dates


def create_teacher_workbook(conn, teacher_name):
    """Create attendance workbook for a specific teacher."""
    wb = Workbook()

    activities = get_teacher_activities(conn, teacher_name)

    if not activities:
        # No activities - create placeholder sheet
        ws = wb.active
        ws.title = "Sem Atividades"
        ws.cell(row=1, column=1, value=f"Nenhuma atividade atribuída para {teacher_name}")
        return wb

    # Group activities by day
    activities_by_day = {}
    for act in activities:
        day = act[1] or 'Outro'
        if day not in activities_by_day:
            activities_by_day[day] = []
        activities_by_day[day].append(act)

    first_sheet = True
    date_columns = generate_date_columns()

    for day, day_activities in activities_by_day.items():
        if first_sheet:
            ws = wb.active
            ws.title = day[:10]  # Limit sheet name length
            first_sheet = False
        else:
            ws = wb.create_sheet(day[:10])

        row_num = 1

        # Day header
        ws.cell(row=row_num, column=1, value=f"{day} - {teacher_name}")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        row_num += 2

        for activity in day_activities:
            act_id, _, act_name, act_type, start_time, end_time, duration, location = activity

            # Activity header
            ws.cell(row=row_num, column=1, value=f"{act_name}")
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=f"{start_time}-{end_time}")
            ws.cell(row=row_num, column=3, value=location)
            row_num += 1

            # Get students for this activity
            students = get_students_for_activity(conn, act_id)

            # Fallback: for group activities, get from group_assignments
            if not students and act_type in ('ensaio_banda', 'ensaio_percussao', 'aula_em_grupo'):
                students = get_group_students_for_activity(conn, act_name, act_type)

            if not students:
                ws.cell(row=row_num, column=1, value="(sem alunos atribuídos)")
                ws.cell(row=row_num, column=1).font = Font(italic=True, color="888888")
                row_num += 2
                continue

            # Column headers: ID | Nome | Date1 | Date2 | ...
            headers = ['ID', 'Nome'] + date_columns
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.fill = SUBHEADER_FILL
                cell.border = THIN_BORDER
                cell.font = Font(bold=True)
                if col <= 2:
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')
            row_num += 1

            # Student rows
            for student_id, student_name in students:
                ws.cell(row=row_num, column=1, value=student_id).border = THIN_BORDER
                ws.cell(row=row_num, column=2, value=student_name).border = THIN_BORDER

                # Empty cells for attendance with borders
                for col in range(3, 3 + len(date_columns)):
                    cell = ws.cell(row=row_num, column=col, value='')
                    cell.border = THIN_BORDER

                row_num += 1

            # Add attendance dropdown validation (P/F/J/A)
            if students:
                dv = DataValidation(
                    type="list",
                    formula1='"P,F,J,A"',
                    allow_blank=True
                )
                dv.error = "Use: P (Presente), F (Falta), J (Justificado), A (Atrasado)"
                dv.errorTitle = "Valor inválido"
                ws.add_data_validation(dv)
                start_row = row_num - len(students)
                end_row = row_num - 1
                for col in range(3, 3 + len(date_columns)):
                    col_letter = get_column_letter(col)
                    dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')

            row_num += 1  # Space between activities

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        for col in range(3, 3 + len(date_columns)):
            ws.column_dimensions[get_column_letter(col)].width = 8

    return wb


def create_saturday_master_sheet(conn):
    """
    Create the Saturday master attendance sheet for Wesley/Iris.
    All Saturday activities in one workbook with all assigned students.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sábado - Curvelo"

    c = conn.cursor()
    c.execute("""
        SELECT id, name, type, start_time, end_time, location, teacher
        FROM activities
        WHERE day_of_week = 'Sábado'
        ORDER BY start_time, name
    """)
    activities = c.fetchall()

    date_columns = generate_date_columns()
    row_num = 1

    # Title
    ws.cell(row=row_num, column=1, value="Presença Sábado - Curvelo")
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=16)
    row_num += 2

    for activity in activities:
        act_id, act_name, act_type, start_time, end_time, location, teacher = activity

        # Activity header
        header_text = f"{act_name} ({start_time}-{end_time})"
        if teacher:
            header_text += f" - {teacher}"
        ws.cell(row=row_num, column=1, value=header_text)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1

        # Get students
        students = get_students_for_activity(conn, act_id)
        if not students and act_type in ('ensaio_banda', 'ensaio_percussao', 'aula_em_grupo', 'aula_teoria', 'iniciacao'):
            students = get_group_students_for_activity(conn, act_name, act_type)

        if not students:
            ws.cell(row=row_num, column=1, value="(sem alunos atribuídos)")
            ws.cell(row=row_num, column=1).font = Font(italic=True, color="888888")
            row_num += 2
            continue

        # Column headers
        headers = ['ID', 'Nome'] + date_columns
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER
            cell.font = Font(bold=True)
        row_num += 1

        # Student rows
        for student_id, student_name in students:
            ws.cell(row=row_num, column=1, value=student_id).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=student_name).border = THIN_BORDER
            for col in range(3, 3 + len(date_columns)):
                ws.cell(row=row_num, column=col, value='').border = THIN_BORDER
            row_num += 1

        # Attendance dropdown
        if students:
            dv = DataValidation(type="list", formula1='"P,F,J,A"', allow_blank=True)
            ws.add_data_validation(dv)
            start_row = row_num - len(students)
            end_row = row_num - 1
            for col in range(3, 3 + len(date_columns)):
                col_letter = get_column_letter(col)
                dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')

        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    for col in range(3, 3 + len(date_columns)):
        ws.column_dimensions[get_column_letter(col)].width = 8

    return wb


def create_school_sheet(conn, school_name, location):
    """
    Create attendance sheet for a school workshop.
    One sheet per school location.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = school_name[:31]  # Excel sheet name limit

    c = conn.cursor()
    c.execute("""
        SELECT id, day_of_week, name, type, start_time, end_time, teacher
        FROM activities
        WHERE location = ? AND type = 'aula_escola'
        ORDER BY
            CASE day_of_week
                WHEN '2a - Segunda' THEN 1
                WHEN '3a - Terça' THEN 2
                WHEN '4a - Quarta' THEN 3
                WHEN '5a - Quinta' THEN 4
                WHEN '6a - Sexta' THEN 5
            END,
            start_time
    """, (location,))
    activities = c.fetchall()

    if not activities:
        ws.cell(row=1, column=1, value=f"Sem aulas registradas em {school_name}")
        return wb

    date_columns = generate_date_columns()
    row_num = 1

    # Title
    ws.cell(row=row_num, column=1, value=f"Presença - {school_name}")
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=16)
    row_num += 2

    for activity in activities:
        act_id, day, act_name, act_type, start_time, end_time, teacher = activity

        # Activity header
        header_text = f"{day} {start_time}-{end_time}"
        if teacher:
            header_text += f" ({teacher})"
        ws.cell(row=row_num, column=1, value=header_text)
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1

        # Get students - for schools, need to check which students go to this school
        # and are assigned to escola program
        c2 = conn.cursor()
        c2.execute("""
            SELECT id, name FROM students
            WHERE school = ? AND status = 'Ativo' AND program_type = 'escola'
            ORDER BY name
        """, (school_name,))
        students = c2.fetchall()

        if not students:
            ws.cell(row=row_num, column=1, value="(sem alunos)")
            ws.cell(row=row_num, column=1).font = Font(italic=True, color="888888")
            row_num += 2
            continue

        # Headers
        headers = ['ID', 'Nome'] + date_columns
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER
            cell.font = Font(bold=True)
        row_num += 1

        # Students
        for student_id, student_name in students:
            ws.cell(row=row_num, column=1, value=student_id).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=student_name).border = THIN_BORDER
            for col in range(3, 3 + len(date_columns)):
                ws.cell(row=row_num, column=col, value='').border = THIN_BORDER
            row_num += 1

        # Dropdown
        if students:
            dv = DataValidation(type="list", formula1='"P,F,J,A"', allow_blank=True)
            ws.add_data_validation(dv)
            start_row = row_num - len(students)
            end_row = row_num - 1
            for col in range(3, 3 + len(date_columns)):
                dv.add(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}')

        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    for col in range(3, 3 + len(date_columns)):
        ws.column_dimensions[get_column_letter(col)].width = 8

    return wb


def main():
    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Connecting to database...")
    conn = sqlite3.connect(DB_PATH)

    # Check if we have activity data
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM activities")
    activity_count = c.fetchone()[0]

    if activity_count == 0:
        print("\n*** WARNING: No activities in database! ***")
        print("Wesley needs to fill in the Atividades sheet in the master spreadsheet first.")
        print("Once activities exist, re-run this script to generate attendance sheets.\n")

    # 1. Generate per-teacher sheets
    print("\n=== Generating per-teacher attendance sheets ===")
    teachers = get_teachers_with_activities(conn)

    if not teachers:
        print("No teachers with activities found.")
    else:
        for teacher_name, status in teachers:
            safe_name = teacher_name.replace(' ', '_').replace('/', '_')
            filename = f"Presenca_{safe_name}.xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)

            wb = create_teacher_workbook(conn, teacher_name)
            wb.save(filepath)
            print(f"  Created: {filename}")

    # 2. Generate Saturday master sheet
    print("\n=== Generating Saturday master sheet ===")
    c.execute("SELECT COUNT(*) FROM activities WHERE day_of_week = 'Sábado'")
    saturday_count = c.fetchone()[0]

    if saturday_count > 0:
        wb = create_saturday_master_sheet(conn)
        filepath = os.path.join(OUTPUT_DIR, "Presenca_Sabado_Curvelo.xlsx")
        wb.save(filepath)
        print(f"  Created: Presenca_Sabado_Curvelo.xlsx")
    else:
        print("  No Saturday activities found.")

    # 3. Generate school sheets
    print("\n=== Generating school attendance sheets ===")
    c.execute("""
        SELECT DISTINCT location FROM activities
        WHERE type = 'aula_escola' AND location IS NOT NULL
    """)
    schools = [row[0] for row in c.fetchall()]

    if not schools:
        # Fallback: get schools from students table
        c.execute("""
            SELECT DISTINCT school FROM students
            WHERE school IS NOT NULL AND school != '' AND program_type = 'escola'
        """)
        schools = [row[0] for row in c.fetchall()]

    for school in schools:
        safe_name = school.replace(' ', '_').replace('.', '').replace('/', '_')
        filename = f"Presenca_Escola_{safe_name}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)

        wb = create_school_sheet(conn, school, school)
        wb.save(filepath)
        print(f"  Created: {filename}")

    conn.close()

    print(f"\n=== Done! ===")
    print(f"Attendance sheets saved to: {OUTPUT_DIR}")
    print(f"\nNext steps:")
    print(f"1. Wesley fills in Atividades with actual timetable")
    print(f"2. Wesley fills in Atribuicao_Atividades linking students to activities")
    print(f"3. Re-run this script to regenerate sheets with correct student lists")
    print(f"4. Upload to Google Drive and share with respective teachers")


if __name__ == "__main__":
    main()
