#!/usr/bin/env python3
"""Eventos — Event production management for Favela Brass.

Served at portal.favelabrass.org/eventos/ behind Google OAuth.
No password gate — OAuth is sufficient.

Run locally:  python3 eventos_web.py
Deploy:       scp eventos_web.py root@45.55.73.116:/root/ && ssh root@45.55.73.116 "systemctl restart eventos-web"
"""

import sqlite3
import os
import json
from pathlib import Path
from datetime import date, timedelta
import csv
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
from flask import Flask, request, redirect, render_template_string, jsonify, Response

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "fb-eventos-2026-change-me")

DB_PATH = Path("/root/favelabrass.db") if Path("/root/favelabrass.db").exists() else Path(__file__).resolve().parent.parent / "data" / "favelabrass.db"

PREFIX = "/eventos"


@app.context_processor
def inject_prefix():
    return dict(P=PREFIX)


def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def format_brl(amount):
    if amount is None:
        return "—"
    try:
        formatted = f"{float(amount):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {formatted}"
    except (TypeError, ValueError):
        return "—"


def format_date(d):
    if not d:
        return "—"
    try:
        parts = str(d).split("-")
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except (IndexError, ValueError):
        return str(d)


def format_phone(p):
    if not p:
        return ""
    digits = "".join(ch for ch in str(p) if ch.isdigit())
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return str(p)


def is_overnight_trip(event):
    """True when an event spans multiple days — i.e. needs a manifest."""
    if not event:
        return False
    end = event["end_date"] if hasattr(event, "keys") else event.get("end_date")
    start = event["date"] if hasattr(event, "keys") else event.get("date")
    return bool(end and end != start)


def participant_missing_fields(p):
    """Return list of missing identity fields for a participant row."""
    missing = []
    if not (p["birth_date"] if "birth_date" in p.keys() else None): missing.append("DOB")
    if not (p["child_rg"] if "child_rg" in p.keys() else None): missing.append("RG")
    if not (p["child_cpf"] if "child_cpf" in p.keys() else None): missing.append("CPF")
    if not (p["guardian1_phone"] if "guardian1_phone" in p.keys() else None): missing.append("Tel resp.")
    return missing


def staff_missing_fields(s):
    """Return list of missing identity fields for an equipe row."""
    missing = []
    has_teacher = s["teacher_id"] if "teacher_id" in s.keys() else None
    if not has_teacher:
        return ["dados pessoais (sem cadastro)"]
    if not (s["t_birth_date"] if "t_birth_date" in s.keys() else None): missing.append("DOB")
    if not (s["t_rg"] if "t_rg" in s.keys() else None): missing.append("RG")
    if not (s["t_cpf"] if "t_cpf" in s.keys() else None): missing.append("CPF")
    if not (s["t_phone"] if "t_phone" in s.keys() else None): missing.append("Tel")
    return missing


TYPE_LABELS = {
    'show': '🎵 Apresentação',
    'festival': '🎪 Festival',
    'viagem': '✈️ Viagem',
    'reuniao': '📋 Reunião',
    'milestone': '📌 Marco',
    'ensaio': '🎶 Ensaio',
    'workshop': '🎓 Workshop',
}
EVENT_TYPES = ['show', 'festival', 'viagem', 'reuniao', 'milestone', 'ensaio', 'workshop']
# Types that need production fichas (cronograma, equipe, orçamento, etc.)
FICHA_TYPES = ['show', 'festival', 'viagem', 'workshop']
EVENT_STATUSES = ['planned', 'confirmed', 'completed', 'cancelled']
STATUS_LABELS = {'planned': 'Planejado', 'confirmed': 'Confirmado', 'completed': 'Já realizado', 'cancelled': 'Cancelado'}

LOGISTICS_CATEGORIES = ['transporte', 'lanche', 'equipamento', 'material', 'geral']
CONTACT_CATEGORIES = ['local', 'transporte', 'alimentacao', 'emergencia', 'parceiro', 'geral']
BUDGET_CATEGORIES = ['Transporte', 'Alimentação', 'Hospedagem', 'Cachê', 'Equipamento', 'Material', 'Produção', 'Comunicação', 'Outro']

CHECKLIST_PHASES = [
    ('pre', 'Pré-produção'),
    ('dia', 'No dia'),
    ('pos', 'Pós-evento'),
]

CHECKLIST_TEMPLATES = {
    'show': {
        'pre': [
            'Confirmar local',
            'Confirmar transporte (van)',
            'Definir repertório',
            'Autorização dos responsáveis',
            'Enviar convites',
            'Separar equipamento de som (Roland PA, mic, cabo, pedestal)',
        ],
        'dia': [
            'Montar e testar PA (Roland, mic, cabo, pedestal)',
            'Soundcheck',
            'Servir lanche',
            'Filmar para redes sociais',
        ],
        'pos': [
            'Postar nas redes sociais',
            'Agradecer parceiros',
            'Desmontar e guardar equipamento de som',
        ],
    },
    'festival': {
        'pre': [
            'Confirmar inscrição / participação',
            'Confirmar transporte',
            'Definir repertório',
            'Autorização dos responsáveis',
            'Preparar materiais de divulgação',
        ],
        'dia': [
            'Montar equipamento',
            'Soundcheck',
            'Servir lanche',
            'Filmar para redes sociais',
        ],
        'pos': [
            'Postar nas redes sociais',
            'Agradecer organizadores',
            'Devolver equipamento',
        ],
    },
    'viagem': {
        'pre': [
            'Confirmar hospedagem',
            'Confirmar transporte',
            'Autorização dos responsáveis',
            'Lista de bagagem / equipamento',
            'Seguro viagem',
            'Definir repertório',
        ],
        'dia': [
            'Conferir lista de presença',
            'Distribuir crachás / identificação',
            'Filmar para redes sociais',
        ],
        'pos': [
            'Postar nas redes sociais',
            'Agradecer anfitriões',
            'Devolver equipamento',
            'Relatório de viagem',
        ],
    },
    'workshop': {
        'pre': [
            'Confirmar local e sala',
            'Confirmar materiais necessários',
            'Enviar convites / divulgação',
        ],
        'dia': [
            'Preparar sala',
            'Lista de presença',
            'Filmar para redes sociais',
        ],
        'pos': [
            'Postar nas redes sociais',
            'Agradecer facilitador',
        ],
    },
}


# ══════════════════════════════════════════════════════════════════════════
# CSS / HTML TEMPLATES
# ══════════════════════════════════════════════════════════════════════════

BASE_CSS = """
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&display=swap');

:root {
    --purple: #5A0E7A;
    --purple-dark: #3E0B59;
    --purple-light: #7B2FA0;
    --yellow: #FEF100;
    --green: #62CC3C;
    --red: #E74C3C;
    --orange: #E67E22;
    --blue: #3498DB;
    --text: #fff;
    --text-muted: rgba(255,255,255,0.6);
    --text-dim: rgba(255,255,255,0.4);
    --border: rgba(255,255,255,0.1);
    --border-light: rgba(255,255,255,0.06);
    --card-bg: rgba(255,255,255,0.05);
    --card-hover: rgba(255,255,255,0.08);
    --input-bg: rgba(255,255,255,0.08);
    --section-bg: rgba(255,255,255,0.03);
}

* { margin: 0; padding: 0; box-sizing: border-box; }

body {
    font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    background: linear-gradient(135deg, var(--purple-dark) 0%, var(--purple) 100%);
    color: var(--text);
    min-height: 100vh;
    line-height: 1.5;
    font-size: 14px;
}

a { color: var(--yellow); text-decoration: none; }
a:hover { text-decoration: underline; }

.container { max-width: 1100px; margin: 0 auto; padding: 0 20px 48px; }

.nav {
    background: rgba(0,0,0,0.35);
    border-bottom: 1px solid var(--border);
    padding: 0 20px;
    display: flex;
    align-items: center;
    gap: 0;
    overflow-x: auto;
    position: sticky;
    top: 0;
    z-index: 100;
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
}

.nav .nav-portal {
    color: var(--text-dim);
    font-size: 16px;
    padding: 12px 8px 12px 4px;
    text-decoration: none;
    transition: color 0.15s;
}
.nav .nav-portal:hover { color: #fff; background: none; }

.nav .brand {
    font-weight: 700;
    font-size: 15px;
    color: var(--yellow);
    padding: 12px 12px 12px 0;
    white-space: nowrap;
    letter-spacing: -0.02em;
    text-decoration: none;
    border-right: 1px solid var(--border);
    margin-right: 4px;
}
.nav .brand:hover { background: none; opacity: 0.85; }

.nav a {
    color: var(--text-muted);
    padding: 12px 14px;
    font-size: 13px;
    font-weight: 500;
    white-space: nowrap;
    transition: color 0.15s, background 0.15s;
    border-radius: 6px 6px 0 0;
    text-decoration: none;
}

.nav a:hover { color: #fff; background: rgba(255,255,255,0.06); }
.nav a.active { color: var(--yellow); }

.page-header {
    padding: 28px 0 20px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 16px;
    flex-wrap: wrap;
}

.page-header h1 {
    font-size: 24px;
    font-weight: 700;
    letter-spacing: -0.02em;
}

.stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin-bottom: 24px; }

.stat-card {
    background: var(--card-bg);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 16px 20px;
}

.stat-card .label { font-size: 11px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-muted); font-weight: 500; }
.stat-card .value { font-size: 28px; font-weight: 700; margin-top: 4px; letter-spacing: -0.02em; }

.section {
    background: var(--section-bg);
    border: 1px solid var(--border-light);
    border-radius: 16px;
    padding: 24px;
    margin-bottom: 20px;
}

.section h2 {
    color: var(--yellow);
    font-size: 16px;
    font-weight: 600;
    margin-bottom: 16px;
    padding-bottom: 10px;
    border-bottom: 1px solid var(--border);
    letter-spacing: -0.01em;
}

.table-wrap { overflow-x: auto; margin: 0 -4px; }

table { width: 100%; border-collapse: collapse; }

th {
    text-align: left;
    color: var(--text-dim);
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    padding: 8px 12px;
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
}

td {
    padding: 10px 12px;
    border-bottom: 1px solid var(--border-light);
    color: rgba(255,255,255,0.9);
    font-size: 13px;
}

tr:hover td { background: rgba(255,255,255,0.02); }

.btn {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 8px 16px;
    border-radius: 8px;
    font-size: 13px;
    font-weight: 600;
    font-family: inherit;
    cursor: pointer;
    transition: all 0.15s;
    border: none;
    text-decoration: none;
}

.btn-primary { background: var(--yellow); color: var(--purple-dark); }
.btn-primary:hover { background: #fff; text-decoration: none; }

.btn-secondary { background: var(--card-bg); color: var(--text); border: 1px solid var(--border); }
.btn-secondary:hover { background: var(--card-hover); text-decoration: none; }

.btn-danger { background: rgba(231,76,60,0.15); color: var(--red); border: 1px solid rgba(231,76,60,0.3); }
.btn-danger:hover { background: rgba(231,76,60,0.25); text-decoration: none; }

.btn-sm { padding: 5px 10px; font-size: 12px; }

.form-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
.form-full { grid-column: 1 / -1; }
.form-group { margin-bottom: 0; }

.form-group label {
    display: block;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--text-muted);
    font-weight: 500;
    margin-bottom: 6px;
}

.form-group input, .form-group select, .form-group textarea {
    width: 100%;
    padding: 10px 12px;
    background: var(--input-bg);
    border: 1px solid var(--border);
    border-radius: 8px;
    color: var(--text);
    font-family: inherit;
    font-size: 14px;
    transition: border-color 0.15s;
}

.form-group input:focus, .form-group select:focus, .form-group textarea:focus {
    outline: none;
    border-color: var(--yellow);
}

.form-group textarea { resize: vertical; min-height: 80px; }
.form-group select { appearance: auto; }
.form-group select option { background: #3E0B59; color: #fff; }

.form-actions { display: flex; gap: 12px; margin-top: 24px; }

.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.02em;
}

.badge-confirmed { background: rgba(98,204,60,0.2); color: var(--green); }
.badge-planned { background: rgba(254,241,0,0.15); color: var(--yellow); }
.badge-cancelled { background: rgba(255,255,255,0.1); color: var(--text-muted); }
.badge-concluido { background: rgba(98,204,60,0.2); color: var(--green); }
.badge-pendente { background: rgba(254,241,0,0.15); color: var(--yellow); }
.badge-em_andamento { background: rgba(52,152,219,0.2); color: var(--blue); }
.badge-rouanet { background: rgba(155,89,182,0.25); color: #c39bd3; }
.badge-associacao { background: rgba(98,204,60,0.15); color: var(--green); }
.badge-alta { background: rgba(254,241,0,0.15); color: var(--yellow); }
.badge-urgente { background: rgba(231,76,60,0.2); color: var(--red); }
.badge-normal { background: rgba(255,255,255,0.08); color: var(--text-muted); }

.filters {
    display: flex;
    gap: 12px;
    align-items: center;
    flex-wrap: wrap;
    margin-bottom: 16px;
}

.filters select, .filters input {
    padding: 7px 12px;
    background: var(--input-bg);
    border: 1px solid var(--border);
    border-radius: 8px;
    color: var(--text);
    font-family: inherit;
    font-size: 13px;
}

.filters select:focus, .filters input:focus { outline: none; border-color: var(--yellow); }

.empty {
    text-align: center;
    padding: 40px 20px;
    color: var(--text-muted);
}

.empty-icon { font-size: 40px; margin-bottom: 8px; }

/* Tabs */
.tabs { margin-bottom: 0; }
.tab-input { display: none; }
.tab-labels {
    display: flex;
    gap: 0;
    border-bottom: 2px solid var(--border);
    margin-bottom: 20px;
    overflow-x: auto;
}
.tab-label {
    padding: 10px 16px;
    font-size: 13px;
    font-weight: 500;
    color: var(--text-muted);
    cursor: pointer;
    white-space: nowrap;
    border-bottom: 2px solid transparent;
    margin-bottom: -2px;
    transition: color 0.15s;
}
.tab-label:hover { color: #fff; }
.tab-content { display: none; }
#tab-resumo:checked ~ .tab-labels .tab-label[for="tab-resumo"],
#tab-cronograma:checked ~ .tab-labels .tab-label[for="tab-cronograma"],
#tab-equipe:checked ~ .tab-labels .tab-label[for="tab-equipe"],
#tab-participantes:checked ~ .tab-labels .tab-label[for="tab-participantes"],
#tab-logistica:checked ~ .tab-labels .tab-label[for="tab-logistica"],
#tab-orcamento:checked ~ .tab-labels .tab-label[for="tab-orcamento"],
#tab-contatos:checked ~ .tab-labels .tab-label[for="tab-contatos"],
#tab-tarefas:checked ~ .tab-labels .tab-label[for="tab-tarefas"] {
    color: var(--yellow);
    border-bottom-color: var(--yellow);
}
#tab-resumo:checked ~ .tab-panels .panel-resumo,
#tab-cronograma:checked ~ .tab-panels .panel-cronograma,
#tab-equipe:checked ~ .tab-panels .panel-equipe,
#tab-participantes:checked ~ .tab-panels .panel-participantes,
#tab-logistica:checked ~ .tab-panels .panel-logistica,
#tab-orcamento:checked ~ .tab-panels .panel-orcamento,
#tab-contatos:checked ~ .tab-panels .panel-contatos,
#tab-tarefas:checked ~ .tab-panels .panel-tarefas { display: block; }

.task-done { text-decoration: line-through; color: var(--text-muted); }

/* ── Cronograma (agile timeline) ──── */
.crono-add {
    display: flex;
    gap: 8px;
    margin-bottom: 20px;
}
.crono-add input {
    flex: 1;
    padding: 12px 16px;
    background: var(--input-bg);
    border: 1px solid var(--border);
    border-radius: 10px;
    color: var(--text);
    font-family: inherit;
    font-size: 15px;
    transition: border-color 0.15s;
}
.crono-add input:focus { outline: none; border-color: var(--yellow); }
.crono-add input::placeholder { color: var(--text-dim); }

.crono-list { display: flex; flex-direction: column; gap: 2px; }

.crono-item {
    display: flex;
    align-items: center;
    gap: 0;
    background: var(--card-bg);
    border: 1px solid var(--border-light);
    border-radius: 8px;
    padding: 0;
    transition: border-color 0.2s, background 0.2s;
    position: relative;
}
.crono-item:hover { border-color: var(--border); background: var(--card-hover); }
.crono-item.crono-saved { border-color: var(--green); }

.crono-error {
    color: #e74c3c;
    background: rgba(231, 76, 60, 0.1);
    border: 1px solid rgba(231, 76, 60, 0.3);
    border-radius: 6px;
    padding: 8px 12px;
    margin-bottom: 8px;
    font-size: 13px;
    animation: fadeIn 0.2s ease;
}

.crono-time {
    min-width: 72px;
    padding: 10px 12px;
    font-family: 'DM Sans', monospace;
    font-weight: 700;
    font-size: 14px;
    color: var(--yellow);
    white-space: nowrap;
    cursor: pointer;
    border-right: 1px solid var(--border-light);
    text-align: center;
}
.crono-time .crono-end { font-weight: 400; font-size: 11px; color: var(--text-dim); display: block; }
.crono-time:hover { background: rgba(255,255,255,0.04); }

.crono-body {
    flex: 1;
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 10px 12px;
    min-width: 0;
    flex-wrap: wrap;
}

.crono-activity {
    font-size: 14px;
    font-weight: 500;
    color: var(--text);
    cursor: pointer;
    min-width: 100px;
    flex: 1;
}
.crono-activity:hover { color: var(--yellow); }

.crono-location, .crono-responsible {
    font-size: 12px;
    color: var(--text-muted);
    cursor: pointer;
    white-space: nowrap;
    padding: 2px 8px;
    border-radius: 4px;
    background: rgba(255,255,255,0.04);
}
.crono-location:hover, .crono-responsible:hover { color: var(--text); background: rgba(255,255,255,0.08); }
.crono-location::before { content: '📍 '; }
.crono-responsible::before { content: '👤 '; }

.crono-actions {
    display: flex;
    align-items: center;
    gap: 2px;
    padding: 4px 8px;
    opacity: 0.3;
    transition: opacity 0.15s;
}
.crono-item:hover .crono-actions { opacity: 1; }

.crono-actions button {
    background: none;
    border: none;
    color: var(--text-muted);
    cursor: pointer;
    padding: 4px 6px;
    border-radius: 4px;
    font-size: 13px;
    transition: background 0.1s, color 0.1s;
}
.crono-actions button:hover { background: rgba(255,255,255,0.1); color: var(--text); }
.crono-actions .crono-del:hover { color: var(--red); }

/* Inline edit inputs */
.crono-edit-input {
    background: var(--input-bg);
    border: 1px solid var(--yellow);
    border-radius: 4px;
    color: var(--text);
    font-family: inherit;
    padding: 4px 8px;
    font-size: inherit;
}
.crono-edit-input:focus { outline: none; }

.crono-empty {
    text-align: center;
    padding: 40px 20px;
    color: var(--text-dim);
    font-size: 14px;
}
.crono-hint {
    font-size: 12px;
    color: var(--text-dim);
    margin-top: 6px;
}

.crono-day-header {
    margin: 16px 0 6px;
    padding: 6px 12px;
    background: var(--card-hover);
    border-left: 3px solid var(--yellow);
    border-radius: 4px;
    font-weight: 600;
    font-size: 12px;
    color: var(--text);
    text-transform: uppercase;
    letter-spacing: 0.5px;
}
.crono-day-header:first-child { margin-top: 0; }
.crono-day-header.crono-day-undated { border-left-color: var(--red); color: var(--text-dim); }
.crono-date-pill {
    display: inline-flex;
    align-items: center;
    padding: 1px 6px;
    border-radius: 3px;
    background: rgba(255,255,255,0.04);
    font-size: 11px;
    color: var(--text-dim);
    cursor: pointer;
}
.crono-date-pill:hover { background: rgba(255,255,255,0.08); color: var(--text); }
.crono-date-pill.crono-date-missing { color: var(--red); background: rgba(244,67,54,0.08); }

@media (max-width: 700px) {
    .form-grid { grid-template-columns: 1fr; }
    .stats { grid-template-columns: repeat(2, 1fr); }
    .page-header { flex-direction: column; align-items: flex-start; }
    .nav { padding: 0 8px; }
    .nav a { padding: 10px 10px; font-size: 12px; }
    .section { padding: 16px; }
    .crono-actions { opacity: 1; }
    .crono-body { flex-direction: column; align-items: flex-start; gap: 4px; }
    .crono-add input { font-size: 14px; padding: 10px 12px; }
}
"""

NAV_HTML = """
<nav class="nav">
    <a href="/" class="nav-portal" title="Voltar ao portal">←</a>
    <a href="{{ P }}/" class="brand">🎪 Eventos</a>
    <a href="{{ P }}/" class="{{ 'active' if page == 'dashboard' else '' }}">Fichas</a>
    <a href="{{ P }}/lista" class="{{ 'active' if page == 'lista' else '' }}">Lista Completa</a>
    <a href="{{ P }}/orcamento" class="{{ 'active' if page == 'orcamento' else '' }}">Orçamento</a>
</nav>
"""

PAGE_OPEN = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }} — Eventos FB</title>
    <style>""" + BASE_CSS + """</style>
</head>
<body>
<div class="app-content">
""" + NAV_HTML

PAGE_CLOSE = """
</div>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════

@app.route("/")
def dashboard():
    conn = get_db()
    today = date.today().isoformat()
    year = date.today().year
    in_90 = (date.today() + timedelta(days=90)).isoformat()

    # Only count production events (shows, festivals, viagens, workshops)
    ficha_placeholders = ','.join('?' * len(FICHA_TYPES))

    total_eventos = conn.execute(
        f"SELECT COUNT(*) FROM events WHERE status != 'cancelled' AND type IN ({ficha_placeholders}) AND strftime('%Y', date) = ?",
        (*FICHA_TYPES, str(year))
    ).fetchone()[0]

    proximos = conn.execute(
        f"SELECT COUNT(*) FROM events WHERE date BETWEEN ? AND ? AND status != 'cancelled' AND type IN ({ficha_placeholders})",
        (today, in_90, *FICHA_TYPES)
    ).fetchone()[0]

    tarefas_pendentes = conn.execute(
        "SELECT COUNT(*) FROM event_tasks WHERE status IN ('pendente', 'em_andamento')"
    ).fetchone()[0]

    orcamento_total = conn.execute("""
        SELECT COALESCE(SUM(eb.estimated_amount), 0)
        FROM event_budget eb
        JOIN events e ON eb.event_id = e.id
        WHERE e.date >= ? AND e.status != 'cancelled' AND eb.is_income = 0
    """, (today,)).fetchone()[0]

    upcoming_events = conn.execute(f"""
        SELECT * FROM events
        WHERE date BETWEEN ? AND ? AND status != 'cancelled' AND type IN ({ficha_placeholders})
        ORDER BY date, time
        LIMIT 15
    """, (today, in_90, *FICHA_TYPES)).fetchall()

    overdue_tasks = conn.execute("""
        SELECT t.*, e.name as event_name, e.id as event_id
        FROM event_tasks t
        JOIN events e ON t.event_id = e.id
        WHERE t.status IN ('pendente', 'em_andamento')
        AND t.deadline IS NOT NULL AND t.deadline < ?
        ORDER BY t.deadline
        LIMIT 10
    """, (today,)).fetchall()

    conn.close()

    return render_template_string(PAGE_OPEN.replace("{{ title }}", "Fichas de Produção") + DASHBOARD_BODY + PAGE_CLOSE,
        page='dashboard',
        total_eventos=total_eventos, proximos=proximos,
        tarefas_pendentes=tarefas_pendentes, orcamento_total=orcamento_total,
        upcoming_events=upcoming_events, overdue_tasks=overdue_tasks,
        format_date=format_date, format_brl=format_brl,
        TYPE_LABELS=TYPE_LABELS, STATUS_LABELS=STATUS_LABELS, today=today)


@app.route("/lista")
def lista():
    conn = get_db()
    type_filter = request.args.get("tipo", "")
    status_filter = request.args.get("status", "")
    year_filter = request.args.get("ano", str(date.today().year))

    query = "SELECT * FROM events WHERE 1=1"
    params = []
    if type_filter:
        if type_filter == "all":
            pass  # No type filter — show everything
        else:
            query += " AND type = ?"
            params.append(type_filter)
    else:
        # Default: only show production-worthy event types
        ficha_placeholders = ','.join('?' * len(FICHA_TYPES))
        query += f" AND type IN ({ficha_placeholders})"
        params.extend(FICHA_TYPES)
    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)
    else:
        query += " AND status != 'cancelled'"
    if year_filter:
        query += " AND strftime('%Y', date) = ?"
        params.append(year_filter)
    query += " ORDER BY date ASC"

    events = conn.execute(query, params).fetchall()
    years = conn.execute("SELECT DISTINCT strftime('%Y', date) as y FROM events WHERE date IS NOT NULL ORDER BY y DESC").fetchall()

    # For overnight trips, count rows with missing identity fields so we can flag them in the list
    trip_missing = {}
    for e in events:
        if not (e["end_date"] and e["end_date"] != e["date"]):
            continue
        a = conn.execute("""
            SELECT COUNT(*) FROM event_students estu
            JOIN students s ON estu.student_id = s.id
            WHERE estu.event_id = ?
              AND (s.birth_date IS NULL OR s.child_rg IS NULL OR s.child_cpf IS NULL OR s.guardian1_phone IS NULL)
        """, (e["id"],)).fetchone()[0]
        sm = conn.execute("""
            SELECT COUNT(*) FROM event_staff es
            LEFT JOIN teachers t ON es.teacher_id = t.id
            WHERE es.event_id = ?
              AND (es.teacher_id IS NULL
                   OR t.birth_date IS NULL OR t.cpf IS NULL OR t.rg IS NULL OR t.phone IS NULL)
        """, (e["id"],)).fetchone()[0]
        trip_missing[e["id"]] = a + sm

    conn.close()

    return render_template_string(PAGE_OPEN.replace("{{ title }}", "Lista de Eventos") + LISTA_BODY + PAGE_CLOSE,
        page='lista', events=events, years=years, trip_missing=trip_missing,
        type_filter=type_filter, status_filter=status_filter, year_filter=year_filter,
        format_date=format_date, TYPE_LABELS=TYPE_LABELS, STATUS_LABELS=STATUS_LABELS,
        EVENT_TYPES=EVENT_TYPES, EVENT_STATUSES=EVENT_STATUSES, FICHA_TYPES=FICHA_TYPES)


@app.route("/novo", methods=["GET", "POST"])
def novo_evento():
    if request.method == "POST":
        conn = get_db()
        banda = request.form.get("banda") or None
        conn.execute("""
            INSERT INTO events (name, date, end_date, time, type, status, location, description, is_public, banda)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            request.form["name"],
            request.form.get("date") or None,
            request.form.get("end_date") or None,
            request.form.get("time") or None,
            request.form.get("type") or "show",
            request.form.get("status") or "planned",
            request.form.get("location") or None,
            request.form.get("description") or None,
            1 if request.form.get("is_public") else 0,
            banda,
        ))
        conn.commit()
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        _auto_populate_students(conn, new_id, banda)
        conn.close()
        return redirect(PREFIX + f"/{new_id}")

    conn = get_db()
    bandas = list(BANDA_GROUP_MAP.keys())
    conn.close()
    return render_template_string(PAGE_OPEN.replace("{{ title }}", "Novo Evento") + NOVO_EVENTO_BODY + PAGE_CLOSE,
        page='novo', event=None, bandas=bandas,
        EVENT_TYPES=EVENT_TYPES, EVENT_STATUSES=EVENT_STATUSES, STATUS_LABELS=STATUS_LABELS, TYPE_LABELS=TYPE_LABELS)


@app.route("/<int:event_id>")
def ficha(event_id):
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        conn.close()
        return redirect(PREFIX + "/lista")

    cronograma = conn.execute(
        "SELECT * FROM event_schedule WHERE event_id = ? ORDER BY sort_order, date, time_start",
        (event_id,)
    ).fetchall()

    equipe = conn.execute("""
        SELECT es.*,
               t.name as teacher_name,
               t.birth_date as t_birth_date,
               t.gender as t_gender,
               t.cpf as t_cpf,
               t.rg as t_rg,
               t.phone as t_phone
        FROM event_staff es
        LEFT JOIN teachers t ON es.teacher_id = t.id
        WHERE es.event_id = ?
        ORDER BY es.id
    """, (event_id,)).fetchall()

    participantes = conn.execute("""
        SELECT estu.*,
               s.name, s.gender, s.birth_date,
               s.child_rg, s.child_cpf,
               s.guardian1_name, s.guardian1_phone,
               s.allergies, s.medical_condition, s.special_needs,
               g.name as group_name
        FROM event_students estu
        JOIN students s ON estu.student_id = s.id
        LEFT JOIN groups g ON estu.group_id = g.id
        WHERE estu.event_id = ?
        ORDER BY estu.room IS NULL, estu.room, g.name, s.name
    """, (event_id,)).fetchall()

    logistica = conn.execute(
        "SELECT * FROM event_logistics WHERE event_id = ? ORDER BY category, sort_order",
        (event_id,)
    ).fetchall()

    orcamento = conn.execute(
        "SELECT * FROM event_budget WHERE event_id = ? ORDER BY is_income DESC, category, sort_order",
        (event_id,)
    ).fetchall()

    contatos = conn.execute(
        "SELECT * FROM event_contacts WHERE event_id = ? ORDER BY category, sort_order",
        (event_id,)
    ).fetchall()

    tarefas = conn.execute(
        "SELECT * FROM event_tasks WHERE event_id = ? ORDER BY CASE category WHEN 'pre' THEN 1 WHEN 'dia' THEN 2 WHEN 'pos' THEN 3 ELSE 4 END, sort_order, id",
        (event_id,)
    ).fetchall()

    teachers = conn.execute("SELECT id, name FROM teachers ORDER BY name").fetchall()
    groups = conn.execute("SELECT id, name FROM groups ORDER BY name").fetchall()
    students_all = conn.execute("SELECT id, name FROM students WHERE status = 'Ativo' ORDER BY name").fetchall()

    # Budget totals
    receitas = sum(r["estimated_amount"] or 0 for r in orcamento if r["is_income"])
    despesas_rouanet = sum(r["estimated_amount"] or 0 for r in orcamento if not r["is_income"] and r["funding_source"] == "rouanet")
    despesas_assoc = sum(r["estimated_amount"] or 0 for r in orcamento if not r["is_income"] and r["funding_source"] != "rouanet")
    despesas_total = despesas_rouanet + despesas_assoc
    saldo = receitas - despesas_total

    total_tarefas = len(tarefas)
    tarefas_done = sum(1 for t in tarefas if t["status"] == "concluido")

    # JSON for cronograma JS (agile timeline)
    cronograma_json = json.dumps([_crono_row_to_dict(r) for r in cronograma], ensure_ascii=False)

    # WhatsApp summary text
    whatsapp_text = _build_whatsapp_summary(event, cronograma, participantes, equipe)

    conn.close()

    is_trip = is_overnight_trip(event)
    missing_alunos = sum(1 for p in participantes if participant_missing_fields(p)) if is_trip else 0
    missing_equipe = sum(1 for s in equipe if staff_missing_fields(s)) if is_trip else 0

    transport_counts = {
        "alunos_fb":      sum(1 for p in participantes if p["transport"] == "fb"),
        "alunos_proprio": sum(1 for p in participantes if p["transport"] == "proprio"),
        "alunos_unset":   sum(1 for p in participantes if not p["transport"]),
        "equipe_fb":      sum(1 for s in equipe if s["transport"] == "fb"),
        "equipe_proprio": sum(1 for s in equipe if s["transport"] == "proprio"),
        "equipe_unset":   sum(1 for s in equipe if not s["transport"]),
    }

    return render_template_string(PAGE_OPEN.replace("{{ title }}", event["name"]) + FICHA_BODY + PAGE_CLOSE,
        page='ficha', event=event,
        cronograma=cronograma, cronograma_json=cronograma_json,
        equipe=equipe, participantes=participantes,
        logistica=logistica, orcamento=orcamento, contatos=contatos, tarefas=tarefas,
        teachers=teachers, groups=groups, students_all=students_all,
        receitas=receitas, despesas_rouanet=despesas_rouanet, despesas_assoc=despesas_assoc,
        despesas_total=despesas_total, saldo=saldo,
        total_tarefas=total_tarefas, tarefas_done=tarefas_done,
        whatsapp_text=whatsapp_text,
        is_trip=is_trip, missing_alunos=missing_alunos, missing_equipe=missing_equipe,
        transport_counts=transport_counts,
        participant_missing_fields=participant_missing_fields,
        staff_missing_fields=staff_missing_fields,
        format_date=format_date, format_brl=format_brl, format_phone=format_phone,
        TYPE_LABELS=TYPE_LABELS, STATUS_LABELS=STATUS_LABELS,
        LOGISTICS_CATEGORIES=LOGISTICS_CATEGORIES, CONTACT_CATEGORIES=CONTACT_CATEGORIES,
        BUDGET_CATEGORIES=BUDGET_CATEGORIES,
        CHECKLIST_PHASES=CHECKLIST_PHASES, CHECKLIST_TEMPLATES=CHECKLIST_TEMPLATES)


@app.route("/<int:event_id>/licoes")
def licoes(event_id):
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    conn.close()
    if not event or not event["lessons_learned"]:
        return redirect(PREFIX + "/lista")
    content = event["lessons_learned"]
    import re
    # Simple markdown → HTML
    lines = content.split('\n')
    html_lines = []
    in_list = False
    for line in lines:
        stripped = line.strip()
        if not stripped:
            if in_list:
                html_lines.append('</ul>')
                in_list = False
            html_lines.append('<br>')
            continue
        if stripped == '---':
            if in_list:
                html_lines.append('</ul>')
                in_list = False
            html_lines.append('<hr style="border: none; border-top: 1px solid var(--border); margin: 24px 0;">')
            continue
        if stripped.startswith('### '):
            if in_list:
                html_lines.append('</ul>')
                in_list = False
            html_lines.append(f'<h3 style="color: var(--yellow); font-size: 16px; margin: 20px 0 12px;">{stripped[4:]}</h3>')
            continue
        if stripped.startswith('## '):
            if in_list:
                html_lines.append('</ul>')
                in_list = False
            html_lines.append(f'<h2 style="color: var(--yellow); font-size: 20px; margin: 0 0 16px;">{stripped[3:]}</h2>')
            continue
        if stripped.startswith('- '):
            if not in_list:
                html_lines.append('<ul style="margin: 8px 0; padding-left: 20px;">')
                in_list = True
            item = stripped[2:]
            item = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', item)
            html_lines.append(f'<li style="margin-bottom: 6px; color: rgba(255,255,255,0.85);">{item}</li>')
            continue
        if in_list:
            html_lines.append('</ul>')
            in_list = False
        stripped = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', stripped)
        html_lines.append(f'<p style="margin-bottom: 8px; color: rgba(255,255,255,0.85);">{stripped}</p>')
    if in_list:
        html_lines.append('</ul>')
    body_html = '\n'.join(html_lines)

    return render_template_string(PAGE_OPEN.replace("{{ title }}", "Lições — " + event["name"]) + """
<div class="container">
    <div class="page-header">
        <div>
            <div style="font-size: 12px; color: var(--text-muted); margin-bottom: 4px;">
                {{ TYPE_LABELS.get(event['type'], event['type'] or '') }}
                · {{ format_date(event['date']) }}
            </div>
            <h1>Lições Aprendidas</h1>
        </div>
        <div style="display: flex; gap: 8px;">
            <a href="{{ P }}/{{ event['id'] }}" class="btn btn-secondary btn-sm">← Voltar à ficha</a>
        </div>
    </div>
    <div class="section">
        {{ body_html | safe }}
    </div>
</div>
""" + PAGE_CLOSE, page='licoes', event=event, body_html=body_html,
        format_date=format_date, TYPE_LABELS=TYPE_LABELS)


@app.route("/<int:event_id>/editar", methods=["GET", "POST"])
def editar_evento(event_id):
    conn = get_db()
    if request.method == "POST":
        banda = request.form.get("banda") or None
        conn.execute("""
            UPDATE events SET name=?, date=?, end_date=?, time=?, type=?, status=?,
            location=?, description=?, is_public=?, banda=?
            WHERE id=?
        """, (
            request.form["name"],
            request.form.get("date") or None,
            request.form.get("end_date") or None,
            request.form.get("time") or None,
            request.form.get("type") or "show",
            request.form.get("status") or "planned",
            request.form.get("location") or None,
            request.form.get("description") or None,
            1 if request.form.get("is_public") else 0,
            banda,
            event_id,
        ))
        conn.commit()
        _auto_populate_students(conn, event_id, banda)
        conn.close()
        return redirect(PREFIX + f"/{event_id}")

    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    conn.close()
    if not event:
        return redirect(PREFIX + "/lista")

    bandas = list(BANDA_GROUP_MAP.keys())
    return render_template_string(PAGE_OPEN.replace("{{ title }}", "Editar Ficha") + NOVO_EVENTO_BODY + PAGE_CLOSE,
        page='ficha', event=event, bandas=bandas,
        EVENT_TYPES=EVENT_TYPES, EVENT_STATUSES=EVENT_STATUSES, STATUS_LABELS=STATUS_LABELS, TYPE_LABELS=TYPE_LABELS)


@app.route("/<int:event_id>/excluir", methods=["POST"])
def excluir_evento(event_id):
    conn = get_db()
    conn.execute("UPDATE events SET status='cancelled' WHERE id=?", (event_id,))
    conn.commit()
    conn.close()
    return redirect(PREFIX + "/lista")


# ── Cronograma ─────────────────────────────────────────────────────────────

@app.route("/<int:event_id>/cronograma/adicionar", methods=["POST"])
def cronograma_add(event_id):
    conn = get_db()
    conn.execute("""
        INSERT INTO event_schedule (event_id, date, time_start, time_end, activity, location, responsible, notes, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        event_id,
        request.form.get("date") or None,
        request.form.get("time_start") or None,
        request.form.get("time_end") or None,
        request.form.get("activity") or "",
        request.form.get("location") or None,
        request.form.get("responsible") or None,
        request.form.get("notes") or None,
        int(request.form.get("sort_order") or 0),
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#cronograma")


@app.route("/<int:event_id>/cronograma/<int:item_id>/editar", methods=["POST"])
def cronograma_edit(event_id, item_id):
    conn = get_db()
    conn.execute("""
        UPDATE event_schedule SET date=?, time_start=?, time_end=?, activity=?, location=?, responsible=?, notes=?
        WHERE id=? AND event_id=?
    """, (
        request.form.get("date") or None,
        request.form.get("time_start") or None,
        request.form.get("time_end") or None,
        request.form.get("activity") or "",
        request.form.get("location") or None,
        request.form.get("responsible") or None,
        request.form.get("notes") or None,
        item_id, event_id,
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#cronograma")


@app.route("/<int:event_id>/cronograma/<int:item_id>/excluir", methods=["POST"])
def cronograma_delete(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_schedule WHERE id=? AND event_id=?", (item_id, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#cronograma")


# ── Cronograma JSON API ──────────────────────────────────────────────────

DIAS_SEMANA = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira', 'Sábado', 'Domingo']
MESES = ['', 'janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho', 'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']

def _disambiguate_first_names(full_names):
    """Short display names: first name when unique, else 'First X.' using second name.

    Uses second name (not surname) because Brazilian surnames are often shared
    (Silva, Santos, Oliveira) while second given names are distinctive.
    Falls back to full second name if initials still collide.
    """
    from collections import Counter
    firsts = [n.split()[0] for n in full_names]
    first_counts = Counter(firsts)

    result = []
    for full in full_names:
        parts = full.split()
        first = parts[0]
        if first_counts[first] == 1 or len(parts) < 2:
            result.append(first)
        else:
            result.append(f"{first} {parts[1][0]}.")

    # Resolve remaining collisions: expand to full second name
    result_counts = Counter(result)
    for i, (full, r) in enumerate(zip(full_names, result)):
        if result_counts[r] > 1:
            parts = full.split()
            if len(parts) >= 2:
                result[i] = f"{parts[0]} {parts[1]}"

    return result


TYPE_EMOJI = {'show': '🎺', 'festival': '🎪', 'viagem': '✈️', 'reuniao': '📋', 'workshop': '🎓', 'ensaio': '🎶'}
TYPE_WHATSAPP_LABEL = {'show': 'SHOW', 'festival': 'FESTIVAL', 'viagem': 'VIAGEM', 'workshop': 'WORKSHOP', 'ensaio': 'ENSAIO'}


def _build_whatsapp_summary(event, cronograma, participantes, equipe=None):
    """Build a WhatsApp-formatted summary for copy-to-clipboard."""
    lines = []

    # Header
    emoji = TYPE_EMOJI.get(event['type'], '📌')
    label = TYPE_WHATSAPP_LABEL.get(event['type'], event['type'] or 'EVENTO')
    lines.append(f"*{emoji} {label} — {event['name']}*")

    # Date line
    if event['date']:
        try:
            from datetime import date as dt_date
            d = dt_date.fromisoformat(str(event['date']))
            dia_semana = DIAS_SEMANA[d.weekday()]
            lines.append(f"*{dia_semana}, {d.day} de {MESES[d.month]}*")
        except (ValueError, IndexError):
            lines.append(f"*{format_date(event['date'])}*")

    lines.append("")

    # Location
    if event['location'] and event['location'] != 'None':
        lines.append(f"📍 {event['location']}")

    # Banda
    if event['banda']:
        lines.append(f"🎵 {event['banda']}")

    if event['location'] or event['banda']:
        lines.append("")

    # Cronograma
    if cronograma:
        lines.append("*Cronograma:*")
        for item in cronograma:
            t = item['time_start'] or ''
            if item['time_end']:
                t += f"–{item['time_end']}"
            activity = item['activity'] or ''
            lines.append(f"{t} — {activity}")
        lines.append("")

    # Equipe (staff with roles)
    if equipe:
        lines.append(f"*Equipe ({len(equipe)}):*")
        for s in equipe:
            nome = s['name'] or s['teacher_name'] or '—'
            role = s['role']
            if role:
                lines.append(f"{nome} ({role})")
            else:
                lines.append(nome)
        lines.append("")

    # Participantes (full names, one per line)
    if participantes:
        lines.append(f"*Alunos ({len(participantes)}):*")
        for p in participantes:
            lines.append(p['name'])
        lines.append("")

    lines.append("Qualquer dúvida, fala comigo! 👍")

    return "\n".join(lines)


def _crono_row_to_dict(row):
    """Convert a sqlite Row to a plain dict for JSON serialization."""
    return {
        "id": row["id"],
        "event_id": row["event_id"],
        "date": row["date"],
        "time_start": row["time_start"],
        "time_end": row["time_end"],
        "activity": row["activity"],
        "location": row["location"],
        "responsible": row["responsible"],
        "notes": row["notes"],
        "sort_order": row["sort_order"],
    }


@app.route("/api/check-duplicates")
def api_check_duplicates():
    """Check for existing events on a given date."""
    dt = request.args.get("date")
    exclude = request.args.get("exclude")  # event ID to exclude (for edit mode)
    if not dt:
        return jsonify([])
    conn = get_db()
    query = "SELECT id, name, banda, type FROM events WHERE date = ? AND status != 'cancelled'"
    params = [dt]
    if exclude:
        query += " AND id != ?"
        params.append(int(exclude))
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/<int:event_id>/api/cronograma", methods=["GET"])
def api_cronograma_list(event_id):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM event_schedule WHERE event_id = ? ORDER BY sort_order, time_start, id",
        (event_id,)
    ).fetchall()
    conn.close()
    return jsonify([_crono_row_to_dict(r) for r in rows])


@app.route("/<int:event_id>/api/cronograma", methods=["POST"])
def api_cronograma_create(event_id):
    data = request.get_json(force=True)
    conn = get_db()
    # Get next sort_order
    max_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order), -1) FROM event_schedule WHERE event_id = ?",
        (event_id,)
    ).fetchone()[0]
    conn.execute("""
        INSERT INTO event_schedule (event_id, date, time_start, time_end, activity, location, responsible, notes, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        event_id,
        data.get("date") or None,
        data.get("time_start") or None,
        data.get("time_end") or None,
        data.get("activity") or "",
        data.get("location") or None,
        data.get("responsible") or None,
        data.get("notes") or None,
        max_order + 1,
    ))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = conn.execute("SELECT * FROM event_schedule WHERE id = ?", (new_id,)).fetchone()
    conn.close()
    return jsonify(_crono_row_to_dict(row)), 201


@app.route("/<int:event_id>/api/cronograma/<int:item_id>", methods=["PUT"])
def api_cronograma_update(event_id, item_id):
    data = request.get_json(force=True)
    conn = get_db()
    # Build dynamic UPDATE from provided fields
    allowed = {"date", "time_start", "time_end", "activity", "location", "responsible", "notes"}
    sets = []
    vals = []
    for key in allowed:
        if key in data:
            sets.append(f"{key} = ?")
            vals.append(data[key] or None)
    if not sets:
        conn.close()
        return jsonify({"error": "No fields to update"}), 400
    vals.extend([item_id, event_id])
    conn.execute(f"UPDATE event_schedule SET {', '.join(sets)} WHERE id = ? AND event_id = ?", vals)
    conn.commit()
    row = conn.execute("SELECT * FROM event_schedule WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify(_crono_row_to_dict(row))


@app.route("/<int:event_id>/api/cronograma/<int:item_id>", methods=["DELETE"])
def api_cronograma_remove(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_schedule WHERE id = ? AND event_id = ?", (item_id, event_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/<int:event_id>/api/cronograma/reorder", methods=["POST"])
def api_cronograma_reorder(event_id):
    data = request.get_json(force=True)
    item_id = data.get("item_id")
    direction = data.get("direction")  # "up" or "down"
    if not item_id or direction not in ("up", "down"):
        return jsonify({"error": "Need item_id and direction (up/down)"}), 400

    conn = get_db()
    items = conn.execute(
        "SELECT id, sort_order FROM event_schedule WHERE event_id = ? ORDER BY sort_order, time_start, id",
        (event_id,)
    ).fetchall()

    # Normalize sort_order to sequential 0,1,2,...
    id_list = [r["id"] for r in items]
    try:
        idx = id_list.index(item_id)
    except ValueError:
        conn.close()
        return jsonify({"error": "Item not found"}), 404

    swap_idx = idx - 1 if direction == "up" else idx + 1
    if swap_idx < 0 or swap_idx >= len(id_list):
        # Already at boundary — return current order
        rows = conn.execute(
            "SELECT * FROM event_schedule WHERE event_id = ? ORDER BY sort_order, time_start, id",
            (event_id,)
        ).fetchall()
        conn.close()
        return jsonify([_crono_row_to_dict(r) for r in rows])

    # Swap sort_order values
    id_a, id_b = id_list[idx], id_list[swap_idx]
    conn.execute("UPDATE event_schedule SET sort_order = ? WHERE id = ?", (swap_idx, id_a))
    conn.execute("UPDATE event_schedule SET sort_order = ? WHERE id = ?", (idx, id_b))
    conn.commit()

    rows = conn.execute(
        "SELECT * FROM event_schedule WHERE event_id = ? ORDER BY sort_order, time_start, id",
        (event_id,)
    ).fetchall()
    conn.close()
    return jsonify([_crono_row_to_dict(r) for r in rows])


# ── Orçamento ──────────────────────────────────────────────────────────────

@app.route("/<int:event_id>/orcamento/adicionar", methods=["POST"])
def orcamento_add(event_id):
    conn = get_db()
    conn.execute("""
        INSERT INTO event_budget (event_id, category, item, quantity, unit_cost, estimated_amount, actual_amount, is_income, funding_source, paid, paid_date, notes, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        event_id,
        request.form.get("category") or "Outro",
        request.form.get("item") or "",
        float(request.form["quantity"]) if request.form.get("quantity") else None,
        float(request.form["unit_cost"]) if request.form.get("unit_cost") else None,
        float(request.form["estimated_amount"]) if request.form.get("estimated_amount") else None,
        float(request.form["actual_amount"]) if request.form.get("actual_amount") else None,
        1 if request.form.get("is_income") else 0,
        request.form.get("funding_source") or "associacao",
        1 if request.form.get("paid") else 0,
        request.form.get("paid_date") or None,
        request.form.get("notes") or None,
        int(request.form.get("sort_order") or 0),
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#orcamento")


@app.route("/<int:event_id>/orcamento/<int:item_id>/editar", methods=["POST"])
def orcamento_edit(event_id, item_id):
    conn = get_db()
    conn.execute("""
        UPDATE event_budget SET category=?, item=?, quantity=?, unit_cost=?, estimated_amount=?,
        actual_amount=?, is_income=?, funding_source=?, paid=?, paid_date=?, notes=?
        WHERE id=? AND event_id=?
    """, (
        request.form.get("category") or "Outro",
        request.form.get("item") or "",
        float(request.form["quantity"]) if request.form.get("quantity") else None,
        float(request.form["unit_cost"]) if request.form.get("unit_cost") else None,
        float(request.form["estimated_amount"]) if request.form.get("estimated_amount") else None,
        float(request.form["actual_amount"]) if request.form.get("actual_amount") else None,
        1 if request.form.get("is_income") else 0,
        request.form.get("funding_source") or "associacao",
        1 if request.form.get("paid") else 0,
        request.form.get("paid_date") or None,
        request.form.get("notes") or None,
        item_id, event_id,
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#orcamento")


@app.route("/<int:event_id>/orcamento/<int:item_id>/excluir", methods=["POST"])
def orcamento_delete(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_budget WHERE id=? AND event_id=?", (item_id, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#orcamento")


# ── Logística ──────────────────────────────────────────────────────────────

@app.route("/<int:event_id>/logistica/adicionar", methods=["POST"])
def logistica_add(event_id):
    conn = get_db()
    conn.execute("""
        INSERT INTO event_logistics (event_id, category, item, details, status, responsible, deadline, notes, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        event_id,
        request.form.get("category") or "geral",
        request.form.get("item") or "",
        request.form.get("details") or None,
        request.form.get("status") or "pendente",
        request.form.get("responsible") or None,
        request.form.get("deadline") or None,
        request.form.get("notes") or None,
        int(request.form.get("sort_order") or 0),
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#logistica")


@app.route("/<int:event_id>/logistica/<int:item_id>/editar", methods=["POST"])
def logistica_edit(event_id, item_id):
    conn = get_db()
    conn.execute("""
        UPDATE event_logistics SET category=?, item=?, details=?, status=?, responsible=?, deadline=?, notes=?
        WHERE id=? AND event_id=?
    """, (
        request.form.get("category") or "geral",
        request.form.get("item") or "",
        request.form.get("details") or None,
        request.form.get("status") or "pendente",
        request.form.get("responsible") or None,
        request.form.get("deadline") or None,
        request.form.get("notes") or None,
        item_id, event_id,
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#logistica")


@app.route("/<int:event_id>/logistica/<int:item_id>/excluir", methods=["POST"])
def logistica_delete(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_logistics WHERE id=? AND event_id=?", (item_id, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#logistica")


# ── Contatos ───────────────────────────────────────────────────────────────

@app.route("/<int:event_id>/contatos/adicionar", methods=["POST"])
def contatos_add(event_id):
    conn = get_db()
    conn.execute("""
        INSERT INTO event_contacts (event_id, category, name, role, phone, email, notes, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        event_id,
        request.form.get("category") or "geral",
        request.form.get("name") or "",
        request.form.get("role") or None,
        request.form.get("phone") or None,
        request.form.get("email") or None,
        request.form.get("notes") or None,
        int(request.form.get("sort_order") or 0),
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#contatos")


@app.route("/<int:event_id>/contatos/<int:item_id>/editar", methods=["POST"])
def contatos_edit(event_id, item_id):
    conn = get_db()
    conn.execute("""
        UPDATE event_contacts SET category=?, name=?, role=?, phone=?, email=?, notes=?
        WHERE id=? AND event_id=?
    """, (
        request.form.get("category") or "geral",
        request.form.get("name") or "",
        request.form.get("role") or None,
        request.form.get("phone") or None,
        request.form.get("email") or None,
        request.form.get("notes") or None,
        item_id, event_id,
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#contatos")


@app.route("/<int:event_id>/contatos/<int:item_id>/excluir", methods=["POST"])
def contatos_delete(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_contacts WHERE id=? AND event_id=?", (item_id, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#contatos")


# ── Tarefas ────────────────────────────────────────────────────────────────

@app.route("/<int:event_id>/tarefas/adicionar", methods=["POST"])
def tarefas_add(event_id):
    conn = get_db()
    conn.execute("""
        INSERT INTO event_tasks (event_id, task, category, deadline, responsible, status, priority, notes, sort_order)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        event_id,
        request.form.get("task") or "",
        request.form.get("category") or "geral",
        request.form.get("deadline") or None,
        request.form.get("responsible") or None,
        "pendente",
        int(request.form.get("priority") or 0),
        request.form.get("notes") or None,
        int(request.form.get("sort_order") or 0),
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#checklist")


@app.route("/<int:event_id>/tarefas/<int:item_id>/editar", methods=["POST"])
def tarefas_edit(event_id, item_id):
    conn = get_db()
    conn.execute("""
        UPDATE event_tasks SET task=?, category=?, deadline=?, responsible=?, priority=?, notes=?
        WHERE id=? AND event_id=?
    """, (
        request.form.get("task") or "",
        request.form.get("category") or "geral",
        request.form.get("deadline") or None,
        request.form.get("responsible") or None,
        int(request.form.get("priority") or 0),
        request.form.get("notes") or None,
        item_id, event_id,
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#checklist")


@app.route("/<int:event_id>/tarefas/<int:item_id>/concluir", methods=["POST"])
def tarefas_toggle(event_id, item_id):
    conn = get_db()
    task = conn.execute("SELECT * FROM event_tasks WHERE id=? AND event_id=?", (item_id, event_id)).fetchone()
    if task:
        if task["status"] == "concluido":
            conn.execute("UPDATE event_tasks SET status='pendente', completed_at=NULL WHERE id=?", (item_id,))
        else:
            conn.execute("UPDATE event_tasks SET status='concluido', completed_at=datetime('now') WHERE id=?", (item_id,))
        conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#checklist")


@app.route("/<int:event_id>/tarefas/<int:item_id>/excluir", methods=["POST"])
def tarefas_delete(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_tasks WHERE id=? AND event_id=?", (item_id, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#checklist")


@app.route("/<int:event_id>/tarefas/template", methods=["POST"])
def tarefas_load_template(event_id):
    conn = get_db()
    event = conn.execute("SELECT type FROM events WHERE id=?", (event_id,)).fetchone()
    if event and event["type"] in CHECKLIST_TEMPLATES:
        template = CHECKLIST_TEMPLATES[event["type"]]
        order = 0
        for phase_key, _label in CHECKLIST_PHASES:
            for task_text in template.get(phase_key, []):
                conn.execute("""
                    INSERT INTO event_tasks (event_id, task, category, status, sort_order)
                    VALUES (?, ?, ?, 'pendente', ?)
                """, (event_id, task_text, phase_key, order))
                order += 1
        conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#checklist")


# ── Equipe ─────────────────────────────────────────────────────────────────

@app.route("/<int:event_id>/equipe/adicionar", methods=["POST"])
def equipe_add(event_id):
    conn = get_db()
    teacher_id = request.form.get("teacher_id") or None
    name = request.form.get("name") or ""
    if teacher_id:
        t = conn.execute("SELECT name FROM teachers WHERE id=?", (int(teacher_id),)).fetchone()
        if t and not name:
            name = t["name"]
    conn.execute("""
        INSERT INTO event_staff (event_id, teacher_id, name, role, hours, fee, confirmed, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        event_id,
        int(teacher_id) if teacher_id else None,
        name,
        request.form.get("role") or None,
        float(request.form["hours"]) if request.form.get("hours") else None,
        float(request.form["fee"]) if request.form.get("fee") else None,
        1 if request.form.get("confirmed") else 0,
        request.form.get("notes") or None,
    ))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#equipe")


@app.route("/<int:event_id>/equipe/<int:item_id>/excluir", methods=["POST"])
def equipe_delete(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_staff WHERE id=? AND event_id=?", (item_id, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#equipe")


@app.route("/<int:event_id>/equipe/<int:item_id>/toggle", methods=["POST"])
def equipe_toggle(event_id, item_id):
    conn = get_db()
    conn.execute(
        "UPDATE event_staff SET confirmed = 1 - COALESCE(confirmed, 0) WHERE id=? AND event_id=?",
        (item_id, event_id),
    )
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#equipe")


# ── Participantes ──────────────────────────────────────────────────────────

BANDA_GROUP_MAP = {
    'Banda Preta':   ('Banda Preta', None),
    'Preta A':       ('Banda Preta', 'A'),
    'Banda Preta 1': ('Banda Preta', 'A'),
    'Preta B':       ('Banda Preta', 'B'),
    'Banda Preta 2': ('Banda Preta', 'B'),
    'Banda Amarela':    ('Banda Amarela', None),
    'Amarela':          ('Banda Amarela', None),
    'Banda Roxa':       ('Banda Roxa', None),
    'Banda Verde':      ('Banda Verde', None),
    'Grupo Semente':    ('Grupo Semente', None),
    'Iniciação Musical': ('Iniciação Musical', None),
}


def _auto_populate_students(conn, event_id, banda):
    """Auto-fill event_students from group_assignments if list is empty."""
    if not banda:
        return
    mapping = BANDA_GROUP_MAP.get(banda)
    if not mapping:
        return
    group_id, sub_group = mapping

    # Only populate if no students yet
    existing = conn.execute(
        "SELECT COUNT(*) as n FROM event_students WHERE event_id = ?", (event_id,)
    ).fetchone()["n"]
    if existing > 0:
        return

    sql = """
        SELECT s.id FROM students s
        JOIN group_assignments ga ON ga.student_id = s.id
        WHERE ga.group_id = ? AND s.status = 'Ativo'
    """
    params = [group_id]
    if sub_group:
        sql += " AND ga.sub_group = ?"
        params.append(sub_group)

    students = conn.execute(sql, params).fetchall()
    for s in students:
        conn.execute(
            "INSERT INTO event_students (event_id, student_id, group_id, confirmed) VALUES (?, ?, ?, 0)",
            (event_id, s["id"], group_id)
        )
    conn.commit()


@app.route("/<int:event_id>/alunos/adicionar", methods=["POST"])
def alunos_add(event_id):
    conn = get_db()
    raw = request.form.get("group_id")
    if raw:
        # Handle "Banda Preta:A" format for sub-groups
        if ':' in raw:
            group_id, sub_group = raw.split(':', 1)
            sql = """
                SELECT s.id FROM students s
                JOIN group_assignments ga ON ga.student_id = s.id
                WHERE ga.group_id = ? AND ga.sub_group = ? AND s.status = 'Ativo'
            """
            students = conn.execute(sql, (group_id, sub_group)).fetchall()
        else:
            group_id = raw
            students = conn.execute("""
                SELECT s.id FROM students s
                JOIN group_assignments ga ON ga.student_id = s.id
                WHERE ga.group_id = ? AND s.status = 'Ativo'
            """, (group_id,)).fetchall()
        for s in students:
            existing = conn.execute(
                "SELECT id FROM event_students WHERE event_id=? AND student_id=?",
                (event_id, s["id"])
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO event_students (event_id, student_id, group_id, confirmed) VALUES (?, ?, ?, 0)",
                    (event_id, s["id"], group_id)
                )
        conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#participantes")


@app.route("/<int:event_id>/alunos/adicionar-individual", methods=["POST"])
def alunos_add_individual(event_id):
    conn = get_db()
    student_id = request.form.get("student_id")
    if student_id:
        existing = conn.execute(
            "SELECT id FROM event_students WHERE event_id=? AND student_id=?",
            (event_id, int(student_id))
        ).fetchone()
        if not existing:
            # Look up student's group if any
            ga = conn.execute(
                "SELECT group_id FROM group_assignments WHERE student_id = ?",
                (int(student_id),)
            ).fetchone()
            group_id = ga["group_id"] if ga else None
            conn.execute(
                "INSERT INTO event_students (event_id, student_id, group_id, confirmed) VALUES (?, ?, ?, 0)",
                (event_id, int(student_id), group_id)
            )
            conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#participantes")


@app.route("/<int:event_id>/alunos/<int:item_id>/excluir", methods=["POST"])
def alunos_delete(event_id, item_id):
    conn = get_db()
    conn.execute("DELETE FROM event_students WHERE id=? AND event_id=?", (item_id, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#participantes")


@app.route("/<int:event_id>/alunos/<int:item_id>/quarto", methods=["POST"])
def alunos_set_room(event_id, item_id):
    room = (request.form.get("room") or "").strip() or None
    conn = get_db()
    conn.execute("UPDATE event_students SET room=? WHERE id=? AND event_id=?", (room, item_id, event_id))
    conn.commit()
    conn.close()
    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "room": room})
    return redirect(PREFIX + f"/{event_id}#participantes")


@app.route("/<int:event_id>/equipe/<int:item_id>/quarto", methods=["POST"])
def equipe_set_room(event_id, item_id):
    room = (request.form.get("room") or "").strip() or None
    conn = get_db()
    conn.execute("UPDATE event_staff SET room=? WHERE id=? AND event_id=?", (room, item_id, event_id))
    conn.commit()
    conn.close()
    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "room": room})
    return redirect(PREFIX + f"/{event_id}#equipe")


@app.route("/<int:event_id>/alunos/<int:item_id>/notas", methods=["POST"])
def alunos_set_notes(event_id, item_id):
    notes = (request.form.get("notes") or "").strip() or None
    conn = get_db()
    conn.execute("UPDATE event_students SET notes=? WHERE id=? AND event_id=?", (notes, item_id, event_id))
    conn.commit()
    conn.close()
    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "notes": notes})
    return redirect(PREFIX + f"/{event_id}#participantes")


@app.route("/<int:event_id>/equipe/<int:item_id>/notas", methods=["POST"])
def equipe_set_notes(event_id, item_id):
    notes = (request.form.get("notes") or "").strip() or None
    conn = get_db()
    conn.execute("UPDATE event_staff SET notes=? WHERE id=? AND event_id=?", (notes, item_id, event_id))
    conn.commit()
    conn.close()
    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "notes": notes})
    return redirect(PREFIX + f"/{event_id}#equipe")


@app.route("/<int:event_id>/alunos/<int:item_id>/transporte", methods=["POST"])
def alunos_set_transport(event_id, item_id):
    raw = (request.form.get("transport") or "").strip()
    transport = raw if raw in ("fb", "proprio") else None
    conn = get_db()
    conn.execute("UPDATE event_students SET transport=? WHERE id=? AND event_id=?", (transport, item_id, event_id))
    conn.commit()
    conn.close()
    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "transport": transport})
    return redirect(PREFIX + f"/{event_id}#participantes")


@app.route("/<int:event_id>/equipe/<int:item_id>/transporte", methods=["POST"])
def equipe_set_transport(event_id, item_id):
    raw = (request.form.get("transport") or "").strip()
    transport = raw if raw in ("fb", "proprio") else None
    conn = get_db()
    conn.execute("UPDATE event_staff SET transport=? WHERE id=? AND event_id=?", (transport, item_id, event_id))
    conn.commit()
    conn.close()
    if request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "transport": transport})
    return redirect(PREFIX + f"/{event_id}#equipe")


@app.route("/<int:event_id>/alunos/transporte-bulk", methods=["POST"])
def alunos_set_transport_bulk(event_id):
    """Set transport on all rostered alunos in one go (e.g. 'all FB' default for shows)."""
    raw = (request.form.get("transport") or "").strip()
    transport = raw if raw in ("fb", "proprio") else None
    only_unset = request.form.get("only_unset") == "1"
    conn = get_db()
    if only_unset:
        conn.execute("UPDATE event_students SET transport=? WHERE event_id=? AND transport IS NULL", (transport, event_id))
    else:
        conn.execute("UPDATE event_students SET transport=? WHERE event_id=?", (transport, event_id))
    conn.commit()
    conn.close()
    return redirect(PREFIX + f"/{event_id}#participantes")


# ── Trip manifest exports ───────────────────────────────────────────────────

def _manifest_rows(conn, event_id):
    """Combined alunos + equipe rows with identity fields, for trip exports."""
    alunos = conn.execute("""
        SELECT s.name, s.gender, s.birth_date, s.child_rg as rg, s.child_cpf as cpf,
               s.guardian1_phone as phone, s.guardian1_name as guardian,
               estu.room, 'Aluno' as kind
        FROM event_students estu JOIN students s ON estu.student_id = s.id
        WHERE estu.event_id = ?
        ORDER BY estu.room IS NULL, estu.room, s.name
    """, (event_id,)).fetchall()
    equipe = conn.execute("""
        SELECT COALESCE(t.name, es.name) as name,
               t.gender, t.birth_date, t.rg, t.cpf, t.phone,
               NULL as guardian, es.room, COALESCE(es.role, t.role, 'Equipe') as kind
        FROM event_staff es LEFT JOIN teachers t ON es.teacher_id = t.id
        WHERE es.event_id = ?
        ORDER BY es.room IS NULL, es.room, name
    """, (event_id,)).fetchall()
    return list(alunos), list(equipe)


@app.route("/<int:event_id>/manifesto.csv")
def manifesto_csv(event_id):
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        conn.close()
        return redirect(PREFIX + "/lista")
    alunos, equipe = _manifest_rows(conn, event_id)
    conn.close()

    out = io.StringIO()
    out.write("﻿")  # UTF-8 BOM for Excel compat
    w = csv.writer(out, delimiter=";")
    w.writerow([f"Manifesto — {event['name']}"])
    w.writerow([f"Data: {format_date(event['date'])}" + (f" → {format_date(event['end_date'])}" if event["end_date"] and event["end_date"] != event["date"] else "")])
    w.writerow([f"Local: {event['location'] or ''}"])
    w.writerow([])
    w.writerow(["Categoria", "Quarto", "Nome", "M/F", "Nascimento", "Idade", "RG", "CPF", "Telefone", "Responsável"])
    today = date.today()
    for r in alunos + equipe:
        bd = r["birth_date"]
        idade = ""
        if bd:
            try:
                y, m, d = [int(x) for x in str(bd).split("-")]
                age = today.year - y - ((today.month, today.day) < (m, d))
                idade = str(age)
            except Exception:
                pass
        w.writerow([
            r["kind"] or "",
            r["room"] or "",
            r["name"] or "",
            r["gender"] or "",
            format_date(bd) if bd else "",
            idade,
            r["rg"] or "",
            r["cpf"] or "",
            format_phone(r["phone"]),
            r["guardian"] or "",
        ])
    safe_name = "".join(c if c.isalnum() else "-" for c in (event["name"] or "evento"))
    return Response(
        out.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="manifesto-{safe_name}.csv"'},
    )


PRINT_BODY = """<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8">
<title>{{ title }}</title>
<style>
body { font-family: -apple-system, system-ui, sans-serif; max-width: 900px; margin: 24px auto; padding: 0 20px; color: #111; }
h1 { font-size: 22px; margin-bottom: 4px; }
.meta { color: #555; font-size: 13px; margin-bottom: 24px; }
table { width: 100%; border-collapse: collapse; margin-bottom: 24px; }
th, td { text-align: left; padding: 6px 10px; border-bottom: 1px solid #ddd; font-size: 13px; }
th { background: #f4f4f4; font-weight: 600; }
h2 { font-size: 15px; margin-top: 20px; padding: 6px 10px; background: #fffbe6; border-left: 4px solid #f5c518; }
.no-print { margin-bottom: 16px; }
@media print { .no-print { display: none; } body { margin: 0; max-width: none; } }
</style></head><body>
<div class="no-print"><button onclick="window.print()">🖨️ Imprimir / Salvar PDF</button> · <a href="{{ back }}">← Voltar à ficha</a></div>
<h1>{{ heading }}</h1>
<div class="meta">{{ event['name'] }} · {{ format_date(event['date']) }}{% if event['end_date'] and event['end_date'] != event['date'] %} → {{ format_date(event['end_date']) }}{% endif %}{% if event['location'] %} · {{ event['location'] }}{% endif %}</div>
{{ body|safe }}
</body></html>
"""


@app.route("/<int:event_id>/lista-quartos")
def lista_quartos(event_id):
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        conn.close()
        return redirect(PREFIX + "/lista")
    alunos, equipe = _manifest_rows(conn, event_id)
    conn.close()
    rows = alunos + equipe
    grouped = {}
    for r in rows:
        key = r["room"] or "— sem quarto —"
        grouped.setdefault(key, []).append(r)
    def sort_key(k):
        if k.startswith("—"):
            return (1, "")
        digits = "".join(ch for ch in k if ch.isdigit())
        return (0, int(digits) if digits else 9999)
    body_parts = []
    for room in sorted(grouped.keys(), key=sort_key):
        body_parts.append(f"<h2>{room} <span style='color:#888;font-weight:400;'>({len(grouped[room])} pessoas)</span></h2>")
        body_parts.append("<table><thead><tr><th>Nome</th><th>M/F</th><th>Categoria</th><th>Telefone</th></tr></thead><tbody>")
        for r in grouped[room]:
            body_parts.append(
                f"<tr><td>{r['name'] or ''}</td><td>{r['gender'] or ''}</td>"
                f"<td>{r['kind'] or ''}</td><td>{format_phone(r['phone'])}</td></tr>"
            )
        body_parts.append("</tbody></table>")
    return render_template_string(
        PRINT_BODY,
        title=f"Lista de Quartos — {event['name']}",
        heading="Lista de Quartos",
        event=event, format_date=format_date,
        back=f"{PREFIX}/{event_id}",
        body="".join(body_parts),
    )


@app.route("/<int:event_id>/lista-embarque")
def lista_embarque(event_id):
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        conn.close()
        return redirect(PREFIX + "/lista")
    alunos, equipe = _manifest_rows(conn, event_id)
    conn.close()
    rows = sorted(alunos + equipe, key=lambda r: (r["kind"] != "Aluno", (r["name"] or "").lower()))
    body_parts = ["<table><thead><tr><th>#</th><th>Nome</th><th>Categoria</th><th>Quarto</th><th>Telefone</th><th>Responsável</th><th>✓</th></tr></thead><tbody>"]
    for i, r in enumerate(rows, 1):
        body_parts.append(
            f"<tr><td>{i}</td><td>{r['name'] or ''}</td><td>{r['kind'] or ''}</td>"
            f"<td>{r['room'] or ''}</td><td>{format_phone(r['phone'])}</td>"
            f"<td>{r['guardian'] or ''}</td><td style='width:24px;'>☐</td></tr>"
        )
    body_parts.append("</tbody></table>")
    body_parts.append(f"<p style='color:#555;font-size:12px;'>Total: {len(rows)} pessoas</p>")
    return render_template_string(
        PRINT_BODY,
        title=f"Lista de Embarque — {event['name']}",
        heading="Lista de Embarque",
        event=event, format_date=format_date,
        back=f"{PREFIX}/{event_id}",
        body="".join(body_parts),
    )


# ── Quartos editor (drag-and-drop) ──────────────────────────────────────────

def _quartos_items(conn, event_id):
    """Return unified list of dicts for the quartos editor (alunos + equipe)."""
    today = date.today()
    def calc_age(bd):
        if not bd:
            return None
        try:
            y, m, d = [int(x) for x in str(bd).split("-")]
            return today.year - y - ((today.month, today.day) < (m, d))
        except Exception:
            return None

    items = []
    for r in conn.execute("""
        SELECT estu.id as id, s.name, s.gender, s.birth_date,
               estu.room,
               COALESCE(NULLIF(estu.group_id, ''), ga.group_id) as group_name
        FROM event_students estu
        JOIN students s ON estu.student_id = s.id
        LEFT JOIN group_assignments ga ON ga.student_id = s.id
        WHERE estu.event_id = ?
        ORDER BY s.name
    """, (event_id,)):
        items.append({
            "kind": "aluno",
            "id": r["id"],
            "name": r["name"] or "",
            "gender": (r["gender"] or "").upper(),
            "age": calc_age(r["birth_date"]),
            "label": r["group_name"] or "",
            "room": (r["room"] or "").strip() or None,
        })
    for r in conn.execute("""
        SELECT es.id as id, COALESCE(t.name, es.name) as name,
               t.gender, t.birth_date, es.room,
               COALESCE(es.role, t.role, 'Equipe') as role
        FROM event_staff es LEFT JOIN teachers t ON es.teacher_id = t.id
        WHERE es.event_id = ?
        ORDER BY name
    """, (event_id,)):
        items.append({
            "kind": "equipe",
            "id": r["id"],
            "name": r["name"] or "",
            "gender": (r["gender"] or "").upper(),
            "age": calc_age(r["birth_date"]),
            "label": r["role"] or "Equipe",
            "room": (r["room"] or "").strip() or None,
        })
    return items


QUARTOS_BODY = """
<div class="container">
  <div class="page-header">
    <div>
      <h1>🛏️ Montar Quartos</h1>
      <div style="color: var(--text-muted); font-size: 13px; margin-top: 4px;">
        {{ event['name'] }} — {{ format_date(event['date']) }}{% if event['end_date'] and event['end_date'] != event['date'] %} → {{ format_date(event['end_date']) }}{% endif %}{% if event['banda'] %} · <span style="color: var(--yellow);">{{ event['banda'] }}</span>{% endif %}
      </div>
    </div>
    <div style="display:flex; gap:8px;">
      <a class="btn btn-secondary btn-sm" href="{{ P }}/{{ event['id'] }}#participantes">← Voltar à ficha</a>
      <a class="btn btn-secondary btn-sm" href="{{ P }}/{{ event['id'] }}/lista-quartos" target="_blank">🖨️ Imprimir</a>
    </div>
  </div>

  <div class="quartos-toolbar">
    <div class="qt-stats">
      <span class="qt-stat"><strong id="qt-total">{{ total }}</strong> pessoas</span>
      <span class="qt-stat"><strong id="qt-rooms">{{ rooms|length }}</strong> quartos</span>
      <span class="qt-stat qt-warn" id="qt-unassigned-stat">
        <strong id="qt-unassigned">{{ unassigned_count }}</strong> sem quarto
      </span>
    </div>
    <div class="qt-actions">
      <input type="text" id="qt-new-room" placeholder="Nº do quarto (ex: 12)" maxlength="20" />
      <button class="btn btn-primary btn-sm" onclick="addRoom()">+ Adicionar quarto</button>
      <span class="qt-help">Arraste os cartões para mover de quarto.</span>
    </div>
  </div>

  <div id="qt-status" class="qt-status"></div>

  <div class="quartos-grid">
    <!-- Sem quarto pool -->
    <div class="quarto-card quarto-pool" data-room="">
      <div class="quarto-head">
        <h3>Sem quarto</h3>
        <span class="quarto-count" data-count></span>
      </div>
      <div class="quarto-list" data-list>
        {% for it in by_room['__none__'] %}
          {{ render_card(it)|safe }}
        {% endfor %}
      </div>
    </div>

    {% for room in rooms %}
    <div class="quarto-card" data-room="{{ room }}">
      <div class="quarto-head">
        <h3>Quarto {{ room }}</h3>
        <span class="quarto-count" data-count></span>
      </div>
      <div class="quarto-list" data-list>
        {% for it in by_room[room] %}
          {{ render_card(it)|safe }}
        {% endfor %}
      </div>
    </div>
    {% endfor %}
  </div>
</div>

<style>
.quartos-toolbar {
    display: flex; flex-wrap: wrap; align-items: center;
    gap: 16px; padding: 12px 16px; margin-bottom: 16px;
    background: var(--card-bg); border: 1px solid var(--border); border-radius: 8px;
}
.qt-stats { display: flex; gap: 18px; flex-wrap: wrap; font-size: 13px; color: var(--text-muted); }
.qt-stats strong { color: var(--text); font-weight: 600; }
.qt-stat.qt-warn strong { color: var(--orange); }
.qt-stat.qt-warn.ok strong { color: var(--green); }
.qt-actions { display: flex; gap: 8px; align-items: center; flex-wrap: wrap; margin-left: auto; }
.qt-actions input {
    background: var(--input-bg); color: var(--text);
    border: 1px solid var(--border); border-radius: 6px;
    padding: 6px 10px; font-size: 13px; width: 160px;
}
.qt-help { color: var(--text-dim); font-size: 12px; }

.qt-status { min-height: 20px; font-size: 12px; color: var(--text-muted); margin-bottom: 8px; }
.qt-status.ok { color: var(--green); }
.qt-status.err { color: var(--red); }

.quartos-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
    gap: 12px;
}

.quarto-card {
    background: var(--card-bg);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 10px 12px 12px;
    min-height: 140px;
    display: flex; flex-direction: column;
}
.quarto-card.quarto-pool { border-color: rgba(254,241,0,0.35); background: rgba(254,241,0,0.04); }
.quarto-card.drag-over { border-color: var(--yellow); background: rgba(254,241,0,0.08); }

.quarto-head {
    display: flex; justify-content: space-between; align-items: baseline;
    border-bottom: 1px solid var(--border-light); padding-bottom: 6px; margin-bottom: 8px;
}
.quarto-head h3 { font-size: 14px; font-weight: 600; color: var(--text); margin: 0; letter-spacing: -0.01em; }
.quarto-count { font-size: 12px; color: var(--text-dim); font-variant-numeric: tabular-nums; }
.quarto-count.over { color: var(--orange); }

.quarto-list { flex: 1; min-height: 60px; display: flex; flex-direction: column; gap: 6px; }

.pessoa-card {
    background: rgba(255,255,255,0.07);
    border: 1px solid var(--border-light);
    border-radius: 6px;
    padding: 7px 9px;
    cursor: grab;
    user-select: none;
    transition: background 0.12s, border-color 0.12s, transform 0.1s;
    touch-action: none;
}
.pessoa-card:hover { background: rgba(255,255,255,0.11); border-color: var(--border); }
.pessoa-card:active { cursor: grabbing; }
.pessoa-card.kind-equipe { border-left: 3px solid var(--blue); }
.pessoa-card.kind-aluno  { border-left: 3px solid var(--green); }
.pessoa-card.saving { background: rgba(98,204,60,0.18); }
.pessoa-card.error  { background: rgba(231,76,60,0.22); border-color: var(--red); }
.pessoa-card.sortable-ghost { opacity: 0.4; }
.pessoa-card.sortable-chosen { transform: scale(1.02); }

.pc-name { font-weight: 600; font-size: 13px; line-height: 1.25; }
.pc-meta { display: flex; gap: 6px; flex-wrap: wrap; margin-top: 4px; font-size: 11px; color: var(--text-muted); }
.pc-pill {
    display: inline-block; padding: 1px 6px; border-radius: 10px;
    background: rgba(255,255,255,0.08); font-size: 11px; line-height: 1.4;
}
.pc-pill.gender-M { background: rgba(52,152,219,0.22); color: #9ed3f5; }
.pc-pill.gender-F { background: rgba(231,76,60,0.18); color: #f5b1ab; }
.pc-pill.label { background: rgba(254,241,0,0.13); color: #fff5a0; }
.pc-pill.role  { background: rgba(123,47,160,0.35); color: #e3c8f0; }

@media (max-width: 600px) {
    .quartos-toolbar { flex-direction: column; align-items: stretch; }
    .qt-actions { margin-left: 0; }
    .qt-actions input { flex: 1; }
}
</style>

<script src="https://cdn.jsdelivr.net/npm/sortablejs@1.15.0/Sortable.min.js"></script>
<script>
(function() {
    const PREFIX = '{{ P }}';
    const EVENT_ID = {{ event['id'] }};
    const statusEl = document.getElementById('qt-status');
    let statusTimer = null;
    function flash(msg, cls) {
        statusEl.textContent = msg;
        statusEl.className = 'qt-status ' + (cls || '');
        if (statusTimer) clearTimeout(statusTimer);
        statusTimer = setTimeout(() => { statusEl.textContent=''; statusEl.className='qt-status'; }, 2500);
    }

    function updateCounts() {
        let unassigned = 0;
        document.querySelectorAll('.quarto-card').forEach(card => {
            const list = card.querySelector('[data-list]');
            const count = list.children.length;
            const countEl = card.querySelector('[data-count]');
            countEl.textContent = count + (count === 1 ? ' pessoa' : ' pessoas');
            countEl.classList.toggle('over', count > 4);
            if (card.dataset.room === '') unassigned = count;
        });
        const u = document.getElementById('qt-unassigned');
        u.textContent = unassigned;
        const stat = document.getElementById('qt-unassigned-stat');
        stat.classList.toggle('ok', unassigned === 0);
    }

    function persist(card, newRoom) {
        const kind = card.dataset.kind;
        const id = card.dataset.id;
        const url = PREFIX + '/' + EVENT_ID + '/' + (kind === 'aluno' ? 'alunos' : 'equipe') + '/' + id + '/quarto';
        const body = new URLSearchParams();
        body.set('room', newRoom || '');
        card.classList.add('saving');
        fetch(url, {
            method: 'POST',
            headers: { 'X-Requested-With': 'fetch', 'Content-Type': 'application/x-www-form-urlencoded' },
            body: body.toString(),
        }).then(r => {
            if (!r.ok) throw new Error('HTTP ' + r.status);
            return r.json();
        }).then(() => {
            card.classList.remove('saving');
            flash(card.querySelector('.pc-name').textContent.trim() + ' → ' + (newRoom ? 'Quarto ' + newRoom : 'sem quarto'), 'ok');
        }).catch(err => {
            card.classList.remove('saving');
            card.classList.add('error');
            flash('Erro ao salvar: ' + err.message + ' — recarregue a página', 'err');
            setTimeout(() => card.classList.remove('error'), 4000);
        });
    }

    function attachSortable(list) {
        Sortable.create(list, {
            group: 'quartos',
            animation: 150,
            ghostClass: 'sortable-ghost',
            chosenClass: 'sortable-chosen',
            onAdd: function(evt) {
                const card = evt.item;
                const newRoom = card.closest('.quarto-card').dataset.room;
                persist(card, newRoom);
                updateCounts();
            },
            onEnd: function() { updateCounts(); }
        });
    }

    document.querySelectorAll('[data-list]').forEach(attachSortable);
    updateCounts();

    function sortRoomCards() {
        const grid = document.querySelector('.quartos-grid');
        const cards = Array.from(grid.querySelectorAll('.quarto-card'));
        cards.sort((a, b) => {
            const ra = a.dataset.room, rb = b.dataset.room;
            if (ra === '' && rb !== '') return -1;
            if (rb === '' && ra !== '') return 1;
            const na = parseInt(ra, 10), nb = parseInt(rb, 10);
            if (!isNaN(na) && !isNaN(nb)) return na - nb;
            if (!isNaN(na)) return -1;
            if (!isNaN(nb)) return 1;
            return ra.localeCompare(rb);
        });
        cards.forEach(c => grid.appendChild(c));
    }

    window.addRoom = function() {
        const input = document.getElementById('qt-new-room');
        const val = input.value.trim();
        if (!val) { input.focus(); return; }
        const existing = document.querySelector('.quarto-card[data-room="' + val.replace(/"/g, '') + '"]');
        if (existing) {
            flash('Quarto ' + val + ' já existe', 'err');
            existing.scrollIntoView({behavior: 'smooth', block: 'center'});
            return;
        }
        const card = document.createElement('div');
        card.className = 'quarto-card';
        card.dataset.room = val;
        card.innerHTML =
            '<div class="quarto-head"><h3>Quarto ' + escapeHtml(val) + '</h3>' +
            '<span class="quarto-count" data-count></span></div>' +
            '<div class="quarto-list" data-list></div>';
        document.querySelector('.quartos-grid').appendChild(card);
        attachSortable(card.querySelector('[data-list]'));
        sortRoomCards();
        document.getElementById('qt-rooms').textContent =
            document.querySelectorAll('.quarto-card:not(.quarto-pool)').length;
        input.value = '';
        updateCounts();
    };

    function escapeHtml(s) {
        return String(s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
    }

    document.getElementById('qt-new-room').addEventListener('keydown', e => {
        if (e.key === 'Enter') { e.preventDefault(); window.addRoom(); }
    });
})();
</script>
"""


@app.route("/<int:event_id>/quartos/editar")
def quartos_editar(event_id):
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        conn.close()
        return redirect(PREFIX + "/lista")
    if not is_overnight_trip(event):
        conn.close()
        return redirect(PREFIX + f"/{event_id}#participantes")

    items = _quartos_items(conn, event_id)
    conn.close()

    by_room = {"__none__": []}
    for it in items:
        key = it["room"] if it["room"] else "__none__"
        by_room.setdefault(key, []).append(it)

    def room_sort_key(k):
        digits = "".join(ch for ch in k if ch.isdigit())
        return (0, int(digits)) if digits else (1, k)
    rooms = sorted([k for k in by_room.keys() if k != "__none__"], key=room_sort_key)

    def render_card(it):
        age_pill = f'<span class="pc-pill">{it["age"]} anos</span>' if it["age"] is not None else ""
        gender_pill = f'<span class="pc-pill gender-{it["gender"]}">{it["gender"]}</span>' if it["gender"] in ("M", "F") else ""
        label_class = "label" if it["kind"] == "aluno" else "role"
        label_pill = f'<span class="pc-pill {label_class}">{(it["label"] or "").strip()}</span>' if it["label"] else ""
        return (
            f'<div class="pessoa-card kind-{it["kind"]}" '
            f'data-kind="{it["kind"]}" data-id="{it["id"]}">'
            f'<div class="pc-name">{(it["name"] or "").strip()}</div>'
            f'<div class="pc-meta">{gender_pill}{age_pill}{label_pill}</div>'
            f'</div>'
        )

    total = len(items)
    unassigned_count = len(by_room["__none__"])

    return render_template_string(
        PAGE_OPEN.replace("{{ title }}", "Montar Quartos — " + event["name"]) + QUARTOS_BODY + PAGE_CLOSE,
        page='quartos',
        event=event, P=PREFIX,
        rooms=rooms, by_room=by_room,
        total=total, unassigned_count=unassigned_count,
        format_date=format_date,
        render_card=render_card,
    )


# ── Configurable manifest export ────────────────────────────────────────────

# Field definitions: (key, label, accessor)
# Accessor takes a row dict and returns a string.
def _idade(row):
    bd = row.get("birth_date")
    if not bd:
        return ""
    try:
        y, m, d = [int(x) for x in str(bd).split("-")]
        today = date.today()
        return str(today.year - y - ((today.month, today.day) < (m, d)))
    except Exception:
        return ""


TRANSPORT_LABELS = {
    "fb": "🚐 FB",
    "proprio": "🏠 Próprio",
}


def _transport_label(v):
    return TRANSPORT_LABELS.get(v, "—")


EXPORT_FIELDS = [
    ("categoria",  "Categoria",      lambda r: r.get("kind") or ""),
    ("quarto",     "Quarto",         lambda r: r.get("room") or ""),
    ("nome",       "Nome",           lambda r: r.get("name") or ""),
    ("gender",     "M/F",            lambda r: r.get("gender") or ""),
    ("nasc",       "Nascimento",     lambda r: format_date(r.get("birth_date")) if r.get("birth_date") else ""),
    ("idade",      "Idade",          _idade),
    ("rg",         "RG",             lambda r: r.get("rg") or ""),
    ("cpf",        "CPF",            lambda r: r.get("cpf") or ""),
    ("phone",      "Telefone",       lambda r: format_phone(r.get("phone"))),
    ("guardian",   "Responsável",    lambda r: r.get("guardian") or ""),
    ("grupo",      "Banda/Grupo",    lambda r: r.get("group_name") or ""),
    ("transporte", "Transporte",     lambda r: _transport_label(r.get("transport"))),
    ("alergias",   "Alergias",       lambda r: r.get("allergies") or ""),
    ("medico",     "Condição médica",lambda r: r.get("medical_condition") or ""),
    ("especiais",  "Necessidades esp.", lambda r: r.get("special_needs") or ""),
    ("notas",      "Notas",          lambda r: r.get("notes") or ""),
]

# "nome" is implicitly always included; presets list the *additional* fields.
EXPORT_PRESETS = {
    "hotel":      ["categoria", "quarto", "gender", "idade", "alergias", "medico", "especiais", "notas"],
    "embarque":   ["quarto", "phone", "guardian", "notas"],
    "seguro":     ["categoria", "gender", "nasc", "idade", "rg", "cpf", "phone"],
    "transporte": ["phone", "guardian", "notas"],   # name+phone+resp+notas, filtered to transport='fb'
    "completo":   ["categoria", "quarto", "gender", "nasc", "idade", "rg", "cpf", "phone", "guardian", "grupo", "transporte", "alergias", "medico", "especiais", "notas"],
}

# Some presets imply a transport filter applied at row-selection time
PRESET_TRANSPORT_FILTER = {
    "transporte": "fb",
}

PRESET_LABELS = {
    "hotel":      "🏨 Hotel",
    "embarque":   "🚌 Ônibus (viagem)",
    "seguro":     "🛡️ Seguradora",
    "transporte": "🚐 Transporte FB",
    "completo":   "📋 Completo",
}

PRESET_DESCRIPTIONS = {
    "hotel":      "quarto, sexo, idade, alergias, condições médicas, necessidades, notas",
    "embarque":   "quarto, telefone, responsável, notas",
    "seguro":     "identidade completa (DOB, RG, CPF, telefone)",
    "transporte": "apenas quem está no transporte FB — telefone, responsável, notas",
    "completo":   "todos os campos",
}


def _manifest_dicts(conn, event_id, include_alunos=True, include_equipe=True, transport_filter=None):
    """Return ordered list of dict rows (alunos first, then equipe).

    transport_filter: if set to 'fb' or 'proprio', only return rows with that transport value.
    """
    rows = []
    if include_alunos:
        for r in conn.execute("""
            SELECT s.name, s.gender, s.birth_date,
                   s.child_rg as rg, s.child_cpf as cpf,
                   s.guardian1_phone as phone, s.guardian1_name as guardian,
                   s.allergies, s.medical_condition, s.special_needs,
                   estu.room, estu.notes, estu.transport,
                   g.name as group_name, 'Aluno' as kind
            FROM event_students estu JOIN students s ON estu.student_id = s.id
            LEFT JOIN groups g ON estu.group_id = g.id
            WHERE estu.event_id = ?
            ORDER BY estu.room IS NULL, CAST(estu.room AS INTEGER), s.name
        """, (event_id,)):
            d = dict(r)
            if transport_filter and d.get("transport") != transport_filter:
                continue
            rows.append(d)
    if include_equipe:
        for r in conn.execute("""
            SELECT COALESCE(t.name, es.name) as name,
                   t.gender, t.birth_date,
                   t.rg, t.cpf, t.phone,
                   NULL as guardian, es.room, es.notes, es.transport,
                   NULL as allergies, NULL as medical_condition, NULL as special_needs,
                   NULL as group_name,
                   COALESCE(es.role, t.role, 'Equipe') as kind
            FROM event_staff es LEFT JOIN teachers t ON es.teacher_id = t.id
            WHERE es.event_id = ?
            ORDER BY es.room IS NULL, CAST(es.room AS INTEGER), name
        """, (event_id,)):
            d = dict(r)
            if transport_filter and d.get("transport") != transport_filter:
                continue
            rows.append(d)
    return rows


def _build_xlsx(event, rows, fields_selected):
    """Build an .xlsx Workbook in memory and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Manifesto"

    title = f"{event['name']}"
    subtitle_parts = [format_date(event["date"])]
    if event["end_date"] and event["end_date"] != event["date"]:
        subtitle_parts.append(f"→ {format_date(event['end_date'])}")
    if event["location"]:
        subtitle_parts.append(event["location"])
    subtitle = " · ".join(subtitle_parts)

    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, color="666666")

    header_row = 4
    headers = ["Nome"] + [label for key, label, _acc in EXPORT_FIELDS if key in fields_selected and key != "nome"]
    accessors = [("nome", lambda r: r.get("name") or "")] + [(key, acc) for key, _label, acc in EXPORT_FIELDS if key in fields_selected and key != "nome"]

    header_fill = PatternFill(start_color="F5C518", end_color="F5C518", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c_idx, (_key, acc) in enumerate(accessors, 1):
            ws.cell(row=r_idx, column=c_idx, value=acc(row))

    # Best-effort column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for r_idx in range(header_row + 1, header_row + 1 + len(rows)):
            v = ws.cell(row=r_idx, column=col_idx).value
            if v and len(str(v)) > max_len:
                max_len = len(str(v))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.route("/<int:event_id>/exportar/<preset>")
def exportar_preset(event_id, preset):
    """One-click direct download for a named preset (xlsx, both alunos+equipe)."""
    if preset not in EXPORT_PRESETS:
        return redirect(PREFIX + f"/{event_id}/exportar")
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        conn.close()
        return redirect(PREFIX + "/lista")
    fields_selected = set(EXPORT_PRESETS[preset]) | {"nome"}
    transport_filter = PRESET_TRANSPORT_FILTER.get(preset)
    rows = _manifest_dicts(conn, event_id, True, True, transport_filter=transport_filter)
    conn.close()
    safe_name = "".join(c if c.isalnum() else "-" for c in (event["name"] or "evento"))
    if HAS_OPENPYXL:
        data = _build_xlsx(event, rows, fields_selected)
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{preset}-{safe_name}.xlsx"'},
        )
    return redirect(PREFIX + f"/{event_id}/exportar")


@app.route("/<int:event_id>/exportar", methods=["GET", "POST"])
def exportar(event_id):
    conn = get_db()
    event = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
    if not event:
        conn.close()
        return redirect(PREFIX + "/lista")

    if request.method == "POST":
        fields_selected = set(request.form.getlist("fields"))
        fields_selected.add("nome")  # always
        scope = request.form.get("scope", "ambos")
        fmt = request.form.get("format", "xlsx")
        transport_filter = request.form.get("transport_filter") or None
        if transport_filter not in ("fb", "proprio"):
            transport_filter = None
        include_alunos = scope in ("ambos", "alunos")
        include_equipe = scope in ("ambos", "equipe")
        rows = _manifest_dicts(conn, event_id, include_alunos, include_equipe, transport_filter=transport_filter)
        conn.close()

        safe_name = "".join(c if c.isalnum() else "-" for c in (event["name"] or "evento"))

        if fmt == "xlsx" and HAS_OPENPYXL:
            data = _build_xlsx(event, rows, fields_selected)
            return Response(
                data,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="lista-{safe_name}.xlsx"'},
            )

        # CSV fallback
        out = io.StringIO()
        out.write("﻿")
        w = csv.writer(out, delimiter=";")
        w.writerow([f"{event['name']}"])
        sub = format_date(event["date"]) + (f" → {format_date(event['end_date'])}" if event["end_date"] and event["end_date"] != event["date"] else "")
        if event["location"]:
            sub += f" · {event['location']}"
        w.writerow([sub])
        w.writerow([])
        headers = ["Nome"] + [label for key, label, _ in EXPORT_FIELDS if key in fields_selected and key != "nome"]
        w.writerow(headers)
        accessors = [lambda r: r.get("name") or ""] + [acc for key, _l, acc in EXPORT_FIELDS if key in fields_selected and key != "nome"]
        for row in rows:
            w.writerow([acc(row) for acc in accessors])
        return Response(
            out.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="lista-{safe_name}.csv"'},
        )

    conn.close()
    return render_template_string(
        PAGE_OPEN.replace("{{ title }}", "Exportar — " + event["name"]) + EXPORT_BODY + PAGE_CLOSE,
        page='ficha', event=event,
        EXPORT_FIELDS=EXPORT_FIELDS, EXPORT_PRESETS=EXPORT_PRESETS, PRESET_LABELS=PRESET_LABELS, PRESET_DESCRIPTIONS=PRESET_DESCRIPTIONS,
        HAS_OPENPYXL=HAS_OPENPYXL,
        format_date=format_date,
    )


EXPORT_BODY = """
<div class="container">
    <div class="page-header">
        <div>
            <h1>📤 Exportar lista personalizada</h1>
            <div style="color: var(--text-muted); font-size: 14px;">{{ event['name'] }} · {{ format_date(event['date']) }}{% if event['end_date'] and event['end_date'] != event['date'] %} → {{ format_date(event['end_date']) }}{% endif %}</div>
        </div>
        <a href="{{ P }}/{{ event['id'] }}" class="btn btn-secondary btn-sm">← Voltar à ficha</a>
    </div>
    <div class="section">
        <p style="color: var(--text-muted); font-size: 13px; margin-bottom: 16px;">Para os modelos prontos (Hotel / Ônibus / Seguradora), use os botões na ficha do evento — esta página é só para casos sob medida.</p>
        <form method="POST" action="{{ P }}/{{ event['id'] }}/exportar">

            <h3 style="margin-top: 0;">Começar com um modelo</h3>
            <div style="display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 24px;">
                {% for key, label in PRESET_LABELS.items() %}
                <button type="button" class="btn btn-secondary btn-sm preset-btn" data-preset="{{ key }}" title="{{ PRESET_DESCRIPTIONS[key] }}">{{ label }}</button>
                {% endfor %}
            </div>

            <h3>Campos a incluir</h3>
            <p style="color: var(--text-muted); font-size: 13px; margin-top: -8px;">O nome é sempre incluído.</p>
            <div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 8px 16px; margin-bottom: 24px;">
                {% for key, label, _ in EXPORT_FIELDS %}
                {% if key != 'nome' %}
                <label style="display: flex; align-items: center; gap: 8px; cursor: pointer;">
                    <input type="checkbox" name="fields" value="{{ key }}" data-key="{{ key }}" class="field-cb" style="width: auto;">
                    {{ label }}
                </label>
                {% endif %}
                {% endfor %}
            </div>

            <h3>Quem incluir</h3>
            <div style="display: flex; gap: 16px; margin-bottom: 16px;">
                <label style="display: flex; align-items: center; gap: 8px; cursor: pointer;">
                    <input type="radio" name="scope" value="ambos" checked style="width: auto;"> Alunos + equipe
                </label>
                <label style="display: flex; align-items: center; gap: 8px; cursor: pointer;">
                    <input type="radio" name="scope" value="alunos" style="width: auto;"> Apenas alunos
                </label>
                <label style="display: flex; align-items: center; gap: 8px; cursor: pointer;">
                    <input type="radio" name="scope" value="equipe" style="width: auto;"> Apenas equipe
                </label>
            </div>
            <div style="display: flex; gap: 16px; margin-bottom: 24px; align-items: center;">
                <span style="color: var(--text-muted); font-size: 13px;">Filtrar por transporte:</span>
                <label style="display: flex; align-items: center; gap: 6px; cursor: pointer; font-size: 13px;">
                    <input type="radio" name="transport_filter" value="" checked style="width: auto;"> Todos
                </label>
                <label style="display: flex; align-items: center; gap: 6px; cursor: pointer; font-size: 13px;">
                    <input type="radio" name="transport_filter" value="fb" style="width: auto;"> 🚐 Apenas FB
                </label>
                <label style="display: flex; align-items: center; gap: 6px; cursor: pointer; font-size: 13px;">
                    <input type="radio" name="transport_filter" value="proprio" style="width: auto;"> 🏠 Apenas próprio
                </label>
            </div>

            <input type="hidden" name="format" value="xlsx" id="format-input">
            <details style="margin-bottom: 24px; color: var(--text-muted); font-size: 13px;">
                <summary style="cursor: pointer;">Mais opções</summary>
                <div style="margin-top: 8px; padding: 8px 0;">
                    <label style="display: flex; align-items: center; gap: 8px; cursor: pointer;">
                        <input type="checkbox" id="csv-toggle" style="width: auto;">
                        Baixar como CSV em vez de Excel
                    </label>
                </div>
            </details>

            <div class="form-actions">
                <button type="submit" class="btn btn-primary">📥 Baixar arquivo</button>
                <a href="{{ P }}/{{ event['id'] }}" class="btn btn-secondary">Cancelar</a>
            </div>
        </form>
    </div>
</div>
<script>
const PRESETS = {{ EXPORT_PRESETS|tojson }};
function clearPresetActive() {
    document.querySelectorAll('.preset-btn').forEach(b => { b.classList.remove('btn-primary'); b.classList.add('btn-secondary'); });
}
document.querySelectorAll('.preset-btn').forEach(btn => {
    btn.addEventListener('click', function() {
        const fields = PRESETS[this.dataset.preset] || [];
        document.querySelectorAll('input[name="fields"]').forEach(cb => {
            cb.checked = fields.includes(cb.dataset.key);
        });
        clearPresetActive();
        this.classList.remove('btn-secondary');
        this.classList.add('btn-primary');
    });
});
// Deselect preset styling once user starts editing checkboxes
document.querySelectorAll('.field-cb').forEach(cb => {
    cb.addEventListener('change', clearPresetActive);
});
// CSV toggle
document.getElementById('csv-toggle')?.addEventListener('change', function() {
    document.getElementById('format-input').value = this.checked ? 'csv' : 'xlsx';
});
// Default: trigger "completo" preset
document.querySelector('.preset-btn[data-preset="completo"]')?.click();
</script>
"""


# ── Special pages ──────────────────────────────────────────────────────────

@app.route("/<int:event_id>/dia")
def dia_view(event_id):
    return render_template_string(PAGE_OPEN.replace("{{ title }}", "Dia do Evento") + """
<div class="container">
    <div class="page-header"><h1>📱 Vista do Dia</h1></div>
    <div class="section">
        <p style="color: var(--text-muted); text-align: center; padding: 40px 0;">
            Vista mobile para o dia do evento — em construção.
        </p>
    </div>
</div>
""" + PAGE_CLOSE, page='ficha')


@app.route("/orcamento")
def orcamento_overview():
    conn = get_db()
    today = date.today().isoformat()
    rows = conn.execute("""
        SELECT e.id, e.name, e.date, e.type, e.status,
            COALESCE(SUM(CASE WHEN eb.is_income=1 THEN eb.estimated_amount ELSE 0 END), 0) as total_receitas,
            COALESCE(SUM(CASE WHEN eb.is_income=0 THEN eb.estimated_amount ELSE 0 END), 0) as total_despesas
        FROM events e
        LEFT JOIN event_budget eb ON eb.event_id = e.id
        WHERE e.status != 'cancelled' AND strftime('%Y', e.date) = strftime('%Y', 'now')
        GROUP BY e.id
        ORDER BY e.date
    """).fetchall()
    conn.close()
    return render_template_string(PAGE_OPEN.replace("{{ title }}", "Orçamento Geral") + ORCAMENTO_OVERVIEW_BODY + PAGE_CLOSE,
        page='orcamento', rows=rows,
        format_date=format_date, format_brl=format_brl,
        TYPE_LABELS=TYPE_LABELS, STATUS_LABELS=STATUS_LABELS)


# ══════════════════════════════════════════════════════════════════════════
# TEMPLATES
# ══════════════════════════════════════════════════════════════════════════

DASHBOARD_BODY = """
<div class="container">
    <div class="page-header">
        <h1>Fichas de Produção</h1>
    </div>

    <div class="stats">
        <div class="stat-card">
            <div class="label">Eventos este ano</div>
            <div class="value">{{ total_eventos }}</div>
        </div>
        <div class="stat-card">
            <div class="label">Próximos 3 meses</div>
            <div class="value">{{ proximos }}</div>
        </div>
        <div class="stat-card">
            <div class="label">Tarefas pendentes</div>
            <div class="value" style="color: {% if tarefas_pendentes > 10 %}var(--red){% elif tarefas_pendentes > 5 %}var(--orange){% else %}var(--yellow){% endif %}">{{ tarefas_pendentes }}</div>
        </div>
        <div class="stat-card">
            <div class="label">Orçamento (próximos)</div>
            <div class="value" style="font-size: 20px;">{{ format_brl(orcamento_total) }}</div>
        </div>
    </div>

    <div class="section">
        <h2>Próximas Fichas</h2>
        {% if upcoming_events %}
        <div class="table-wrap">
        <table>
            <thead>
                <tr>
                    <th>Data</th>
                    <th>Evento</th>
                    <th>Banda</th>
                    <th>Local</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
            {% for e in upcoming_events %}
            <tr>
                <td>{{ format_date(e['date']) }}</td>
                <td><a href="{{ P }}/{{ e['id'] }}">{{ e['name'] }}</a></td>
                <td>{{ e['banda'] or '—' }}</td>
                <td>{{ e['location'] or '—' }}</td>
                <td><span class="badge badge-{{ e['status'] }}">{{ STATUS_LABELS.get(e['status'], e['status']) }}</span></td>
            </tr>
            {% endfor %}
            </tbody>
        </table>
        </div>
        {% else %}
        <div class="empty"><div class="empty-icon">🎉</div><p>Sem eventos nos próximos 3 meses</p></div>
        {% endif %}
    </div>

    {% if overdue_tasks %}
    <div class="section">
        <h2>Tarefas Atrasadas</h2>
        <div class="table-wrap">
        <table>
            <thead>
                <tr><th>Evento</th><th>Tarefa</th><th>Prazo</th><th>Responsável</th></tr>
            </thead>
            <tbody>
            {% for t in overdue_tasks %}
            <tr>
                <td><a href="{{ P }}/{{ t['event_id'] }}#tarefas">{{ t['event_name'] }}</a></td>
                <td>{{ t['task'] }}</td>
                <td style="color: var(--red);">{{ format_date(t['deadline']) }}</td>
                <td>{{ t['responsible'] or '—' }}</td>
            </tr>
            {% endfor %}
            </tbody>
        </table>
        </div>
    </div>
    {% endif %}
</div>
"""

LISTA_BODY = """
<div class="container">
    <div class="page-header">
        <h1>Lista Completa</h1>
    </div>

    <form method="GET" class="filters">
        <select name="tipo" onchange="this.form.submit()">
            <option value="">Apresentações</option>
            <option value="all" {% if type_filter == 'all' %}selected{% endif %}>Todos os tipos</option>
            {% for t in FICHA_TYPES %}
            <option value="{{ t }}" {% if type_filter == t %}selected{% endif %}>{{ TYPE_LABELS.get(t, t) }}</option>
            {% endfor %}
        </select>
        <select name="status" onchange="this.form.submit()">
            <option value="">Todos os status</option>
            {% for s in EVENT_STATUSES %}
            <option value="{{ s }}" {% if status_filter == s %}selected{% endif %}>{{ STATUS_LABELS.get(s, s) }}</option>
            {% endfor %}
        </select>
        <select name="ano" onchange="this.form.submit()">
            {% for y in years %}
            <option value="{{ y['y'] }}" {% if year_filter == y['y'] %}selected{% endif %}>{{ y['y'] }}</option>
            {% endfor %}
        </select>
    </form>

    <div class="section">
        {% if events %}
        <div class="table-wrap">
        <table>
            <thead>
                <tr>
                    <th>Data</th>
                    <th>Evento</th>
                    <th>Banda</th>
                    <th>Local</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
            {% for e in events %}
            <tr>
                <td>{{ format_date(e['date']) }}{% if e['end_date'] and e['end_date'] != e['date'] %}<div style="color: var(--text-muted); font-size: 11px;">→ {{ format_date(e['end_date']) }}</div>{% endif %}</td>
                <td><a href="{{ P }}/{{ e['id'] }}">{{ e['name'] }}</a>
                    {% if e['end_date'] and e['end_date'] != e['date'] %}
                        {% set m = trip_missing.get(e['id'], 0) %}
                        {% if m > 0 %}<span class="badge badge-pendente" style="margin-left: 8px;" title="Pessoas com dados incompletos para manifesto">⚠ {{ m }} sem dados</span>{% endif %}
                    {% endif %}
                </td>
                <td>{{ e['banda'] or '—' }}</td>
                <td>{{ e['location'] or '—' }}</td>
                <td><span class="badge badge-{{ e['status'] }}">{{ STATUS_LABELS.get(e['status'], e['status']) }}</span></td>
            </tr>
            {% endfor %}
            </tbody>
        </table>
        </div>
        {% else %}
        <div class="empty"><div class="empty-icon">📭</div><p>Nenhum evento encontrado</p></div>
        {% endif %}
    </div>
</div>
"""

NOVO_EVENTO_BODY = """
<div class="container">
    <div class="page-header">
        <h1>{% if event %}Editar Ficha{% else %}Novo Evento{% endif %}</h1>
    </div>
    <div class="section">
        <form method="POST" action="{{ P }}/{% if event %}{{ event['id'] }}/editar{% else %}novo{% endif %}">
            <div class="form-grid">
                <div class="form-group form-full">
                    <label>Nome do evento *</label>
                    <input type="text" name="name" value="{{ event['name'] if event else '' }}" required>
                </div>
                <div class="form-group">
                    <label>Data de início *</label>
                    <input type="date" name="date" value="{{ event['date'] if event else '' }}" required>
                </div>
                <div class="form-group">
                    <label>Data de fim</label>
                    <input type="date" name="end_date" value="{{ event['end_date'] if event else '' }}">
                </div>
                <div class="form-group">
                    <label>Horário</label>
                    <input type="time" name="time" value="{{ event['time'] if event else '' }}">
                </div>
                <div class="form-group">
                    <label>Tipo</label>
                    <select name="type">
                        {% for t in EVENT_TYPES %}
                        <option value="{{ t }}" {% if event and event['type'] == t %}selected{% endif %}>{{ TYPE_LABELS.get(t, t) }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="form-group">
                    <label>Status</label>
                    <select name="status">
                        {% for s in EVENT_STATUSES %}
                        <option value="{{ s }}" {% if event and event['status'] == s %}selected{% endif %}>{{ STATUS_LABELS.get(s, s) }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="form-group">
                    <label>Local</label>
                    <input type="text" name="location" value="{{ event['location'] if event else '' }}">
                </div>
                <div class="form-group">
                    <label>Banda</label>
                    <select name="banda">
                        <option value="">—</option>
                        {% for b in bandas %}
                        <option value="{{ b }}" {% if event and event['banda'] == b %}selected{% endif %}>{{ b }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="form-group form-full">
                    <label>Descrição</label>
                    <textarea name="description">{{ event['description'] if event else '' }}</textarea>
                </div>
                <div class="form-group">
                    <label style="display: flex; align-items: center; gap: 8px; cursor: pointer;">
                        <input type="checkbox" name="is_public" {% if event and event['is_public'] %}checked{% endif %} style="width: auto;">
                        Evento público
                    </label>
                </div>
            </div>
            <div id="dupe-warning" style="display:none; background: #553300; border: 1px solid var(--yellow); border-radius: 8px; padding: 12px 16px; margin-bottom: 12px;">
                <strong style="color: var(--yellow);">Atenção:</strong> <span id="dupe-msg" style="color: var(--text);"></span>
            </div>
            <div class="form-actions">
                <button type="submit" class="btn btn-primary">Salvar</button>
                <a href="{{ P }}/{% if event %}{{ event['id'] }}{% else %}lista{% endif %}" class="btn btn-secondary">Cancelar</a>
                {% if event %}
                <form method="POST" action="{{ P }}/{{ event['id'] }}/excluir" style="margin-left: auto;">
                    <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Cancelar este evento?')">Cancelar evento</button>
                </form>
                {% endif %}
            </div>
        </form>
    </div>
</div>
<script>
(function() {
    const dateInput = document.querySelector('input[name="date"]');
    const warn = document.getElementById('dupe-warning');
    const msg = document.getElementById('dupe-msg');
    const eventId = '{{ event.id if event else "" }}';
    if (!dateInput) return;
    dateInput.addEventListener('change', async function() {
        const dt = this.value;
        if (!dt) { warn.style.display = 'none'; return; }
        const url = '{{ P }}/api/check-duplicates?date=' + dt + (eventId ? '&exclude=' + eventId : '');
        try {
            const res = await fetch(url);
            const dupes = await res.json();
            if (dupes.length > 0) {
                const names = dupes.map(d => d.name + (d.banda ? ' (' + d.banda + ')' : '')).join(', ');
                msg.textContent = 'Já existem eventos nesta data: ' + names + '. Verifique antes de criar um duplicado.';
                warn.style.display = 'block';
            } else {
                warn.style.display = 'none';
            }
        } catch(e) { warn.style.display = 'none'; }
    });
    // Trigger check on page load for edit mode
    if (dateInput.value && eventId) dateInput.dispatchEvent(new Event('change'));
})();
</script>
"""

FICHA_BODY = """
<div class="container">
    <div class="page-header">
        <div>
            <div style="font-size: 12px; color: var(--text-muted); margin-bottom: 4px;">
                {{ TYPE_LABELS.get(event['type'], event['type'] or '') }}
                {% if event['date'] %} · {{ format_date(event['date']) }}{% endif %}
                {% if event['end_date'] and event['end_date'] != event['date'] %} → {{ format_date(event['end_date']) }}{% endif %}
            </div>
            <h1>{{ event['name'] }}</h1>
        </div>
        <div style="display: flex; gap: 8px; flex-wrap: wrap;">
            <span class="badge badge-{{ event['status'] }}">{{ STATUS_LABELS.get(event['status'], event['status']) }}</span>
            <a href="{{ P }}/{{ event['id'] }}/editar" class="btn btn-secondary btn-sm">✏️ Editar Ficha</a>
            <a href="{{ P }}/{{ event['id'] }}/dia" class="btn btn-secondary btn-sm">📱 Dia</a>
            <button id="btn-whatsapp" class="btn btn-secondary btn-sm" onclick="copyWhatsApp()">📋 WhatsApp</button>
            {% if event['lessons_learned'] %}
            <a href="{{ P }}/{{ event['id'] }}/licoes" class="btn btn-sm" style="background: var(--green); color: #fff;">📝 Lições</a>
            {% endif %}
        </div>
    </div>

    <div class="tabs">
        <input type="radio" name="tab" id="tab-resumo" class="tab-input" checked>
        <input type="radio" name="tab" id="tab-cronograma" class="tab-input">
        <input type="radio" name="tab" id="tab-equipe" class="tab-input">
        <input type="radio" name="tab" id="tab-participantes" class="tab-input">
        <input type="radio" name="tab" id="tab-logistica" class="tab-input">
        <input type="radio" name="tab" id="tab-orcamento" class="tab-input">
        <input type="radio" name="tab" id="tab-contatos" class="tab-input">
        <input type="radio" name="tab" id="tab-tarefas" class="tab-input">

        <div class="tab-labels">
            <label for="tab-resumo" class="tab-label">Resumo</label>
            <label for="tab-cronograma" class="tab-label">Cronograma</label>
            <label for="tab-equipe" class="tab-label">Equipe ({{ equipe|length }})</label>
            <label for="tab-participantes" class="tab-label">Alunos ({{ participantes|length }})</label>
            <label for="tab-logistica" class="tab-label">Logística ({{ logistica|length }})</label>
            <label for="tab-orcamento" class="tab-label">Orçamento</label>
            <label for="tab-contatos" class="tab-label">Contatos ({{ contatos|length }})</label>
            <label for="tab-tarefas" class="tab-label">Checklist ({{ tarefas_done }}/{{ total_tarefas }})</label>
        </div>

        <div class="tab-panels">

        <!-- RESUMO -->
        <div class="tab-content panel-resumo">
            <div class="stats" style="margin-bottom: 20px;">
                <div class="stat-card">
                    <div class="label">Orçamento estimado</div>
                    <div class="value" style="font-size: 20px;">{{ format_brl(despesas_total) }}</div>
                </div>
                <div class="stat-card">
                    <div class="label">Receitas previstas</div>
                    <div class="value" style="font-size: 20px; color: var(--green);">{{ format_brl(receitas) }}</div>
                </div>
                <div class="stat-card">
                    <div class="label">Checklist</div>
                    <div class="value">{{ tarefas_done }}<span style="font-size: 14px; color: var(--text-muted);">/{{ total_tarefas }}</span></div>
                </div>
                <div class="stat-card">
                    <div class="label">Equipe / Alunos</div>
                    <div class="value">{{ equipe|length }}<span style="font-size: 14px; color: var(--text-muted);"> / {{ participantes|length }}</span></div>
                </div>
            </div>

            <div class="section">
                <h2>Detalhes</h2>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 16px;">
                    {% if event['date'] %}
                    <div>
                        <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-muted); margin-bottom: 4px;">Data</div>
                        <div>{{ format_date(event['date']) }}{% if event['end_date'] and event['end_date'] != event['date'] %} → {{ format_date(event['end_date']) }}{% endif %}</div>
                    </div>
                    {% endif %}
                    {% if event['time'] %}
                    <div>
                        <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-muted); margin-bottom: 4px;">Horário</div>
                        <div>{{ event['time'] }}</div>
                    </div>
                    {% endif %}
                    {% if event['location'] %}
                    <div>
                        <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-muted); margin-bottom: 4px;">Local</div>
                        <div>{{ event['location'] }}</div>
                    </div>
                    {% endif %}
                    {% if event['description'] %}
                    <div style="grid-column: 1 / -1;">
                        <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-muted); margin-bottom: 4px;">Descrição</div>
                        <div style="white-space: pre-line; color: rgba(255,255,255,0.8);">{{ event['description'] }}</div>
                    </div>
                    {% endif %}
                </div>
            </div>
        </div>

        <!-- CRONOGRAMA -->
        <div class="tab-content panel-cronograma">
            <div class="crono-add">
                <input type="text" id="crono-input" placeholder="09:30 Banda Preta | local | responsável" autocomplete="off">
            </div>
            <div class="crono-hint">Formato: <code>HH:MM Atividade</code> · Separar local e responsável com <code>|</code> · Horário opcional · Enter para adicionar</div>
            <div id="crono-list" class="crono-list" style="margin-top: 16px;"></div>
            <script>var CRONO_DATA = {{ cronograma_json|safe }}; var CRONO_EVENT_ID = {{ event['id'] }}; var CRONO_PREFIX = "{{ P }}";</script>
        </div>

        <!-- EQUIPE -->
        <div class="tab-content panel-equipe">
            <div style="text-align:right; margin-bottom: 12px;">
                <button class="btn btn-primary btn-sm" onclick="toggleShow('add-equipe')">+ Adicionar</button>
            </div>
            <div id="add-equipe" style="display:none;" class="section">
                <form method="POST" action="{{ P }}/{{ event['id'] }}/equipe/adicionar">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Equipe (opcional)</label>
                            <select name="teacher_id" id="teacher-select" onchange="fillTeacherName(this)">
                                <option value="">— Selecionar —</option>
                                {% for t in teachers %}
                                <option value="{{ t['id'] }}" data-name="{{ t['name'] }}">{{ t['name'] }}</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Nome</label>
                            <input type="text" name="name" id="teacher-name-field" placeholder="ou digite nome de freelance/voluntário">
                        </div>
                        <div class="form-group">
                            <label>Função do dia</label>
                            <input type="text" name="role" placeholder="ex: Regente, Músico, Produtora, Técnico">
                        </div>
                        <div class="form-group">
                            <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; padding-top: 24px;">
                                <input type="checkbox" name="confirmed" style="width: auto;">
                                Confirmado
                            </label>
                        </div>
                        <div class="form-group form-full">
                            <label>Notas</label>
                            <input type="text" name="notes">
                        </div>
                    </div>
                    <div class="form-actions">
                        <button type="submit" class="btn btn-primary">Adicionar</button>
                        <button type="button" class="btn btn-secondary" onclick="toggleShow('add-equipe')">Cancelar</button>
                    </div>
                </form>
            </div>

            {% if equipe %}
            {% if is_trip %}
            {% if missing_equipe > 0 %}
            <div style="margin-bottom: 12px;"><span class="badge badge-pendente">⚠️ {{ missing_equipe }} integrante{{ 's' if missing_equipe != 1 else '' }} sem dados completos</span> — para preencher, edite o cadastro do(a) professor(a) em /folha (ou peça ao Tom).</div>
            {% endif %}
            <div class="table-wrap"><table>
                <thead><tr>
                    <th style="width: 70px;">Quarto</th>
                    <th style="width: 110px;">Transporte</th>
                    <th>Nome</th>
                    <th>Função</th>
                    <th>M/F</th>
                    <th>Nasc.</th>
                    <th>RG</th>
                    <th>CPF</th>
                    <th>Telefone</th>
                    <th>Notas</th>
                    <th>Status</th>
                    <th>Faltam</th>
                    <th></th>
                </tr></thead><tbody>
                {% for item in equipe %}
                {% set smissing = staff_missing_fields(item) %}
                <tr {% if smissing %}style="background: rgba(180, 80, 0, 0.10);"{% endif %}>
                    <td>
                        <input type="text" value="{{ item['room'] or '' }}" data-id="{{ item['id'] }}" data-kind="equipe" class="room-input" placeholder="—" style="width: 60px; padding: 4px 6px; font-size: 13px;">
                    </td>
                    <td>
                        <select class="transporte-select" data-id="{{ item['id'] }}" data-kind="equipe" style="width: 100px; padding: 3px 4px; font-size: 12px;">
                            <option value="" {% if not item['transport'] %}selected{% endif %}>—</option>
                            <option value="fb" {% if item['transport'] == 'fb' %}selected{% endif %}>🚐 FB</option>
                            <option value="proprio" {% if item['transport'] == 'proprio' %}selected{% endif %}>🏠 Próprio</option>
                        </select>
                    </td>
                    <td>{{ item['name'] or item['teacher_name'] or '—' }}</td>
                    <td>{{ item['role'] or '—' }}</td>
                    <td>{{ item['t_gender'] or '—' }}</td>
                    <td>{{ format_date(item['t_birth_date']) }}</td>
                    <td style="font-size: 12px;">{{ item['t_rg'] or '—' }}</td>
                    <td style="font-size: 12px;">{{ item['t_cpf'] or '—' }}</td>
                    <td style="font-size: 12px;">{{ format_phone(item['t_phone']) or '—' }}</td>
                    <td>
                        <input type="text" value="{{ item['notes'] or '' }}" data-id="{{ item['id'] }}" data-kind="equipe" class="notes-input" placeholder="—" style="width: 140px; padding: 4px 6px; font-size: 12px;">
                    </td>
                    <td>
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/equipe/{{ item['id'] }}/toggle" style="display:inline;">
                            {% if item['confirmed'] %}
                            <button type="submit" class="badge badge-confirmed" style="border:none; cursor:pointer;" title="Clique para marcar pendente">✓</button>
                            {% else %}
                            <button type="submit" class="badge badge-pendente" style="border:none; cursor:pointer;" title="Clique para confirmar">?</button>
                            {% endif %}
                        </form>
                    </td>
                    <td>{% if smissing %}<span style="color: #f5a623; font-size: 12px;">⚠ {{ smissing|join(', ') }}</span>{% else %}<span style="color: #4caf50;">✓</span>{% endif %}</td>
                    <td>
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/equipe/{{ item['id'] }}/excluir" style="display:inline;">
                            <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Remover?')">🗑️</button>
                        </form>
                    </td>
                </tr>
                {% endfor %}
            </tbody></table></div>
            {% else %}
            <div class="table-wrap">
            <table>
                <thead>
                    <tr><th>Nome</th><th>Função do dia</th><th style="width: 110px;">Transporte</th><th>Status</th><th></th></tr>
                </thead>
                <tbody>
                {% for item in equipe %}
                <tr>
                    <td>{{ item['name'] or item['teacher_name'] or '—' }}</td>
                    <td>{{ item['role'] or '—' }}</td>
                    <td>
                        <select class="transporte-select" data-id="{{ item['id'] }}" data-kind="equipe" style="width: 100px; padding: 3px 4px; font-size: 12px;">
                            <option value="" {% if not item['transport'] %}selected{% endif %}>—</option>
                            <option value="fb" {% if item['transport'] == 'fb' %}selected{% endif %}>🚐 FB</option>
                            <option value="proprio" {% if item['transport'] == 'proprio' %}selected{% endif %}>🏠 Próprio</option>
                        </select>
                    </td>
                    <td>
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/equipe/{{ item['id'] }}/toggle" style="display:inline;">
                            {% if item['confirmed'] %}
                            <button type="submit" class="badge badge-confirmed" style="border:none; cursor:pointer;" title="Clique para marcar pendente">Confirmado</button>
                            {% else %}
                            <button type="submit" class="badge badge-pendente" style="border:none; cursor:pointer;" title="Clique para confirmar">Pendente</button>
                            {% endif %}
                        </form>
                    </td>
                    <td>
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/equipe/{{ item['id'] }}/excluir" style="display:inline;">
                            <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Remover?')">🗑️</button>
                        </form>
                    </td>
                </tr>
                {% endfor %}
                </tbody>
            </table>
            </div>
            {% endif %}
            {% else %}
            <div class="empty"><div class="empty-icon">👥</div><p>Sem equipe definida</p></div>
            {% endif %}
        </div>

        <!-- ALUNOS -->
        <div class="tab-content panel-participantes">
            {% set tc = transport_counts %}
            {% if participantes %}
            <div style="background: var(--bg-secondary, #1a1a1a); border: 1px solid var(--border, #333); border-radius: 8px; padding: 12px 16px; margin-bottom: 16px; display: flex; flex-wrap: wrap; gap: 12px; align-items: center;">
                <strong style="color: var(--text);">🚐 Transporte:</strong>
                <span><span style="color: #4caf50;">🚐 {{ tc['alunos_fb'] }}</span> FB</span>
                <span><span style="color: #2196f3;">🏠 {{ tc['alunos_proprio'] }}</span> próprio</span>
                {% if tc['alunos_unset'] > 0 %}
                <span style="color: var(--text-muted);">❓ {{ tc['alunos_unset'] }} não definido</span>
                {% endif %}
                <div style="margin-left: auto; display: flex; gap: 6px; flex-wrap: wrap; align-items: center;">
                    {% if tc['alunos_unset'] > 0 %}
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/alunos/transporte-bulk" style="display:inline;">
                        <input type="hidden" name="transport" value="fb">
                        <input type="hidden" name="only_unset" value="1">
                        <button type="submit" class="btn btn-secondary btn-sm" onclick="return confirm('Marcar os {{ tc[\"alunos_unset\"] }} alunos pendentes como transporte FB?')">Marcar pendentes como FB</button>
                    </form>
                    {% endif %}
                    <a class="btn btn-primary btn-sm" href="{{ P }}/{{ event['id'] }}/exportar/transporte" title="Apenas alunos+equipe no transporte FB — telefone, responsável, notas">📥 Lista de transporte FB.xlsx</a>
                </div>
            </div>
            {% endif %}
            {% if is_trip %}
            <div style="background: var(--bg-secondary, #1a1a1a); border: 1px solid var(--border, #333); border-radius: 8px; padding: 12px 16px; margin-bottom: 16px; display: flex; flex-wrap: wrap; gap: 12px; align-items: center;">
                <strong style="color: var(--yellow);">🧳 Viagem com pernoite</strong>
                {% if missing_alunos > 0 %}
                <span class="badge badge-pendente" title="Alunos com dados pessoais incompletos">⚠️ {{ missing_alunos }} aluno{{ 's' if missing_alunos != 1 else '' }} sem dados completos</span>
                {% else %}
                <span class="badge badge-confirmed">✓ Dados pessoais completos</span>
                {% endif %}
                <div style="margin-left: auto; display: flex; gap: 6px; flex-wrap: wrap; align-items: center;">
                    <span style="font-size: 12px; color: var(--text-muted); margin-right: 4px;">Baixar para:</span>
                    <a class="btn btn-primary btn-sm" href="{{ P }}/{{ event['id'] }}/exportar/hotel" title="quarto, sexo, idade, alergias, condições médicas, notas">🏨 Hotel.xlsx</a>
                    <a class="btn btn-primary btn-sm" href="{{ P }}/{{ event['id'] }}/exportar/embarque" title="quarto, telefone, responsável, notas">🚌 Ônibus.xlsx</a>
                    <a class="btn btn-primary btn-sm" href="{{ P }}/{{ event['id'] }}/exportar/seguro" title="identidade completa: DOB, RG, CPF, telefone">🛡️ Seguradora.xlsx</a>
                    <a class="btn btn-secondary btn-sm" href="{{ P }}/{{ event['id'] }}/exportar">Outro modelo…</a>
                    <span style="border-left: 1px solid var(--border); height: 20px; margin: 0 4px;"></span>
                    <a class="btn btn-primary btn-sm" href="{{ P }}/{{ event['id'] }}/quartos/editar" title="Montar quartos arrastando">🛏️ Montar quartos</a>
                    <a class="btn btn-secondary btn-sm" href="{{ P }}/{{ event['id'] }}/lista-quartos" target="_blank" title="Versão para imprimir">🖨️ Quartos</a>
                    <a class="btn btn-secondary btn-sm" href="{{ P }}/{{ event['id'] }}/lista-embarque" target="_blank" title="Versão para imprimir">🖨️ Embarque</a>
                </div>
            </div>
            {% endif %}
            <div style="text-align:right; margin-bottom: 12px; display: flex; gap: 8px; justify-content: flex-end;">
                <button class="btn btn-secondary btn-sm" onclick="toggleShow('add-aluno-individual')">+ Adicionar Aluno</button>
                <button class="btn btn-primary btn-sm" onclick="toggleShow('add-participantes')">+ Adicionar Grupo</button>
            </div>
            <div id="add-aluno-individual" style="display:none;" class="section">
                <form method="POST" action="{{ P }}/{{ event['id'] }}/alunos/adicionar-individual">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Aluno</label>
                            <select name="student_id" required>
                                <option value="">— Selecionar aluno —</option>
                                {% for s in students_all %}
                                <option value="{{ s['id'] }}">{{ s['name'] }}</option>
                                {% endfor %}
                            </select>
                        </div>
                    </div>
                    <div class="form-actions">
                        <button type="submit" class="btn btn-primary">Adicionar</button>
                        <button type="button" class="btn btn-secondary" onclick="toggleShow('add-aluno-individual')">Cancelar</button>
                    </div>
                </form>
            </div>
            <div id="add-participantes" style="display:none;" class="section">
                <form method="POST" action="{{ P }}/{{ event['id'] }}/alunos/adicionar">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Grupo</label>
                            <select name="group_id" required>
                                <option value="">— Selecionar grupo —</option>
                                {% for g in groups %}
                                <option value="{{ g['id'] }}">{{ g['name'] }}</option>
                                {% endfor %}
                                <option value="Banda Preta:A">Banda Preta A</option>
                                <option value="Banda Preta:B">Banda Preta B</option>
                            </select>
                        </div>
                    </div>
                    <div class="form-actions">
                        <button type="submit" class="btn btn-primary">Adicionar todos do grupo</button>
                        <button type="button" class="btn btn-secondary" onclick="toggleShow('add-participantes')">Cancelar</button>
                    </div>
                </form>
            </div>

            {% if participantes %}
            {% if is_trip %}
            <div class="table-wrap"><table>
                <thead><tr>
                    <th style="width: 70px;">Quarto</th>
                    <th style="width: 110px;">Transporte</th>
                    <th>Nome</th>
                    <th>M/F</th>
                    <th>Nasc.</th>
                    <th>RG</th>
                    <th>CPF</th>
                    <th>Responsável</th>
                    <th>Telefone</th>
                    <th>Notas</th>
                    <th>Faltam</th>
                    <th></th>
                </tr></thead><tbody>
                {% for p in participantes %}
                {% set missing = participant_missing_fields(p) %}
                <tr {% if missing %}style="background: rgba(180, 80, 0, 0.10);"{% endif %}>
                    <td>
                        <input type="text" value="{{ p['room'] or '' }}" data-id="{{ p['id'] }}" data-kind="aluno" class="room-input" placeholder="—" style="width: 60px; padding: 4px 6px; font-size: 13px;">
                    </td>
                    <td>
                        <select class="transporte-select" data-id="{{ p['id'] }}" data-kind="aluno" style="width: 100px; padding: 3px 4px; font-size: 12px;">
                            <option value="" {% if not p['transport'] %}selected{% endif %}>—</option>
                            <option value="fb" {% if p['transport'] == 'fb' %}selected{% endif %}>🚐 FB</option>
                            <option value="proprio" {% if p['transport'] == 'proprio' %}selected{% endif %}>🏠 Próprio</option>
                        </select>
                    </td>
                    <td>{{ p['name'] }}
                        {% if p['group_name'] %}<div style="color: var(--text-muted); font-size: 11px;">{{ p['group_name'] }}</div>{% endif %}
                        {% set warns = [] %}
                        {% if p['allergies'] %}{% set _ = warns.append('Alergia: ' + p['allergies']) %}{% endif %}
                        {% if p['medical_condition'] %}{% set _ = warns.append('Méd: ' + p['medical_condition']) %}{% endif %}
                        {% if p['special_needs'] %}{% set _ = warns.append('Esp: ' + p['special_needs']) %}{% endif %}
                        {% if warns %}<div style="color: #f5a623; font-size: 11px;" title="Vem do cadastro do aluno">⚕ {{ warns|join(' · ') }}</div>{% endif %}
                    </td>
                    <td>{{ p['gender'] or '—' }}</td>
                    <td>{{ format_date(p['birth_date']) }}</td>
                    <td style="font-size: 12px;">{{ p['child_rg'] or '—' }}</td>
                    <td style="font-size: 12px;">{{ p['child_cpf'] or '—' }}</td>
                    <td style="font-size: 12px;">{{ p['guardian1_name'] or '—' }}</td>
                    <td style="font-size: 12px;">{{ format_phone(p['guardian1_phone']) or '—' }}</td>
                    <td>
                        <input type="text" value="{{ p['notes'] or '' }}" data-id="{{ p['id'] }}" data-kind="aluno" class="notes-input" placeholder="—" style="width: 140px; padding: 4px 6px; font-size: 12px;" title="Notas para esta viagem (vai pra exportação)">
                    </td>
                    <td>{% if missing %}<span title="Editar no perfil do aluno" style="color: #f5a623;">⚠ {{ missing|join(', ') }}</span>{% else %}<span style="color: #4caf50;">✓</span>{% endif %}</td>
                    <td>
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/alunos/{{ p['id'] }}/excluir" style="display:inline;">
                            <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Remover?')">🗑️</button>
                        </form>
                    </td>
                </tr>
                {% endfor %}
            </tbody></table></div>
            {% else %}
            {% set current_group = namespace(name='') %}
            {% for p in participantes %}
            {% if p['group_name'] != current_group.name %}
            {% if not loop.first %}</tbody></table></div>{% endif %}
            {% set current_group.name = p['group_name'] %}
            <div style="margin-bottom: 8px; margin-top: {{ '20px' if not loop.first else '0' }};">
                <strong style="color: var(--yellow);">{{ p['group_name'] or 'Sem grupo' }}</strong>
            </div>
            <div class="table-wrap"><table>
                <thead><tr><th>Nome</th><th style="width: 110px;">Transporte</th><th></th></tr></thead>
                <tbody>
            {% endif %}
            <tr>
                <td>{{ p['name'] }}</td>
                <td>
                    <select class="transporte-select" data-id="{{ p['id'] }}" data-kind="aluno" style="width: 100px; padding: 3px 4px; font-size: 12px;">
                        <option value="" {% if not p['transport'] %}selected{% endif %}>—</option>
                        <option value="fb" {% if p['transport'] == 'fb' %}selected{% endif %}>🚐 FB</option>
                        <option value="proprio" {% if p['transport'] == 'proprio' %}selected{% endif %}>🏠 Próprio</option>
                    </select>
                </td>
                <td>
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/alunos/{{ p['id'] }}/excluir" style="display:inline;">
                        <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Remover?')">🗑️</button>
                    </form>
                </td>
            </tr>
            {% endfor %}
            </tbody></table></div>
            {% endif %}
            {% else %}
            <div class="empty"><div class="empty-icon">🎺</div><p>Sem alunos. Use o botão acima para adicionar um grupo inteiro.</p></div>
            {% endif %}
            <script>
            (function() {
                function postUpdate(elem, field, urlSuffix, ackEl) {
                    const id = elem.dataset.id;
                    const kind = elem.dataset.kind;
                    const path = kind === 'aluno' ? 'alunos' : 'equipe';
                    const fd = new FormData();
                    fd.append(field, elem.value);
                    fetch('{{ P }}/{{ event["id"] }}/' + path + '/' + id + '/' + urlSuffix, {
                        method: 'POST', body: fd, headers: {'X-Requested-With': 'fetch'}
                    }).then(r => {
                        const target = ackEl || elem;
                        if (r.ok) { target.style.background = '#0a4'; setTimeout(()=>target.style.background='', 600); }
                    });
                }
                function bindBlur(selector, field, urlSuffix) {
                    document.querySelectorAll(selector).forEach(function(inp) {
                        let prev = inp.value;
                        inp.addEventListener('blur', function() {
                            if (this.value === prev) return;
                            prev = this.value;
                            postUpdate(this, field, urlSuffix);
                        });
                    });
                }
                function bindChange(selector, field, urlSuffix) {
                    document.querySelectorAll(selector).forEach(function(sel) {
                        sel.addEventListener('change', function() { postUpdate(this, field, urlSuffix); });
                    });
                }
                {% if is_trip %}
                bindBlur('.room-input', 'room', 'quarto');
                bindBlur('.notes-input', 'notes', 'notas');
                {% endif %}
                bindChange('.transporte-select', 'transport', 'transporte');
            })();
            </script>
        </div>

        <!-- LOGÍSTICA -->
        <div class="tab-content panel-logistica">
            <div style="text-align:right; margin-bottom: 12px;">
                <button class="btn btn-primary btn-sm" onclick="toggleShow('add-logistica')">+ Adicionar</button>
            </div>
            <div id="add-logistica" style="display:none;" class="section">
                <form method="POST" action="{{ P }}/{{ event['id'] }}/logistica/adicionar">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Categoria</label>
                            <select name="category">
                                {% for c in LOGISTICS_CATEGORIES %}
                                <option value="{{ c }}">{{ c.capitalize() }}</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Status</label>
                            <select name="status">
                                <option value="pendente">Pendente</option>
                                <option value="em_andamento">Em andamento</option>
                                <option value="concluido">Concluído</option>
                            </select>
                        </div>
                        <div class="form-group form-full">
                            <label>Item *</label>
                            <input type="text" name="item" required>
                        </div>
                        <div class="form-group form-full">
                            <label>Detalhes</label>
                            <input type="text" name="details">
                        </div>
                        <div class="form-group">
                            <label>Responsável</label>
                            <input type="text" name="responsible">
                        </div>
                        <div class="form-group">
                            <label>Prazo</label>
                            <input type="date" name="deadline">
                        </div>
                        <div class="form-group form-full">
                            <label>Notas</label>
                            <input type="text" name="notes">
                        </div>
                    </div>
                    <div class="form-actions">
                        <button type="submit" class="btn btn-primary">Salvar</button>
                        <button type="button" class="btn btn-secondary" onclick="toggleShow('add-logistica')">Cancelar</button>
                    </div>
                </form>
            </div>

            {% if logistica %}
            {% set current_cat = namespace(name='') %}
            {% for item in logistica %}
            {% if item['category'] != current_cat.name %}
            {% if not loop.first %}</tbody></table></div>{% endif %}
            {% set current_cat.name = item['category'] %}
            <div style="margin-bottom: 8px; margin-top: {{ '20px' if not loop.first else '0' }};">
                <strong style="color: var(--yellow);">{{ item['category'].capitalize() }}</strong>
            </div>
            <div class="table-wrap"><table>
                <thead><tr><th>Item</th><th>Detalhes</th><th>Responsável</th><th>Prazo</th><th>Status</th><th></th></tr></thead>
                <tbody>
            {% endif %}
            <tr>
                <td>{{ item['item'] }}</td>
                <td>{{ item['details'] or '—' }}</td>
                <td>{{ item['responsible'] or '—' }}</td>
                <td>{{ format_date(item['deadline']) }}</td>
                <td><span class="badge badge-{{ item['status'] }}">{{ item['status'].replace('_', ' ').capitalize() }}</span></td>
                <td style="white-space: nowrap;">
                    <button class="btn btn-secondary btn-sm" onclick="toggleEdit('logistica-{{ item['id'] }}')">✏️</button>
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/logistica/{{ item['id'] }}/excluir" style="display:inline;">
                        <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Excluir?')">🗑️</button>
                    </form>
                </td>
            </tr>
            <tr id="edit-logistica-{{ item['id'] }}" style="display:none;">
                <td colspan="99">
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/logistica/{{ item['id'] }}/editar" style="padding: 12px 0;">
                        <div class="form-grid">
                            <div class="form-group">
                                <label>Categoria</label>
                                <select name="category">
                                    {% for c in LOGISTICS_CATEGORIES %}
                                    <option value="{{ c }}" {% if c == item['category'] %}selected{% endif %}>{{ c.capitalize() }}</option>
                                    {% endfor %}
                                </select>
                            </div>
                            <div class="form-group">
                                <label>Status</label>
                                <select name="status">
                                    <option value="pendente" {% if item['status'] == 'pendente' %}selected{% endif %}>Pendente</option>
                                    <option value="em_andamento" {% if item['status'] == 'em_andamento' %}selected{% endif %}>Em andamento</option>
                                    <option value="concluido" {% if item['status'] == 'concluido' %}selected{% endif %}>Concluído</option>
                                </select>
                            </div>
                            <div class="form-group form-full">
                                <label>Item</label>
                                <input type="text" name="item" value="{{ item['item'] or '' }}">
                            </div>
                            <div class="form-group form-full">
                                <label>Detalhes</label>
                                <input type="text" name="details" value="{{ item['details'] or '' }}">
                            </div>
                            <div class="form-group">
                                <label>Responsável</label>
                                <input type="text" name="responsible" value="{{ item['responsible'] or '' }}">
                            </div>
                            <div class="form-group">
                                <label>Prazo</label>
                                <input type="date" name="deadline" value="{{ item['deadline'] or '' }}">
                            </div>
                            <div class="form-group form-full">
                                <label>Notas</label>
                                <input type="text" name="notes" value="{{ item['notes'] or '' }}">
                            </div>
                        </div>
                        <div class="form-actions">
                            <button type="submit" class="btn btn-primary btn-sm">Salvar</button>
                            <button type="button" class="btn btn-secondary btn-sm" onclick="toggleEdit('logistica-{{ item['id'] }}')">Cancelar</button>
                        </div>
                    </form>
                </td>
            </tr>
            {% endfor %}
            {% if logistica %}</tbody></table></div>{% endif %}
            {% else %}
            <div class="empty"><div class="empty-icon">📦</div><p>Sem itens de logística</p></div>
            {% endif %}
        </div>

        <!-- ORÇAMENTO -->
        <div class="tab-content panel-orcamento">
            <div style="text-align:right; margin-bottom: 12px;">
                <button class="btn btn-primary btn-sm" onclick="toggleShow('add-orcamento')">+ Adicionar</button>
            </div>
            <div id="add-orcamento" style="display:none;" class="section">
                <form method="POST" action="{{ P }}/{{ event['id'] }}/orcamento/adicionar">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Categoria</label>
                            <select name="category">
                                {% for c in BUDGET_CATEGORIES %}
                                <option value="{{ c }}">{{ c }}</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Fonte</label>
                            <select name="funding_source">
                                <option value="associacao">🏠 Associação</option>
                                <option value="rouanet">🏛️ Rouanet</option>
                            </select>
                        </div>
                        <div class="form-group form-full">
                            <label>Item *</label>
                            <input type="text" name="item" required>
                        </div>
                        <div class="form-group">
                            <label>Qtd</label>
                            <input type="number" name="quantity" step="0.01" min="0">
                        </div>
                        <div class="form-group">
                            <label>Custo unitário (R$)</label>
                            <input type="number" name="unit_cost" step="0.01" min="0">
                        </div>
                        <div class="form-group">
                            <label>Valor estimado (R$)</label>
                            <input type="number" name="estimated_amount" step="0.01" min="0">
                        </div>
                        <div class="form-group">
                            <label>Valor real (R$)</label>
                            <input type="number" name="actual_amount" step="0.01" min="0">
                        </div>
                        <div class="form-group">
                            <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; padding-top: 24px;">
                                <input type="checkbox" name="is_income" style="width: auto;">
                                É receita
                            </label>
                        </div>
                        <div class="form-group">
                            <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; padding-top: 24px;">
                                <input type="checkbox" name="paid" style="width: auto;">
                                Pago
                            </label>
                        </div>
                        <div class="form-group">
                            <label>Data pagamento</label>
                            <input type="date" name="paid_date">
                        </div>
                        <div class="form-group form-full">
                            <label>Notas</label>
                            <input type="text" name="notes">
                        </div>
                    </div>
                    <div class="form-actions">
                        <button type="submit" class="btn btn-primary">Salvar</button>
                        <button type="button" class="btn btn-secondary" onclick="toggleShow('add-orcamento')">Cancelar</button>
                    </div>
                </form>
            </div>

            {% if orcamento %}
            <div class="table-wrap">
            <table>
                <thead>
                    <tr><th>Cat.</th><th>Item</th><th>Fonte</th><th>Estimado</th><th>Real</th><th>Pago</th><th></th></tr>
                </thead>
                <tbody>
                {% for item in orcamento %}
                <tr {% if item['is_income'] %}style="background: rgba(98,204,60,0.04);"{% endif %}>
                    <td>{{ item['category'] or '—' }}</td>
                    <td>
                        {% if item['is_income'] %}<span style="color: var(--green);">↑</span>{% else %}<span style="color: var(--text-muted);">↓</span>{% endif %}
                        {{ item['item'] }}
                    </td>
                    <td>
                        {% if item['funding_source'] == 'rouanet' %}
                        <span class="badge badge-rouanet">🏛️ Rouanet</span>
                        {% else %}
                        <span class="badge badge-associacao">🏠 Assoc.</span>
                        {% endif %}
                    </td>
                    <td>{{ format_brl(item['estimated_amount']) }}</td>
                    <td>{{ format_brl(item['actual_amount']) }}</td>
                    <td>
                        {% if item['paid'] %}
                        <span class="badge badge-confirmed">✓ Pago</span>
                        {% else %}
                        <span class="badge badge-pendente">Pendente</span>
                        {% endif %}
                    </td>
                    <td style="white-space: nowrap;">
                        <button class="btn btn-secondary btn-sm" onclick="toggleEdit('orcamento-{{ item['id'] }}')">✏️</button>
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/orcamento/{{ item['id'] }}/excluir" style="display:inline;">
                            <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Excluir?')">🗑️</button>
                        </form>
                    </td>
                </tr>
                <tr id="edit-orcamento-{{ item['id'] }}" style="display:none;">
                    <td colspan="99">
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/orcamento/{{ item['id'] }}/editar" style="padding: 12px 0;">
                            <div class="form-grid">
                                <div class="form-group">
                                    <label>Categoria</label>
                                    <select name="category">
                                        {% for c in BUDGET_CATEGORIES %}
                                        <option value="{{ c }}" {% if c == item['category'] %}selected{% endif %}>{{ c }}</option>
                                        {% endfor %}
                                    </select>
                                </div>
                                <div class="form-group">
                                    <label>Fonte</label>
                                    <select name="funding_source">
                                        <option value="associacao" {% if item['funding_source'] != 'rouanet' %}selected{% endif %}>🏠 Associação</option>
                                        <option value="rouanet" {% if item['funding_source'] == 'rouanet' %}selected{% endif %}>🏛️ Rouanet</option>
                                    </select>
                                </div>
                                <div class="form-group form-full">
                                    <label>Item</label>
                                    <input type="text" name="item" value="{{ item['item'] or '' }}">
                                </div>
                                <div class="form-group">
                                    <label>Qtd</label>
                                    <input type="number" name="quantity" step="0.01" value="{{ item['quantity'] or '' }}">
                                </div>
                                <div class="form-group">
                                    <label>Custo unitário (R$)</label>
                                    <input type="number" name="unit_cost" step="0.01" value="{{ item['unit_cost'] or '' }}">
                                </div>
                                <div class="form-group">
                                    <label>Estimado (R$)</label>
                                    <input type="number" name="estimated_amount" step="0.01" value="{{ item['estimated_amount'] or '' }}">
                                </div>
                                <div class="form-group">
                                    <label>Real (R$)</label>
                                    <input type="number" name="actual_amount" step="0.01" value="{{ item['actual_amount'] or '' }}">
                                </div>
                                <div class="form-group">
                                    <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; padding-top: 24px;">
                                        <input type="checkbox" name="is_income" style="width: auto;" {% if item['is_income'] %}checked{% endif %}>
                                        É receita
                                    </label>
                                </div>
                                <div class="form-group">
                                    <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; padding-top: 24px;">
                                        <input type="checkbox" name="paid" style="width: auto;" {% if item['paid'] %}checked{% endif %}>
                                        Pago
                                    </label>
                                </div>
                                <div class="form-group">
                                    <label>Data pagamento</label>
                                    <input type="date" name="paid_date" value="{{ item['paid_date'] or '' }}">
                                </div>
                                <div class="form-group form-full">
                                    <label>Notas</label>
                                    <input type="text" name="notes" value="{{ item['notes'] or '' }}">
                                </div>
                            </div>
                            <div class="form-actions">
                                <button type="submit" class="btn btn-primary btn-sm">Salvar</button>
                                <button type="button" class="btn btn-secondary btn-sm" onclick="toggleEdit('orcamento-{{ item['id'] }}')">Cancelar</button>
                            </div>
                        </form>
                    </td>
                </tr>
                {% endfor %}
                </tbody>
                <tfoot>
                    <tr style="border-top: 2px solid var(--border);">
                        <td colspan="3" style="font-weight: 600; color: var(--text-muted); font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em;">Totais</td>
                        <td colspan="4"></td>
                    </tr>
                    <tr>
                        <td colspan="3" style="color: var(--green);">Receitas</td>
                        <td colspan="4" style="color: var(--green); font-weight: 600;">{{ format_brl(receitas) }}</td>
                    </tr>
                    <tr>
                        <td colspan="3" style="color: var(--text-muted);">Despesas (Associação)</td>
                        <td colspan="4">{{ format_brl(despesas_assoc) }}</td>
                    </tr>
                    <tr>
                        <td colspan="3" style="color: var(--text-muted);">Despesas (Rouanet)</td>
                        <td colspan="4"><span class="badge badge-rouanet">{{ format_brl(despesas_rouanet) }}</span></td>
                    </tr>
                    <tr>
                        <td colspan="3" style="font-weight: 700;">Saldo</td>
                        <td colspan="4" style="font-weight: 700; color: {% if saldo >= 0 %}var(--green){% else %}var(--red){% endif %};">{{ format_brl(saldo) }}</td>
                    </tr>
                </tfoot>
            </table>
            </div>
            {% else %}
            <div class="empty"><div class="empty-icon">💰</div><p>Sem linhas de orçamento</p></div>
            {% endif %}
        </div>

        <!-- CONTATOS -->
        <div class="tab-content panel-contatos">
            <div style="text-align:right; margin-bottom: 12px;">
                <button class="btn btn-primary btn-sm" onclick="toggleShow('add-contatos')">+ Adicionar</button>
            </div>
            <div id="add-contatos" style="display:none;" class="section">
                <form method="POST" action="{{ P }}/{{ event['id'] }}/contatos/adicionar">
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Categoria</label>
                            <select name="category">
                                {% for c in CONTACT_CATEGORIES %}
                                <option value="{{ c }}">{{ c.capitalize() }}</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Nome *</label>
                            <input type="text" name="name" required>
                        </div>
                        <div class="form-group">
                            <label>Função</label>
                            <input type="text" name="role">
                        </div>
                        <div class="form-group">
                            <label>Telefone</label>
                            <input type="tel" name="phone">
                        </div>
                        <div class="form-group">
                            <label>Email</label>
                            <input type="email" name="email">
                        </div>
                        <div class="form-group form-full">
                            <label>Notas</label>
                            <input type="text" name="notes">
                        </div>
                    </div>
                    <div class="form-actions">
                        <button type="submit" class="btn btn-primary">Salvar</button>
                        <button type="button" class="btn btn-secondary" onclick="toggleShow('add-contatos')">Cancelar</button>
                    </div>
                </form>
            </div>

            {% if contatos %}
            {% set current_cat = namespace(name='') %}
            {% for item in contatos %}
            {% if item['category'] != current_cat.name %}
            {% if not loop.first %}</tbody></table></div>{% endif %}
            {% set current_cat.name = item['category'] %}
            <div style="margin-bottom: 8px; margin-top: {{ '20px' if not loop.first else '0' }};">
                <strong style="color: var(--yellow);">{{ item['category'].capitalize() }}</strong>
            </div>
            <div class="table-wrap"><table>
                <thead><tr><th>Nome</th><th>Função</th><th>Telefone</th><th>Email</th><th>Notas</th><th></th></tr></thead>
                <tbody>
            {% endif %}
            <tr>
                <td>{{ item['name'] }}</td>
                <td>{{ item['role'] or '—' }}</td>
                <td>{% if item['phone'] %}<a href="tel:{{ item['phone'] }}">{{ item['phone'] }}</a>{% else %}—{% endif %}</td>
                <td>{% if item['email'] %}<a href="mailto:{{ item['email'] }}">{{ item['email'] }}</a>{% else %}—{% endif %}</td>
                <td>{{ item['notes'] or '—' }}</td>
                <td style="white-space: nowrap;">
                    <button class="btn btn-secondary btn-sm" onclick="toggleEdit('contato-{{ item['id'] }}')">✏️</button>
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/contatos/{{ item['id'] }}/excluir" style="display:inline;">
                        <button type="submit" class="btn btn-danger btn-sm" onclick="return confirm('Excluir?')">🗑️</button>
                    </form>
                </td>
            </tr>
            <tr id="edit-contato-{{ item['id'] }}" style="display:none;">
                <td colspan="99">
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/contatos/{{ item['id'] }}/editar" style="padding: 12px 0;">
                        <div class="form-grid">
                            <div class="form-group">
                                <label>Categoria</label>
                                <select name="category">
                                    {% for c in CONTACT_CATEGORIES %}
                                    <option value="{{ c }}" {% if c == item['category'] %}selected{% endif %}>{{ c.capitalize() }}</option>
                                    {% endfor %}
                                </select>
                            </div>
                            <div class="form-group">
                                <label>Nome</label>
                                <input type="text" name="name" value="{{ item['name'] or '' }}">
                            </div>
                            <div class="form-group">
                                <label>Função</label>
                                <input type="text" name="role" value="{{ item['role'] or '' }}">
                            </div>
                            <div class="form-group">
                                <label>Telefone</label>
                                <input type="tel" name="phone" value="{{ item['phone'] or '' }}">
                            </div>
                            <div class="form-group">
                                <label>Email</label>
                                <input type="email" name="email" value="{{ item['email'] or '' }}">
                            </div>
                            <div class="form-group form-full">
                                <label>Notas</label>
                                <input type="text" name="notes" value="{{ item['notes'] or '' }}">
                            </div>
                        </div>
                        <div class="form-actions">
                            <button type="submit" class="btn btn-primary btn-sm">Salvar</button>
                            <button type="button" class="btn btn-secondary btn-sm" onclick="toggleEdit('contato-{{ item['id'] }}')">Cancelar</button>
                        </div>
                    </form>
                </td>
            </tr>
            {% endfor %}
            {% if contatos %}</tbody></table></div>{% endif %}
            {% else %}
            <div class="empty"><div class="empty-icon">📞</div><p>Sem contatos</p></div>
            {% endif %}
        </div>

        <!-- CHECKLIST -->
        <div class="tab-content panel-tarefas">

            {% if not tarefas and event['type'] in CHECKLIST_TEMPLATES %}
            <div style="text-align:center; padding: 24px 0;">
                <p style="color: var(--text-muted); margin-bottom: 12px;">Nenhum item no checklist.</p>
                <form method="POST" action="{{ P }}/{{ event['id'] }}/tarefas/template" style="display:inline;">
                    <button type="submit" class="btn btn-primary">Carregar template de {{ TYPE_LABELS.get(event['type'], event['type'] or 'evento') }}</button>
                </form>
            </div>
            {% elif not tarefas %}
            <div class="empty"><div class="empty-icon">✅</div><p>Nenhum item no checklist.</p></div>
            {% endif %}

            {% for phase_key, phase_label in CHECKLIST_PHASES %}
            {% set phase_items = tarefas|selectattr('category', 'equalto', phase_key)|list %}
            {% set phase_done = phase_items|selectattr('status', 'equalto', 'concluido')|list|length %}
            <details class="checklist-phase" {% if phase_items and phase_done < phase_items|length %}open{% endif %}>
                <summary style="cursor:pointer; font-weight:600; font-size:16px; padding:10px 0; border-bottom:1px solid var(--border); margin-bottom:8px; color: {% if phase_items and phase_done == phase_items|length %}var(--text-muted){% else %}var(--text){% endif %};">
                    {{ phase_label }}
                    <span style="font-weight:400; font-size:13px; color:var(--text-muted);">({{ phase_done }}/{{ phase_items|length }})</span>
                </summary>

                {% for item in phase_items %}
                <div class="checklist-item" style="display:flex; align-items:center; gap:8px; padding:6px 0; {% if item['status'] == 'concluido' %}opacity:0.5;{% endif %}">
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/tarefas/{{ item['id'] }}/concluir" style="display:inline; flex-shrink:0;">
                        <button type="submit" class="btn btn-secondary btn-sm" style="padding:2px 6px;" title="{% if item['status'] == 'concluido' %}Reabrir{% else %}Concluir{% endif %}">
                            {% if item['status'] == 'concluido' %}✅{% else %}⬜{% endif %}
                        </button>
                    </form>
                    <span class="{% if item['status'] == 'concluido' %}task-done{% endif %}" style="flex:1;">{{ item['task'] }}</span>
                    {% if item['responsible'] %}
                    <span style="font-size:12px; color:var(--text-muted); white-space:nowrap;">{{ item['responsible'] }}</span>
                    {% endif %}
                    <span style="white-space:nowrap; flex-shrink:0;">
                        <button class="btn btn-secondary btn-sm" style="padding:2px 6px;" onclick="toggleEdit('tarefa-{{ item['id'] }}')">✏️</button>
                        <form method="POST" action="{{ P }}/{{ event['id'] }}/tarefas/{{ item['id'] }}/excluir" style="display:inline;">
                            <button type="submit" class="btn btn-danger btn-sm" style="padding:2px 6px;" onclick="return confirm('Excluir?')">🗑️</button>
                        </form>
                    </span>
                </div>
                <div id="edit-tarefa-{{ item['id'] }}" style="display:none; padding:8px 0 8px 32px;">
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/tarefas/{{ item['id'] }}/editar" style="display:flex; gap:8px; align-items:end; flex-wrap:wrap;">
                        <input type="hidden" name="category" value="{{ phase_key }}">
                        <div style="flex:2; min-width:200px;">
                            <label style="font-size:12px;">Item</label>
                            <input type="text" name="task" value="{{ item['task'] or '' }}" style="width:100%;">
                        </div>
                        <div style="flex:1; min-width:120px;">
                            <label style="font-size:12px;">Responsável</label>
                            <input type="text" name="responsible" value="{{ item['responsible'] or '' }}" style="width:100%;">
                        </div>
                        <button type="submit" class="btn btn-primary btn-sm">Salvar</button>
                        <button type="button" class="btn btn-secondary btn-sm" onclick="toggleEdit('tarefa-{{ item['id'] }}')">Cancelar</button>
                    </form>
                </div>
                {% endfor %}

                <div id="add-checklist-{{ phase_key }}" style="display:none; padding:8px 0 8px 32px;">
                    <form method="POST" action="{{ P }}/{{ event['id'] }}/tarefas/adicionar" style="display:flex; gap:8px; align-items:end; flex-wrap:wrap;">
                        <input type="hidden" name="category" value="{{ phase_key }}">
                        <div style="flex:2; min-width:200px;">
                            <label style="font-size:12px;">Item</label>
                            <input type="text" name="task" required style="width:100%;">
                        </div>
                        <div style="flex:1; min-width:120px;">
                            <label style="font-size:12px;">Responsável</label>
                            <input type="text" name="responsible" style="width:100%;">
                        </div>
                        <button type="submit" class="btn btn-primary btn-sm">Adicionar</button>
                        <button type="button" class="btn btn-secondary btn-sm" onclick="toggleShow('add-checklist-{{ phase_key }}')">Cancelar</button>
                    </form>
                </div>
                <div style="padding:4px 0 12px 0;">
                    <button class="btn btn-secondary btn-sm" onclick="toggleShow('add-checklist-{{ phase_key }}')" style="font-size:12px;">+ Adicionar item</button>
                </div>
            </details>
            {% endfor %}

        </div>

        </div><!-- tab-panels -->
    </div><!-- tabs -->
</div><!-- container -->

<script>var WHATSAPP_TEXT = {{ whatsapp_text|tojson }};</script>
<script>
function copyWhatsApp() {
    var btn = document.getElementById('btn-whatsapp');
    navigator.clipboard.writeText(WHATSAPP_TEXT).then(function() {
        btn.textContent = '✅ Copiado!';
        setTimeout(function() { btn.textContent = '📋 WhatsApp'; }, 2000);
    });
}

function toggleShow(id) {
    var el = document.getElementById(id);
    el.style.display = el.style.display === 'none' ? 'block' : 'none';
}

function toggleEdit(id) {
    var el = document.getElementById('edit-' + id);
    if (!el) return;
    el.style.display = el.style.display === 'none' ? 'table-row' : 'none';
}

function fillTeacherName(sel) {
    var opt = sel.options[sel.selectedIndex];
    var nameField = document.getElementById('teacher-name-field');
    if (nameField && opt && opt.dataset.name) {
        nameField.value = opt.dataset.name;
    }
}

// Tab from URL hash
var hash = window.location.hash.replace('#', '');
if (hash === 'checklist') hash = 'tarefas';  // alias
if (hash) {
    var tab = document.getElementById('tab-' + hash);
    if (tab) tab.checked = true;
}
// Update hash on tab click
document.querySelectorAll('.tab-label').forEach(function(label) {
    label.addEventListener('click', function() {
        var h = this.getAttribute('for').replace('tab-', '');
        if (h === 'tarefas') h = 'checklist';
        window.location.hash = h;
    });
});

// ── Cronograma (agile timeline) ──────────────────────────
(function() {
    if (typeof CRONO_DATA === 'undefined') return;
    var items = CRONO_DATA;
    var eventId = CRONO_EVENT_ID;
    var P = CRONO_PREFIX;
    var listEl = document.getElementById('crono-list');
    var inputEl = document.getElementById('crono-input');
    if (!listEl || !inputEl) return;

    function esc(s) { var d = document.createElement('div'); d.textContent = s || ''; return d.innerHTML; }

    function dayLabel(dateStr) {
        if (!dateStr) return 'Sem data';
        var parts = dateStr.split('-');
        if (parts.length !== 3) return dateStr;
        var d = new Date(parseInt(parts[0]), parseInt(parts[1])-1, parseInt(parts[2]));
        var dows = ['Dom','Seg','Ter','Qua','Qui','Sex','Sáb'];
        return dows[d.getDay()] + ' ' + parts[2] + '/' + parts[1];
    }

    function shortDate(dateStr) {
        if (!dateStr) return 'sem data';
        var parts = dateStr.split('-');
        if (parts.length !== 3) return dateStr;
        return parts[2] + '/' + parts[1];
    }

    function parseCronoInput(text) {
        text = text.trim();
        if (!text) return null;
        var result = {time_start: null, time_end: null, activity: '', location: null, responsible: null, date: null, notes: null};
        // Try to extract time from beginning: HH:MM or HH:MM-HH:MM
        var timeMatch = text.match(/^(\\d{1,2}:\\d{2})(?:\\s*[-–]\\s*(\\d{1,2}:\\d{2}))?\\s+(.+)/);
        if (timeMatch) {
            result.time_start = timeMatch[1].length === 4 ? '0' + timeMatch[1] : timeMatch[1];
            if (timeMatch[2]) {
                result.time_end = timeMatch[2].length === 4 ? '0' + timeMatch[2] : timeMatch[2];
            }
            text = timeMatch[3];
        }
        // Split by pipe: activity | location | responsible
        var parts = text.split('|').map(function(s) { return s.trim(); });
        result.activity = parts[0] || '';
        if (parts[1]) result.location = parts[1];
        if (parts[2]) result.responsible = parts[2];
        return result.activity ? result : null;
    }

    function renderItem(item) {
        var timeHtml = '';
        if (item.time_start) {
            timeHtml = esc(item.time_start);
            if (item.time_end) timeHtml += '<span class="crono-end">- ' + esc(item.time_end) + '</span>';
        } else {
            timeHtml = '<span style="color:var(--text-dim)">—</span>';
        }
        var dateClass = item.date ? 'crono-date-pill' : 'crono-date-pill crono-date-missing';
        var dateLabel = item.date ? esc(shortDate(item.date)) : 'sem data';
        var dateHtml = '<span class="' + dateClass + '" data-field="date" data-id="' + item.id + '" title="Editar data">' + dateLabel + '</span>';
        var metaHtml = dateHtml;
        if (item.location) metaHtml += '<span class="crono-location" data-field="location" data-id="' + item.id + '">' + esc(item.location) + '</span>';
        if (item.responsible) metaHtml += '<span class="crono-responsible" data-field="responsible" data-id="' + item.id + '">' + esc(item.responsible) + '</span>';

        return '<div class="crono-item" data-id="' + item.id + '">' +
            '<div class="crono-time" data-field="time" data-id="' + item.id + '">' + timeHtml + '</div>' +
            '<div class="crono-body">' +
                '<span class="crono-activity" data-field="activity" data-id="' + item.id + '">' + esc(item.activity) + '</span>' +
                metaHtml +
            '</div>' +
            '<div class="crono-actions">' +
                '<button title="Mover para cima" onclick="cronoMove(' + item.id + ',\\'up\\')">&uarr;</button>' +
                '<button title="Mover para baixo" onclick="cronoMove(' + item.id + ',\\'down\\')">&darr;</button>' +
                '<button class="crono-del" title="Excluir" onclick="cronoDelete(' + item.id + ')">&#x1f5d1;</button>' +
            '</div>' +
        '</div>';
    }

    function renderList() {
        if (!items.length) {
            listEl.innerHTML = '<div class="crono-empty">Cronograma vazio. Digite acima para adicionar o primeiro item.</div>';
            return;
        }
        var groups = {};
        items.forEach(function(it) {
            var key = it.date || '';
            if (!groups[key]) groups[key] = [];
            groups[key].push(it);
        });
        var keys = Object.keys(groups).sort(function(a, b) {
            if (!a) return 1;
            if (!b) return -1;
            return a < b ? -1 : a > b ? 1 : 0;
        });
        var html = '';
        keys.forEach(function(k) {
            var headerCls = k ? 'crono-day-header' : 'crono-day-header crono-day-undated';
            html += '<div class="' + headerCls + '">' + esc(dayLabel(k)) + '</div>';
            html += groups[k].map(renderItem).join('');
        });
        listEl.innerHTML = html;
        attachEditHandlers();
    }

    function flashSaved(el) {
        var item = el.closest('.crono-item');
        if (!item) return;
        item.classList.add('crono-saved');
        setTimeout(function() { item.classList.remove('crono-saved'); }, 800);
    }

    function apiCall(method, path, body) {
        var opts = {method: method, headers: {'Content-Type': 'application/json'}};
        if (body) opts.body = JSON.stringify(body);
        return fetch(P + '/' + eventId + '/api/cronograma' + (path ? '/' + path : ''), opts)
            .then(function(r) {
                var ct = r.headers.get('content-type') || '';
                if (!ct.includes('application/json')) {
                    throw new Error('Sessão expirada. Recarregue a página.');
                }
                if (!r.ok) throw new Error('Erro ' + r.status);
                return r.json();
            });
    }

    function showError(msg) {
        // Remove any existing error
        var old = listEl.parentNode.querySelector('.crono-error');
        if (old) old.remove();
        var el = document.createElement('div');
        el.className = 'crono-error';
        el.textContent = msg || 'Erro ao salvar. Tente novamente.';
        listEl.parentNode.insertBefore(el, listEl);
        setTimeout(function() { if (el.parentNode) el.remove(); }, 4000);
    }

    // Quick-add
    inputEl.addEventListener('keydown', function(e) {
        if (e.key !== 'Enter') return;
        var parsed = parseCronoInput(inputEl.value);
        if (!parsed) return;
        inputEl.disabled = true;
        apiCall('POST', '', parsed).then(function(newItem) {
            items.push(newItem);
            renderList();
            inputEl.value = '';
            inputEl.disabled = false;
            inputEl.focus();
        }).catch(function(err) { inputEl.disabled = false; showError(err.message); });
    });

    // Inline edit
    function attachEditHandlers() {
        listEl.querySelectorAll('[data-field]').forEach(function(el) {
            el.addEventListener('click', function(e) {
                if (el.querySelector('input')) return; // already editing
                startEdit(el);
            });
        });
    }

    function startEdit(el) {
        var field = el.dataset.field;
        var itemId = parseInt(el.dataset.id);
        var item = items.find(function(i) { return i.id === itemId; });
        if (!item) return;

        if (field === 'date') {
            var dateVal = item.date || '';
            el.innerHTML = '<input type="date" class="crono-edit-input" value="' + esc(dateVal) + '" style="width:140px;">';
            var dInp = el.querySelector('input');
            dInp.focus();
            dInp.addEventListener('blur', function() {
                saveField(itemId, {date: dInp.value || null}, el);
            });
            dInp.addEventListener('keydown', function(ev) {
                if (ev.key === 'Enter') { dInp.blur(); }
                if (ev.key === 'Escape') { renderList(); }
            });
        } else if (field === 'time') {
            // Show two time inputs side by side
            var startVal = item.time_start || '';
            var endVal = item.time_end || '';
            el.innerHTML = '<input type="time" class="crono-edit-input" value="' + esc(startVal) + '" style="width:80px;" data-subfield="time_start">' +
                           '<input type="time" class="crono-edit-input" value="' + esc(endVal) + '" style="width:80px; margin-left:4px;" data-subfield="time_end">';
            var inputs = el.querySelectorAll('input');
            inputs[0].focus();
            inputs.forEach(function(inp) {
                inp.addEventListener('blur', function() {
                    setTimeout(function() {
                        // Check if another input in same cell has focus
                        if (el.contains(document.activeElement)) return;
                        var data = {};
                        data.time_start = inputs[0].value || null;
                        data.time_end = inputs[1].value || null;
                        saveField(itemId, data, el);
                    }, 100);
                });
                inp.addEventListener('keydown', function(ev) {
                    if (ev.key === 'Enter') { inp.blur(); }
                    if (ev.key === 'Escape') { renderList(); }
                });
            });
        } else {
            var val = item[field] || '';
            var inp = document.createElement('input');
            inp.type = 'text';
            inp.className = 'crono-edit-input';
            inp.value = val;
            inp.style.width = Math.max(100, el.offsetWidth) + 'px';
            el.textContent = '';
            el.appendChild(inp);
            inp.focus();
            inp.select();
            inp.addEventListener('blur', function() {
                var data = {};
                data[field] = inp.value || null;
                saveField(itemId, data, el);
            });
            inp.addEventListener('keydown', function(ev) {
                if (ev.key === 'Enter') { inp.blur(); }
                if (ev.key === 'Escape') { renderList(); }
                if (ev.key === 'Tab') {
                    ev.preventDefault();
                    inp.blur();
                    // Move to next editable field
                    var itemEl = el.closest('.crono-item');
                    if (!itemEl) return;
                    var fields = itemEl.querySelectorAll('[data-field]');
                    var idx = Array.prototype.indexOf.call(fields, el);
                    var next = fields[idx + 1];
                    if (next) setTimeout(function() { startEdit(next); }, 50);
                }
            });
        }
    }

    function saveField(itemId, data, el) {
        apiCall('PUT', itemId, data).then(function(updated) {
            var idx = items.findIndex(function(i) { return i.id === itemId; });
            if (idx >= 0) items[idx] = updated;
            renderList();
            // Flash the saved item
            var savedEl = listEl.querySelector('[data-id="' + itemId + '"].crono-item');
            if (savedEl) { savedEl.classList.add('crono-saved'); setTimeout(function() { savedEl.classList.remove('crono-saved'); }, 800); }
        }).catch(function(err) { renderList(); showError(err.message); });
    }

    // Reorder
    window.cronoMove = function(itemId, direction) {
        apiCall('POST', 'reorder', {item_id: itemId, direction: direction}).then(function(newList) {
            items = newList;
            renderList();
        }).catch(function(err) { showError(err.message); });
    };

    // Delete
    window.cronoDelete = function(itemId) {
        if (!confirm('Excluir este item?')) return;
        apiCall('DELETE', itemId).then(function() {
            items = items.filter(function(i) { return i.id !== itemId; });
            renderList();
        }).catch(function(err) { showError(err.message); });
    };

    // Initial render
    renderList();
})();
</script>
"""

ORCAMENTO_OVERVIEW_BODY = """
<div class="container">
    <div class="page-header">
        <h1>Orçamento Geral</h1>
        <span style="color: var(--text-muted); font-size: 13px;">Eventos não cancelados — ano corrente</span>
    </div>
    <div class="section">
        {% if rows %}
        <div class="table-wrap">
        <table>
            <thead>
                <tr><th>Data</th><th>Evento</th><th>Tipo</th><th>Status</th><th>Receitas</th><th>Despesas</th><th>Saldo</th></tr>
            </thead>
            <tbody>
            {% for r in rows %}
            {% set saldo = r['total_receitas'] - r['total_despesas'] %}
            <tr>
                <td>{{ format_date(r['date']) }}</td>
                <td><a href="{{ P }}/{{ r['id'] }}#orcamento">{{ r['name'] }}</a></td>
                <td>{{ TYPE_LABELS.get(r['type'], r['type'] or '—') }}</td>
                <td><span class="badge badge-{{ r['status'] }}">{{ STATUS_LABELS.get(r['status'], r['status']) }}</span></td>
                <td style="color: var(--green);">{{ format_brl(r['total_receitas']) if r['total_receitas'] else '—' }}</td>
                <td>{{ format_brl(r['total_despesas']) if r['total_despesas'] else '—' }}</td>
                <td style="color: {% if saldo >= 0 %}var(--green){% else %}var(--red){% endif %}; font-weight: 600;">{{ format_brl(saldo) }}</td>
            </tr>
            {% endfor %}
            </tbody>
        </table>
        </div>
        {% else %}
        <div class="empty"><div class="empty-icon">💰</div><p>Sem dados de orçamento</p></div>
        {% endif %}
    </div>
</div>
"""


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5007, debug=True)
